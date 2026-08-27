from datetime import date, datetime, timedelta
from typing import Any

from sqlalchemy.ext.asyncio import AsyncSession

from app.core.constants import AppointmentStatus, BillingStatus
from app.core.exceptions import BadRequestException, NotFoundException, ConflictException
from app.models.billing_model import BillItem, Billing, Insurance, InsuranceClaim, Payment
from app.repositories.audit_repository import AuditRepository
from app.repositories.billing_repository import BillingRepository, InsuranceClaimRepository, InsuranceRepository
from app.repositories.patient_repository import PatientRepository
from app.schemas.billing_schema import (
    BillType,
    BillingCreate,
    BillingResponse,
    BillingSummary,
    BillingUpdate,
    BillItemResponse,
    DailyCollectionSummary,
    InsuranceClaimCreate,
    InsuranceClaimResponse,
    InsuranceCreate,
    InsuranceResponse,
    PaymentCreate,
    PaymentResponse,
    RefundCreate,
    RevenueReport,
)
from app.utils.helpers import (
    calculate_bill_totals,
    calculate_line_total,
    generate_bill_number,
    generate_claim_number,
    utc_now,
)
from app.utils.pagination import build_paginated_result
from app.utils.pdf_generator import generate_invoice_pdf


class BillingService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.repo = BillingRepository(db)
        self.patient_repo = PatientRepository(db)
        self.audit_repo = AuditRepository(db)

    def _to_response(self, billing: Billing) -> BillingResponse:
        data = BillingResponse.model_validate(billing)
        data.items = [BillItemResponse.model_validate(i) for i in billing.items]
        data.source = "billing"
        if billing.bill_number and str(billing.bill_number).upper().startswith("REC"):
            data.bill_type = "pharmacy"
        else:
            data.bill_type = "consultation"
        return data

    def _pharmacy_invoice_to_billing_response(self, invoice) -> BillingResponse:
        bill_items = []
        for item in (invoice.items or []):
            medicine_name = item.medicine.name if (hasattr(item, "medicine") and item.medicine) else f"Medicine #{item.medicine_id}"
            bill_items.append(
                BillItemResponse(
                    id=item.id,
                    billing_id=invoice.id,
                    description=f"{medicine_name} (Qty: {item.quantity})",
                    quantity=item.quantity,
                    unit_price=item.unit_price,
                    gst_rate=invoice.tax_percentage,
                    gst_amount=invoice.gst_amount / len(invoice.items) if invoice.items else 0.0,
                    line_total=item.line_total,
                    item_type="pharmacy_item",
                    created_at=item.created_at or invoice.created_at or utc_now(),
                )
            )
        
        return BillingResponse(
            id=invoice.id,
            patient_id=invoice.patient_id or 0,
            bill_number=invoice.invoice_number,
            subtotal=invoice.subtotal,
            discount_percent=invoice.discount_percentage,
            discount_amount=invoice.discount_amount,
            gst_rate=invoice.tax_percentage,
            gst_amount=invoice.gst_amount,
            tax_amount=invoice.tax_amount,
            total_amount=invoice.total_amount,
            paid_amount=invoice.paid_amount,
            balance_amount=round(max(invoice.total_amount - invoice.paid_amount, 0.0), 2),
            status=invoice.status,
            due_date=invoice.created_at,
            notes=f"Pharmacy Invoice (Prescription ID: {invoice.prescription_id})",
            invoice_path=None,
            appointment_id=None,
            items=bill_items,
            created_at=invoice.created_at or utc_now(),
            updated_at=invoice.updated_at or utc_now(),
            source="pharmacy",
            bill_type="pharmacy",
        )

    async def _recalculate_billing(self, billing: Billing) -> Billing:
        from sqlalchemy import inspect
        state = inspect(billing)
        if "items" in state.unloaded and billing.id is not None:
            billing = await self.repo.get_by_id(billing.id)

        subtotal = sum(
            (item.quantity * item.unit_price) for item in billing.items
        ) if billing.items else billing.subtotal
        totals = calculate_bill_totals(
            subtotal=subtotal,
            discount_percent=billing.discount_percent,
            discount_amount=billing.discount_amount,
            gst_rate=billing.gst_rate if billing.gst_rate is not None else 18.0,
            tax_amount=billing.tax_amount,
        )
        billing.subtotal = totals["subtotal"]
        billing.discount_amount = totals["discount_amount"]
        billing.tax_amount = totals["tax_amount"]

        if billing.items:
            total_gst = 0.0
            for item in billing.items:
                # Recalculate each item's GST and line total dynamically
                _, gst_amt, line_total = calculate_line_total(
                    item.quantity, item.unit_price, item.gst_rate
                )
                item.gst_amount = gst_amt
                item.line_total = line_total
                total_gst += gst_amt
            billing.gst_amount = round(total_gst, 2)
            
            # Recalculate the effective gst_rate for the bill
            if len(billing.items) == 1:
                billing.gst_rate = billing.items[0].gst_rate
            elif billing.subtotal > 0:
                billing.gst_rate = round((billing.gst_amount / billing.subtotal) * 100, 2)
            else:
                billing.gst_rate = 0.0
        else:
            billing.gst_amount = totals["gst_amount"]

        due_date_naive = billing.due_date
        if due_date_naive and due_date_naive.tzinfo is not None:
            from datetime import timezone as py_timezone
            due_date_naive = due_date_naive.astimezone(py_timezone.utc).replace(tzinfo=None)

        billing.total_amount = round(max(billing.subtotal - billing.discount_amount, 0.0) + billing.gst_amount + billing.tax_amount, 2)
        billing.balance_amount = round(billing.total_amount - billing.paid_amount, 2)
        
        # Check if there are any completed refund payments
        has_refund = any(p.is_refund and p.status == "completed" for p in billing.payments) if billing.payments else False

        if billing.balance_amount <= 0:
            billing.status = BillingStatus.PAID
        elif has_refund and billing.paid_amount <= 0:
            billing.status = BillingStatus.REFUNDED
        elif billing.paid_amount > 0:
            billing.status = BillingStatus.PARTIAL
        elif due_date_naive and due_date_naive < utc_now():
            billing.status = BillingStatus.OVERDUE
        else:
            billing.status = BillingStatus.PENDING
        return await self.repo.update(billing)

    async def create(self, data: BillingCreate, user_id: int) -> BillingResponse:
        patient = await self.patient_repo.get_by_id(data.patient_id)
        if not patient:
            raise NotFoundException("Patient not found")

        if data.appointment_id is not None:
            from sqlalchemy import select
            from app.models.appointment_model import Appointment
            stmt = select(Appointment).where(Appointment.id == data.appointment_id)
            apt_result = await self.db.execute(stmt)
            appointment = apt_result.scalar_one_or_none()
            if not appointment:
                raise NotFoundException(f"Appointment with ID {data.appointment_id} not found")
            
            if appointment.patient_id != data.patient_id:
                raise BadRequestException("Appointment does not belong to the supplied patient")

            if appointment.appointment_status == AppointmentStatus.CANCELLED:
                raise BadRequestException(
                    "Billing cannot be created for a cancelled appointment."
                )

            is_completed = (
                appointment.appointment_status == AppointmentStatus.COMPLETED
                or appointment.appointment_status == "Checked-Out"
                or appointment.queue_status == "COMPLETED"
                or appointment.check_out_time is not None
            )
            if not is_completed:
                raise BadRequestException(
                    "Billing can only be created for completed appointments."
                )

            existing_billing = await self.repo.get_by_patient_and_appointment(data.patient_id, data.appointment_id)
            if existing_billing:
                raise ConflictException("Billing already exists for this patient and appointment.")

        due_date = data.due_date
        if due_date and due_date.tzinfo is not None:
            from datetime import timezone as py_timezone
            due_date = due_date.astimezone(py_timezone.utc).replace(tzinfo=None)

        from app.models.user_model import User
        from app.services.settings_service import SettingsService
        from sqlalchemy import select
        from app.utils.helpers import generate_code

        user = await self.db.scalar(select(User).where(User.id == user_id))
        hospital_id = user.hospital_id if user and user.hospital_id else 1
        billing_settings = await SettingsService(self.db).get_billing_settings(hospital_id)

        billing = Billing(
            patient_id=data.patient_id,
            bill_number=generate_code(billing_settings.get("invoice_prefix", "BIL")),
            discount_percent=data.discount_percent,
            discount_amount=data.discount_amount,
            due_date=due_date,
            notes=data.notes,
            insurance_id=None,
            appointment_id=data.appointment_id,
            created_by=user_id,
        )
        billing = await self.repo.create(billing)

        subtotal = 0.0
        for item_data in data.items:
            _, gst_amt, line_total = calculate_line_total(
                item_data.quantity, item_data.unit_price, item_data.gst_rate
            )
            item = BillItem(
                billing_id=billing.id,
                description=item_data.description,
                quantity=item_data.quantity,
                unit_price=item_data.unit_price,
                gst_rate=item_data.gst_rate,
                gst_amount=gst_amt,
                line_total=line_total,
                item_type=item_data.item_type,
            )
            await self.repo.add_item(item)
            subtotal += item_data.quantity * item_data.unit_price

        billing.subtotal = subtotal
        billing = await self._recalculate_billing(billing)
        billing = await self.repo.get_by_id(billing.id)
        await self.audit_repo.create("create", "billing", user_id=user_id, resource_id=str(billing.id))

        from app.services.transaction_history_service import TransactionHistoryService
        await TransactionHistoryService(self.db).create_event(
            event_type="INVOICE_CREATED",
            reference_no=billing.bill_number,
            description=f"Billing Invoice Created: {billing.notes or ''}",
            amount=billing.total_amount,
            source_module="billing",
            source_id=billing.id,
            status="completed",
            user_id=user_id
        )

        return self._to_response(billing)

    async def _paginate_combined_billings(
        self,
        page: int,
        size: int,
        sort_by: str,
        sort_order: str,
        status: str | None,
        patient_id: int | None,
        q: str | None,
        start_date: date | None = None,
        end_date: date | None = None,
        bill_type: Any | None = None,
    ):
        from app.models.billing_model import Billing
        from app.models.pharmacy_model import PharmacyInvoice, PharmacyInvoiceItem
        from sqlalchemy import select, or_, func, union_all
        from sqlalchemy.sql.expression import literal
        from sqlalchemy.orm import selectinload

        sort_field_b = None
        sort_field_p = None
        
        if sort_by in ["created_at", "status", "total_amount"]:
            sort_field_b = getattr(Billing, sort_by)
            sort_field_p = getattr(PharmacyInvoice, sort_by)
        elif sort_by == "bill_number":
            sort_field_b = Billing.bill_number
            sort_field_p = PharmacyInvoice.invoice_number
        else:
            sort_field_b = Billing.created_at
            sort_field_p = PharmacyInvoice.created_at

        b_query = select(
            Billing.id.label("id"),
            literal("billing").label("source"),
            sort_field_b.label("sort_val"),
            Billing.id.label("tie_breaker")
        ).where(Billing.is_deleted == False)

        p_query = select(
            PharmacyInvoice.id.label("id"),
            literal("pharmacy").label("source"),
            sort_field_p.label("sort_val"),
            PharmacyInvoice.id.label("tie_breaker")
        ).where(PharmacyInvoice.is_deleted == False)

        if status:
            b_query = b_query.where(Billing.status == status)
            p_query = p_query.where(PharmacyInvoice.status == status)
        if patient_id:
            b_query = b_query.where(Billing.patient_id == patient_id)
            p_query = p_query.where(PharmacyInvoice.patient_id == patient_id)
        if start_date:
            start_datetime = datetime.combine(start_date, datetime.min.time())
            b_query = b_query.where(Billing.created_at >= start_datetime)
            p_query = p_query.where(PharmacyInvoice.created_at >= start_datetime)
        if end_date:
            end_datetime = datetime.combine(end_date, datetime.max.time())
            b_query = b_query.where(Billing.created_at <= end_datetime)
            p_query = p_query.where(PharmacyInvoice.created_at <= end_datetime)
        if bill_type:
            bt_val = bill_type.value if hasattr(bill_type, "value") else str(bill_type).lower().strip()
            if bt_val == "pharmacy":
                b_query = b_query.where(func.lower(Billing.bill_number).like("rec%"))
                p_query = p_query.where(func.lower(PharmacyInvoice.invoice_number).like("rec%"))
            elif bt_val == "consultation":
                b_query = b_query.where(
                    or_(
                        func.lower(Billing.bill_number).like("bill%"),
                        func.lower(Billing.bill_number).like("bil%"),
                    )
                )
                p_query = p_query.where(
                    or_(
                        func.lower(PharmacyInvoice.invoice_number).like("bill%"),
                        func.lower(PharmacyInvoice.invoice_number).like("bil%"),
                    )
                )
        if q:
            pattern = f"%{q.lower()}%"
            b_query = b_query.where(
                or_(
                    func.lower(Billing.bill_number).like(pattern),
                    func.lower(Billing.notes).like(pattern),
                )
            )
            p_query = p_query.where(
                func.lower(PharmacyInvoice.invoice_number).like(pattern)
            )

        combined_query = union_all(b_query, p_query)

        subq = combined_query.subquery()
        count_query = select(func.count()).select_from(subq)
        total = await self.db.scalar(count_query) or 0

        skip = (page - 1) * size
        
        order_col = subq.c.sort_val.desc() if sort_order == "desc" else subq.c.sort_val.asc()
        tie_col = subq.c.tie_breaker.desc() if sort_order == "desc" else subq.c.tie_breaker.asc()
        
        paginated_query = select(subq.c.id, subq.c.source).order_by(order_col, tie_col).offset(skip).limit(size)
        
        rows = await self.db.execute(paginated_query)
        id_source_list = list(rows.all())
        
        billing_ids = [r[0] for r in id_source_list if r[1] == "billing"]
        pharmacy_ids = [r[0] for r in id_source_list if r[1] == "pharmacy"]

        billing_map = {}
        if billing_ids:
            b_full = await self.db.execute(
                select(Billing).where(Billing.id.in_(billing_ids)).options(
                    selectinload(Billing.items), selectinload(Billing.payments)
                )
            )
            for b in b_full.scalars().unique().all():
                billing_map[b.id] = self._to_response(b)

        pharmacy_map = {}
        if pharmacy_ids:
            p_full = await self.db.execute(
                select(PharmacyInvoice).where(PharmacyInvoice.id.in_(pharmacy_ids)).options(
                    selectinload(PharmacyInvoice.items).selectinload(PharmacyInvoiceItem.medicine)
                )
            )
            for p in p_full.scalars().unique().all():
                pharmacy_map[p.id] = self._pharmacy_invoice_to_billing_response(p)

        ordered_items = []
        for r_id, r_source in id_source_list:
            if r_source == "billing" and r_id in billing_map:
                ordered_items.append(billing_map[r_id])
            elif r_source == "pharmacy" and r_id in pharmacy_map:
                ordered_items.append(pharmacy_map[r_id])

        return build_paginated_result(ordered_items, total, page, size)

    async def list_billings(
        self, page: int = 1, size: int = 20, sort_by: str = "created_at",
        sort_order: str = "desc", status: str | None = None, patient_id: int | None = None,
        start_date: date | None = None, end_date: date | None = None,
        bill_type: Any | None = None,
    ):
        return await self._paginate_combined_billings(
            page=page, size=size, sort_by=sort_by, sort_order=sort_order,
            status=status, patient_id=patient_id, q=None,
            start_date=start_date, end_date=end_date,
            bill_type=bill_type,
        )

    async def search(
        self, q: str, page: int = 1, size: int = 20, status: str | None = None,
        start_date: date | None = None, end_date: date | None = None,
        bill_type: Any | None = None,
    ):
        return await self._paginate_combined_billings(
            page=page, size=size, sort_by="created_at", sort_order="desc",
            status=status, patient_id=None, q=q,
            start_date=start_date, end_date=end_date,
            bill_type=bill_type,
        )

    async def get_by_id(self, billing_id: int) -> BillingResponse:
        billing = await self.repo.get_by_id(billing_id)
        if not billing:
            from app.models.pharmacy_model import PharmacyInvoice, PharmacyInvoiceItem
            from sqlalchemy import select
            from sqlalchemy.orm import selectinload
            stmt = select(PharmacyInvoice).where(
                PharmacyInvoice.id == billing_id,
                PharmacyInvoice.is_deleted == False
            ).options(
                selectinload(PharmacyInvoice.items).selectinload(PharmacyInvoiceItem.medicine)
            )
            res = await self.db.execute(stmt)
            invoice = res.scalar_one_or_none()
            if not invoice:
                raise NotFoundException("Billing record not found")
            return self._pharmacy_invoice_to_billing_response(invoice)
        return self._to_response(billing)

    async def update(self, billing_id: int, data: BillingUpdate, user_id: int) -> BillingResponse:
        billing = await self.repo.get_by_id(billing_id)
        if not billing:
            from app.models.pharmacy_model import PharmacyInvoice, PharmacyInvoiceItem
            from sqlalchemy import select
            from sqlalchemy.orm import selectinload
            stmt = select(PharmacyInvoice).where(
                PharmacyInvoice.id == billing_id,
                PharmacyInvoice.is_deleted == False
            ).options(
                selectinload(PharmacyInvoice.items).selectinload(PharmacyInvoiceItem.medicine)
            )
            res = await self.db.execute(stmt)
            invoice = res.scalar_one_or_none()
            if not invoice:
                raise NotFoundException("Billing record not found")

            dump = data.model_dump(exclude_unset=True)
            if "status" in dump and dump["status"] is not None:
                invoice.status = dump["status"]

            if "discount_percent" in dump and dump["discount_percent"] is not None:
                invoice.discount_percentage = float(dump["discount_percent"])
            elif "discount_percentage" in dump and dump["discount_percentage"] is not None:
                invoice.discount_percentage = float(dump["discount_percentage"])

            if "gst_rate" in dump and dump["gst_rate"] is not None:
                invoice.tax_percentage = float(dump["gst_rate"])
            elif "tax_percentage" in dump and dump["tax_percentage"] is not None:
                invoice.tax_percentage = float(dump["tax_percentage"])

            subtotal = float(invoice.subtotal or 0.0)
            discount_amount = round((subtotal * invoice.discount_percentage) / 100, 2)
            tax_amount = round((subtotal - discount_amount) * invoice.tax_percentage / 100, 2)

            invoice.discount_amount = discount_amount
            invoice.tax_amount = tax_amount
            invoice.gst_amount = tax_amount
            invoice.total_amount = round(subtotal - discount_amount + tax_amount, 2)
            if invoice.status == "paid":
                invoice.paid_amount = invoice.total_amount

            await self.db.flush()
            await self.audit_repo.create("update", "pharmacy_invoice", user_id=user_id, resource_id=str(invoice.id))
            return self._pharmacy_invoice_to_billing_response(invoice)

        non_nullable_fields = ["discount_percent", "discount_amount", "gst_rate", "tax_amount", "status"]
        for field in non_nullable_fields:
            if field in data.model_fields_set and getattr(data, field) is None:
                raise BadRequestException(f"Field '{field}' cannot be null")

        # Ignore and do not validate insurance_id from data
        dump = data.model_dump(exclude_unset=True)
        dump.pop("insurance_id", None)  # Ignore insurance_id payload

        if "due_date" in dump and dump["due_date"] is not None:
            dt = dump["due_date"]
            if dt.tzinfo is not None:
                from datetime import timezone as py_timezone
                dump["due_date"] = dt.astimezone(py_timezone.utc).replace(tzinfo=None)

        from app.models.user_model import User
        from app.services.settings_service import SettingsService
        from sqlalchemy import select
        
        user = await self.db.scalar(select(User).where(User.id == user_id))
        hospital_id = user.hospital_id if user and user.hospital_id else 1
        billing_settings = await SettingsService(self.db).get_billing_settings(hospital_id)

        if "items" in dump:
            items_data = dump.pop("items")
            # Remove old items
            billing.items.clear()
            if items_data:
                for item_data in items_data:
                    desc = item_data.get("description")
                    qty = item_data.get("quantity", 1)
                    price = item_data.get("unit_price")
                    gst_r = item_data.get("gst_rate", billing_settings.get("gst_percentage", 18.0))
                    item_t = item_data.get("item_type", "service")
                    _, gst_amt, line_total = calculate_line_total(qty, price, gst_r)
                    item = BillItem(
                        billing_id=billing.id,
                        description=desc,
                        quantity=qty,
                        unit_price=price,
                        gst_rate=gst_r,
                        gst_amount=gst_amt,
                        line_total=line_total,
                        item_type=item_t
                    )
                    billing.items.append(item)

        for key, value in dump.items():
            setattr(billing, key, value)

        # Always set insurance_id to None during billing update
        billing.insurance_id = None

        billing = await self._recalculate_billing(billing)
        billing = await self.repo.get_by_id(billing.id)
        await self.audit_repo.create("update", "billing", user_id=user_id, resource_id=str(billing.id))
        return self._to_response(billing)

    async def delete(self, billing_id: int, user_id: int) -> None:
        billing = await self.repo.get_by_id(billing_id)
        if not billing:
            from app.models.pharmacy_model import PharmacyInvoice
            from sqlalchemy import select
            stmt = select(PharmacyInvoice).where(PharmacyInvoice.id == billing_id, PharmacyInvoice.is_deleted == False)
            res = await self.db.execute(stmt)
            invoice = res.scalar_one_or_none()
            if not invoice:
                raise NotFoundException("Billing record not found")
            from app.services.pharmacy_service import PharmacyService
            await PharmacyService(self.db).delete_invoice(billing_id)
            await self.audit_repo.create("delete", "pharmacy_invoice", user_id=user_id, resource_id=str(billing_id))
            return
        await self.repo.soft_delete(billing)
        await self.audit_repo.create("delete", "billing", user_id=user_id, resource_id=str(billing.id))

    async def collect_payment(self, billing_id: int, data: PaymentCreate, user_id: int) -> PaymentResponse:
        billing = await self.repo.get_by_id_for_update(billing_id)
        if not billing:
            raise NotFoundException("Billing record not found")
        if billing.status == BillingStatus.CANCELLED:
            raise BadRequestException("Cannot collect payment on cancelled bill")
        total_amt = billing.total_amount or 0.0
        paid_amt = billing.paid_amount or 0.0
        balance_due = round(total_amt - paid_amt, 2)
        if data.amount > balance_due:
            raise BadRequestException("Payment amount exceeds balance due")

        # Defensive normalization
        method = data.payment_method.strip().lower()
        if method == "cheques":
            method = "cheque"
        data.payment_method = method

        if data.transaction_ref is not None:
            data.transaction_ref = data.transaction_ref.strip()
            if not data.transaction_ref:
                data.transaction_ref = None

        payment = Payment(
            billing_id=billing_id,
            amount=data.amount,
            payment_method=data.payment_method,
            transaction_ref=data.transaction_ref,
            payment_date=utc_now(),
            received_by=user_id,
        )
        payment = await self.repo.add_payment(payment)
        billing.paid_amount = round(billing.paid_amount + data.amount, 2)
        await self._recalculate_billing(billing)
        await self.audit_repo.create("payment", "billing", user_id=user_id, resource_id=str(billing_id))

        from app.services.transaction_history_service import TransactionHistoryService
        await TransactionHistoryService(self.db).create_event(
            event_type="PAYMENT_RECEIVED",
            reference_no=payment.transaction_ref or f"PAY-{payment.id}",
            description=f"Payment Received on bill {billing.bill_number} via {payment.payment_method}",
            amount=payment.amount,
            source_module="payments",
            source_id=payment.id,
            status="completed",
            user_id=user_id
        )

        return PaymentResponse.model_validate(payment)

    async def process_refund(self, billing_id: int, data: RefundCreate, user_id: int) -> PaymentResponse:
        billing = await self.repo.get_by_id_for_update(billing_id)
        if not billing:
            raise NotFoundException("Billing record not found")
        if data.amount > billing.paid_amount:
            raise BadRequestException("Refund amount exceeds paid amount")

        payment = Payment(
            billing_id=billing_id,
            amount=data.amount,
            payment_method="refund",
            payment_date=utc_now(),
            is_refund=True,
            refund_reason=data.refund_reason,
            received_by=user_id,
            status="completed",
        )
        payment = await self.repo.add_payment(payment)
        billing.paid_amount = round(billing.paid_amount - data.amount, 2)
        billing = await self._recalculate_billing(billing)
        if billing.paid_amount == 0 and billing.balance_amount > 0:
            billing.status = BillingStatus.REFUNDED
        await self.audit_repo.create("refund", "billing", user_id=user_id, resource_id=str(billing_id))

        from app.services.transaction_history_service import TransactionHistoryService
        await TransactionHistoryService(self.db).create_event(
            event_type="REFUND_ISSUED",
            reference_no=payment.transaction_ref or f"REF-{payment.id}",
            description=f"Refund Issued for bill {billing.bill_number}: {payment.refund_reason or ''}",
            amount=payment.amount,
            source_module="refunds",
            source_id=payment.id,
            status="completed",
            user_id=user_id
        )

        return PaymentResponse.model_validate(payment)

    async def generate_invoice(self, billing_id: int, user_id: int) -> tuple[str, bytes]:
        billing = await self.repo.get_by_id(billing_id)
        if not billing:
            raise NotFoundException("Billing record not found")
        patient = await self.patient_repo.get_by_id(billing.patient_id)
        patient_name = f"{patient.first_name or ''} {patient.last_name or ''}".strip() if patient else "Walk-in Patient"
        patient_phone = patient.phone if patient else "-"
        patient_email = patient.email if patient else "-"

        items = [
            {
                "description": i.description,
                "quantity": i.quantity,
                "unit_price": f"{i.unit_price:.2f}",
                "gst_rate": f"{i.gst_rate:.1f}",
                "line_total": f"{i.line_total:.2f}",
            }
            for i in billing.items
        ]

        path, pdf_bytes = await generate_invoice_pdf(
            billing.bill_number,
            {
                "patient_name": patient_name,
                "patient_phone": patient_phone,
                "patient_email": patient_email,
                "date": utc_now().strftime("%Y-%m-%d"),
                "items": items,
                "subtotal": f"{billing.subtotal:.2f}",
                "discount_percent": f"{billing.discount_percent:.1f}",
                "discount_amount": f"{billing.discount_amount:.2f}",
                "gst_rate": f"{billing.gst_rate:.1f}",
                "gst_amount": f"{billing.gst_amount:.2f}",
                "tax_amount": f"{billing.tax_amount:.2f}",
                "total_amount": f"{billing.total_amount:.2f}",
                "paid_amount": f"{billing.paid_amount:.2f}",
                "balance_amount": f"{billing.balance_amount:.2f}",
                "status": billing.status.title(),
                "notes": billing.notes or "",
            },
        )
        billing.invoice_path = path
        await self.repo.update(billing)
        await self.audit_repo.create("export", "billing", user_id=user_id, resource_id=str(billing.id))
        return path, pdf_bytes

    async def get_pending_payments(self, page: int = 1, size: int = 20):
        skip = (page - 1) * size
        items = await self.repo.get_pending_payments(skip=skip, limit=size)
        total = await self.repo.count_pending_payments()
        return build_paginated_result([self._to_response(b) for b in items], total, page, size)

    async def get_revenue_summary(self) -> BillingSummary:
        data = await self.repo.get_revenue_summary()

        from app.models.pharmacy_model import PharmacyInvoice
        from sqlalchemy import select, func

        pharmacy_stmt = select(
            func.coalesce(func.sum(PharmacyInvoice.total_amount), 0.0),
            func.coalesce(func.sum(PharmacyInvoice.paid_amount), 0.0)
        ).where(
            PharmacyInvoice.is_deleted == False
        )
        res = await self.db.execute(pharmacy_stmt)
        row = res.first()

        pharm_billed = float(row[0] if row else 0.0)
        pharm_collected = float(row[1] if row else 0.0)
        pharm_pending = max(0.0, pharm_billed - pharm_collected)

        pharm_pending_count_stmt = select(func.count(PharmacyInvoice.id)).where(
            PharmacyInvoice.is_deleted == False,
            PharmacyInvoice.paid_amount < PharmacyInvoice.total_amount
        )
        pharm_pending_count = await self.db.scalar(pharm_pending_count_stmt) or 0

        pharm_overdue_count_stmt = select(func.count(PharmacyInvoice.id)).where(
            PharmacyInvoice.is_deleted == False,
            PharmacyInvoice.status == "overdue"
        )
        pharm_overdue = await self.db.scalar(pharm_overdue_count_stmt) or 0

        data["total_revenue"] = round(data["total_revenue"] + pharm_collected, 2)
        data["total_paid"] = round(data["total_paid"] + pharm_collected, 2)
        data["total_pending"] = round(data["total_pending"] + pharm_pending, 2)
        data["overdue_count"] = data["overdue_count"] + pharm_overdue
        data["pending_count"] = data["pending_count"] + pharm_pending_count

        return BillingSummary(**data)

    async def get_daily_report(self, target_date: date | None = None) -> DailyCollectionSummary:
        target = target_date or date.today()
        data = await self.repo.get_daily_collection(target)

        start = datetime.combine(target, datetime.min.time())
        end = datetime.combine(target, datetime.max.time())

        from app.models.pharmacy_model import PharmacyInvoice
        from sqlalchemy import select, func

        pharmacy_stmt = select(
            func.coalesce(func.sum(PharmacyInvoice.paid_amount), 0.0),
            func.count(PharmacyInvoice.id)
        ).where(
            PharmacyInvoice.is_deleted == False,
            PharmacyInvoice.created_at >= start,
            PharmacyInvoice.created_at <= end,
            PharmacyInvoice.paid_amount > 0.0
        )
        res = await self.db.execute(pharmacy_stmt)
        row = res.first()
        pharm_collected = float(row[0] if row else 0.0)
        pharm_count = int(row[1] if row else 0)

        if pharm_collected > 0:
            data["total_collected"] = round(data["total_collected"] + pharm_collected, 2)
            data["payment_count"] = data["payment_count"] + pharm_count

        # Ensure total_collected cannot be negative
        data["total_collected"] = max(0.0, round(float(data.get("total_collected", 0.0)), 2))

        # Ensure by_method contains only rounded non-pharmacy payment methods with positive amounts
        by_method = {}
        for k, v in data.get("by_method", {}).items():
            if k and str(k).lower() != "pharmacy":
                by_method[str(k)] = round(abs(float(v)), 2)
        data["by_method"] = by_method

        return DailyCollectionSummary(date=str(target), **data)

    async def get_yearly_report(self, year: int | None = None) -> RevenueReport:
        target_year = year or utc_now().year
        start = datetime.combine(date(target_year, 1, 1), datetime.min.time())
        end = datetime.combine(date(target_year, 12, 31), datetime.max.time())
        label = str(target_year)

        data = await self.repo.get_period_report(start, end)

        from app.models.pharmacy_model import PharmacyInvoice
        from sqlalchemy import select, func

        pharmacy_stmt = select(
            func.coalesce(func.sum(PharmacyInvoice.total_amount), 0.0),
            func.coalesce(func.sum(PharmacyInvoice.paid_amount), 0.0),
            func.count(PharmacyInvoice.id)
        ).where(
            PharmacyInvoice.is_deleted == False,
            PharmacyInvoice.created_at >= start,
            PharmacyInvoice.created_at <= end
        )
        res = await self.db.execute(pharmacy_stmt)
        row = res.first()

        pharm_billed = float(row[0] if row else 0.0)
        pharm_collected = float(row[1] if row else 0.0)
        pharm_bill_count = int(row[2] if row else 0)

        pharm_pending = max(0.0, pharm_billed - pharm_collected)

        pharm_pay_count_stmt = select(func.count(PharmacyInvoice.id)).where(
            PharmacyInvoice.is_deleted == False,
            PharmacyInvoice.created_at >= start,
            PharmacyInvoice.created_at <= end,
            PharmacyInvoice.paid_amount > 0.0
        )
        pharm_payment_count = await self.db.scalar(pharm_pay_count_stmt) or 0

        data["total_billed"] = round(data["total_billed"] + pharm_billed, 2)
        data["total_collected"] = round(data["total_collected"] + pharm_collected, 2)
        data["total_pending"] = round(data["total_pending"] + pharm_pending, 2)
        data["total_refunded"] = round(float(data.get("total_refunded", 0.0)), 2)
        data["bill_count"] = data["bill_count"] + pharm_bill_count
        data["payment_count"] = data["payment_count"] + pharm_payment_count

        return RevenueReport(period=label, **data)

    async def get_period_report(
        self,
        period: str,
        target_date: date | None = None,
        year: int | None = None,
        month: int | None = None,
    ) -> RevenueReport:
        if period == "monthly":
            import calendar
            now = utc_now()
            y = year or (target_date.year if target_date else now.year)
            m = month or (target_date.month if target_date else now.month)
            last_day = calendar.monthrange(y, m)[1]
            start = datetime.combine(date(y, m, 1), datetime.min.time())
            end = datetime.combine(date(y, m, last_day), datetime.max.time())
            label = start.strftime("%Y-%m")
        elif target_date:
            y, m = target_date.year, target_date.month
            if period == "daily":
                start = datetime.combine(target_date, datetime.min.time())
                end = datetime.combine(target_date, datetime.max.time())
                label = str(target_date)
            else: # yearly
                start = datetime.combine(date(y, 1, 1), datetime.min.time())
                end = datetime.combine(date(y, 12, 31), datetime.max.time())
                label = str(y)
        else:
            now = utc_now()
            if period == "daily":
                start = datetime.combine(now.date(), datetime.min.time())
                end = now
                label = str(now.date())
            else:
                y = year or now.year
                start = datetime.combine(date(y, 1, 1), datetime.min.time())
                end = datetime.combine(date(y, 12, 31), datetime.max.time())
                label = str(y)
        data = await self.repo.get_period_report(start, end)

        from app.models.pharmacy_model import PharmacyInvoice
        from sqlalchemy import select, func

        pharmacy_stmt = select(
            func.coalesce(func.sum(PharmacyInvoice.total_amount), 0.0),
            func.coalesce(func.sum(PharmacyInvoice.paid_amount), 0.0),
            func.count(PharmacyInvoice.id)
        ).where(
            PharmacyInvoice.is_deleted == False,
            PharmacyInvoice.created_at >= start,
            PharmacyInvoice.created_at <= end
        )
        res = await self.db.execute(pharmacy_stmt)
        row = res.first()

        pharm_billed = float(row[0] if row else 0.0)
        pharm_collected = float(row[1] if row else 0.0)
        pharm_bill_count = int(row[2] if row else 0)

        pharm_pending = max(0.0, pharm_billed - pharm_collected)

        pharm_pay_count_stmt = select(func.count(PharmacyInvoice.id)).where(
            PharmacyInvoice.is_deleted == False,
            PharmacyInvoice.created_at >= start,
            PharmacyInvoice.created_at <= end,
            PharmacyInvoice.paid_amount > 0.0
        )
        pharm_payment_count = await self.db.scalar(pharm_pay_count_stmt) or 0

        data["total_billed"] = round(data["total_billed"] + pharm_billed, 2)
        data["total_collected"] = round(data["total_collected"] + pharm_collected, 2)
        data["total_pending"] = round(data["total_pending"] + pharm_pending, 2)
        data["total_refunded"] = round(float(data.get("total_refunded", 0.0)), 2)
        data["bill_count"] = data["bill_count"] + pharm_bill_count
        data["payment_count"] = data["payment_count"] + pharm_payment_count

        return RevenueReport(period=label, **data)

    async def generate_billing_bulk_template(self):
        from io import BytesIO
        import openpyxl
        from sqlalchemy import select
        from app.models.patient_model import Patient
        
        # Try to find an actual active patient name
        stmt = select(Patient).where(Patient.is_deleted == False).limit(1)
        res = await self.db.execute(stmt)
        patient = res.scalar_one_or_none()
        patient_name = f"{patient.first_name} {patient.last_name}".strip() if patient else "Rahul Sharma"
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Billings Bulk Import"
        
        headers = [
            "group_key", "Patient Name", "discount_percent", "discount_amount", "due_date", "notes",
            "Appointment Number", "item_description", "item_quantity", "item_unit_price", "item_gst_rate", "item_type"
        ]
        ws.append(headers)
        
        # One valid sample billing with 2 items sharing group_key=1
        ws.append([
            1,
            patient_name,
            10.0,
            0.0,
            "2026-08-30 12:00:00",
            "Routine clinic visit",
            "APT-20260825E57D0B",
            "Consultation Fee",
            1,
            500.00,
            18.0,
            "service"
        ])
        ws.append([
            1,
            patient_name,
            10.0,
            0.0,
            "2026-08-30 12:00:00",
            "Routine clinic visit",
            "APT-20260825E57D0B",
            "Disposable Syringe",
            2,
            25.00,
            12.0,
            "pharmacy_item"
        ])
        
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream

    async def import_billing_from_excel(self, file, user_id: int) -> dict:
        from io import BytesIO
        from pydantic import ValidationError
        import openpyxl
        
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents))
        ws = wb.active
        
        header_row = next(ws.iter_rows(max_row=1, values_only=True), None)
        if not header_row:
            raise BadRequestException("The uploaded file is empty or has no headers.")
            
        headers = [str(h).strip().lower() for h in header_row if h is not None]
        required_headers = {"group_key", "patient name", "item_description", "item_unit_price"}
        if not required_headers.issubset(set(headers)):
            raise BadRequestException("Missing required headers in the upload template.")
            
        total_rows = 0
        created = 0
        failed = 0
        errors = []
        
        groups = {}
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None for cell in row):
                continue
                
            total_rows += 1
            row_dict = {}
            for header, val in zip(headers, row):
                if val is None or str(val).strip() == "":
                    row_dict[header] = None
                else:
                    row_dict[header] = val
                    
            gk = row_dict.get("group_key")
            if gk is None or str(gk).strip() == "":
                errors.append({"group_key": "N/A", "row": row_idx, "error": "Missing group_key in row."})
                failed += 1
                continue
                
            gk_str = str(gk).strip()
            if gk_str not in groups:
                groups[gk_str] = []
            groups[gk_str].append((row_idx, row_dict))
            
        total_bills = len(groups)
        
        for gk_str, rows in groups.items():
            # Check consistency of billing-level parameters
            first_idx, first_row = rows[0]
            
            patient_name_raw = first_row.get("patient name")
            discount_percent_raw = first_row.get("discount_percent")
            discount_amount_raw = first_row.get("discount_amount")
            due_date_raw = first_row.get("due_date")
            notes_raw = first_row.get("notes")
            appointment_number_raw = first_row.get("appointment number")
            
            def val_to_str(val):
                if val is None:
                    return ""
                return str(val).strip().lower()
                
            conflict_found = False
            for idx, r in rows[1:]:
                if (val_to_str(r.get("patient name")) != val_to_str(patient_name_raw) or
                    val_to_str(r.get("discount_percent")) != val_to_str(discount_percent_raw) or
                    val_to_str(r.get("discount_amount")) != val_to_str(discount_amount_raw) or
                    val_to_str(r.get("due_date")) != val_to_str(due_date_raw) or
                    val_to_str(r.get("notes")) != val_to_str(notes_raw) or
                    val_to_str(r.get("appointment number")) != val_to_str(appointment_number_raw)):
                    conflict_found = True
                    conflict_idx = idx
                    break
                    
            if conflict_found:
                failed += 1
                errors.append({
                    "group_key": gk_str,
                    "row": conflict_idx,
                    "error": "Conflict in billing-level fields within the same group_key."
                })
                continue
                
            try:
                # 1. patient resolution
                patient_name_str = str(patient_name_raw).strip() if patient_name_raw is not None else ""
                if not patient_name_str:
                    raise BadRequestException("Patient Name is required.")
                
                from sqlalchemy import select, or_
                from app.models.patient_model import Patient
                
                parts = [p for p in patient_name_str.split() if p]
                if not parts:
                    raise BadRequestException("Patient Name cannot be empty.")
                
                stmt = select(Patient).where(Patient.is_deleted == False)
                filters = []
                for part in parts:
                    filters.append(Patient.first_name.ilike(f"%{part}%"))
                    filters.append(Patient.last_name.ilike(f"%{part}%"))
                if filters:
                    stmt = stmt.where(or_(*filters))
                    
                result = await self.db.execute(stmt)
                candidates = result.scalars().all()
                
                matching_patients = []
                normalized_target = " ".join(parts).lower()
                for p in candidates:
                    p_full_name = f"{p.first_name} {p.last_name}".strip().lower()
                    if p_full_name == normalized_target:
                        matching_patients.append(p)
                        
                if not matching_patients:
                    raise BadRequestException(f"Patient not found: {patient_name_str}")
                if len(matching_patients) > 1:
                    raise BadRequestException(f"Multiple patients found with name {patient_name_str}. Use a unique patient.")
                    
                patient_id = matching_patients[0].id
                
                appointment_id = None
                if appointment_number_raw is not None:
                    appt_num_str = str(appointment_number_raw).strip()
                    if appt_num_str:
                        from app.models.appointment_model import Appointment
                        appt_stmt = select(Appointment.id).where(Appointment.appointment_number == appt_num_str)
                        appt_res = await self.db.execute(appt_stmt)
                        appointment_id = appt_res.scalar_one_or_none()
                        if not appointment_id:
                            raise BadRequestException(f"Appointment with number {appt_num_str} not found")
                
                # 2. discount_percent
                discount_percent = 0.0
                if discount_percent_raw is not None:
                    try:
                        discount_percent = float(discount_percent_raw)
                    except (ValueError, TypeError):
                        raise BadRequestException(f"Invalid discount_percent: {discount_percent_raw}")
                        
                # 3. discount_amount
                discount_amount = 0.0
                if discount_amount_raw is not None:
                    try:
                        discount_amount = float(discount_amount_raw)
                    except (ValueError, TypeError):
                        raise BadRequestException(f"Invalid discount_amount: {discount_amount_raw}")
                        
                # 4. due_date parsing
                due_date = None
                if due_date_raw is not None:
                    if isinstance(due_date_raw, datetime):
                        due_date = due_date_raw
                    elif isinstance(due_date_raw, date) and not isinstance(due_date_raw, datetime):
                        due_date = datetime.combine(due_date_raw, datetime.min.time())
                    else:
                        try:
                            p_str = str(due_date_raw).strip()
                            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d"):
                                try:
                                    due_date = datetime.strptime(p_str, fmt)
                                    break
                                except ValueError:
                                    continue
                            if due_date is None:
                                due_date = datetime.fromisoformat(p_str)
                        except Exception:
                            raise BadRequestException(f"Invalid datetime format for due_date: {due_date_raw}")
                            
                notes = str(notes_raw).strip() if notes_raw is not None else None
                
                # 5. Build nested items
                items = []
                for idx, r in rows:
                    desc_raw = r.get("item_description")
                    if desc_raw is None:
                        raise BadRequestException("item_description is required.")
                    description = str(desc_raw).strip()
                    
                    qty_raw = r.get("item_quantity")
                    quantity = 1
                    if qty_raw is not None:
                        try:
                            f_qty = float(qty_raw)
                            if not f_qty.is_integer() or f_qty < 1:
                                raise ValueError()
                            quantity = int(f_qty)
                        except (ValueError, TypeError):
                            raise BadRequestException(f"Invalid quantity: {qty_raw} at row {idx}")
                            
                    up_raw = r.get("item_unit_price")
                    if up_raw is None:
                        raise BadRequestException("item_unit_price is required.")
                    try:
                        unit_price = float(up_raw)
                    except (ValueError, TypeError):
                        raise BadRequestException(f"Invalid unit_price: {up_raw} at row {idx}")
                        
                    gst_raw = r.get("item_gst_rate")
                    gst_rate = 18.0
                    if gst_raw is not None:
                        try:
                            gst_rate = float(gst_raw)
                        except (ValueError, TypeError):
                            raise BadRequestException(f"Invalid gst_rate: {gst_raw} at row {idx}")
                            
                    type_raw = r.get("item_type")
                    item_type = str(type_raw).strip() if type_raw is not None else "service"
                    
                    from app.schemas.billing_schema import BillItemCreate
                    items.append(
                        BillItemCreate(
                            description=description,
                            quantity=quantity,
                            unit_price=unit_price,
                            gst_rate=gst_rate,
                            item_type=item_type
                        )
                    )
                    
                # Validate the entire BillingCreate object before creation
                from app.schemas.billing_schema import BillingCreate
                bill_create = BillingCreate(
                    patient_id=patient_id,
                    discount_percent=discount_percent,
                    discount_amount=discount_amount,
                    due_date=due_date,
                    notes=notes,
                    appointment_id=appointment_id,
                    items=items
                )
                
                # Call billing create service method
                await self.create(bill_create, user_id)
                created += 1
                
            except ValidationError as e:
                failed += 1
                err_msg = "; ".join([f"{'.'.join(str(loc) for loc in error['loc'])}: {error['msg']}" for error in e.errors()])
                errors.append({
                    "group_key": gk_str,
                    "row": first_idx,
                    "error": err_msg
                })
            except BadRequestException as e:
                failed += 1
                errors.append({
                    "group_key": gk_str,
                    "row": first_idx,
                    "error": str(e.detail)
                })
            except NotFoundException as e:
                failed += 1
                errors.append({
                    "group_key": gk_str,
                    "row": first_idx,
                    "error": str(e.detail)
                })
            except ConflictException as e:
                failed += 1
                errors.append({
                    "group_key": gk_str,
                    "row": first_idx,
                    "error": str(e.detail)
                })
            except Exception as e:
                failed += 1
                errors.append({
                    "group_key": gk_str,
                    "row": first_idx,
                    "error": str(e)
                })
                
        await self.db.flush()
        return {
            "total_rows": total_rows,
            "total_bills": total_bills,
            "created": created,
            "failed": failed,
            "errors": errors
        }

    async def export_billings(
        self,
        format_type: str,
        status: str | None = None,
        patient_id: int | None = None,
        start_date: date | None = None,
        end_date: date | None = None,
        q: str | None = None,
        sort_by: str = "created_at",
        sort_order: str = "desc"
    ):
        from app.models.billing_model import Billing
        from app.models.pharmacy_model import PharmacyInvoice, PharmacyInvoiceItem
        from sqlalchemy import select, or_, func, union_all
        from sqlalchemy.sql.expression import literal
        from sqlalchemy.orm import selectinload
        from datetime import datetime, date
        from io import BytesIO

        sort_field_b = None
        sort_field_p = None
        
        if sort_by in ["created_at", "status", "total_amount"]:
            sort_field_b = getattr(Billing, sort_by)
            sort_field_p = getattr(PharmacyInvoice, sort_by)
        elif sort_by == "bill_number":
            sort_field_b = Billing.bill_number
            sort_field_p = PharmacyInvoice.invoice_number
        else:
            sort_field_b = Billing.created_at
            sort_field_p = PharmacyInvoice.created_at

        b_query = select(
            Billing.id.label("id"),
            literal("billing").label("source"),
            sort_field_b.label("sort_val"),
            Billing.id.label("tie_breaker")
        ).where(Billing.is_deleted == False)

        p_query = select(
            PharmacyInvoice.id.label("id"),
            literal("pharmacy").label("source"),
            sort_field_p.label("sort_val"),
            PharmacyInvoice.id.label("tie_breaker")
        ).where(PharmacyInvoice.is_deleted == False)

        if status:
            b_query = b_query.where(Billing.status == status)
            p_query = p_query.where(PharmacyInvoice.status == status)
        if patient_id:
            b_query = b_query.where(Billing.patient_id == patient_id)
            p_query = p_query.where(PharmacyInvoice.patient_id == patient_id)
        if start_date:
            start_datetime = datetime.combine(start_date, datetime.min.time())
            b_query = b_query.where(Billing.created_at >= start_datetime)
            p_query = p_query.where(PharmacyInvoice.created_at >= start_datetime)
        if end_date:
            end_datetime = datetime.combine(end_date, datetime.max.time())
            b_query = b_query.where(Billing.created_at <= end_datetime)
            p_query = p_query.where(PharmacyInvoice.created_at <= end_datetime)
        if q:
            pattern = f"%{q.lower()}%"
            b_query = b_query.where(
                or_(
                    func.lower(Billing.bill_number).like(pattern),
                    func.lower(Billing.notes).like(pattern),
                )
            )
            p_query = p_query.where(
                func.lower(PharmacyInvoice.invoice_number).like(pattern)
            )

        combined_query = union_all(b_query, p_query)
        subq = combined_query.subquery()
        
        order_col = subq.c.sort_val.desc() if sort_order == "desc" else subq.c.sort_val.asc()
        tie_col = subq.c.tie_breaker.desc() if sort_order == "desc" else subq.c.tie_breaker.asc()
        
        non_paginated_query = select(subq.c.id, subq.c.source).order_by(order_col, tie_col)
        
        rows = await self.db.execute(non_paginated_query)
        id_source_list = list(rows.all())
        
        billing_ids = [r[0] for r in id_source_list if r[1] == "billing"]
        pharmacy_ids = [r[0] for r in id_source_list if r[1] == "pharmacy"]

        billing_map = {}
        if billing_ids:
            b_full = await self.db.execute(
                select(Billing).where(Billing.id.in_(billing_ids)).options(
                    selectinload(Billing.items), selectinload(Billing.payments)
                )
            )
            for b in b_full.scalars().unique().all():
                billing_map[b.id] = self._to_response(b)

        pharmacy_map = {}
        if pharmacy_ids:
            p_full = await self.db.execute(
                select(PharmacyInvoice).where(PharmacyInvoice.id.in_(pharmacy_ids)).options(
                    selectinload(PharmacyInvoice.items).selectinload(PharmacyInvoiceItem.medicine)
                )
            )
            for p in p_full.scalars().unique().all():
                pharmacy_map[p.id] = self._pharmacy_invoice_to_billing_response(p)

        ordered_items = []
        for r_id, r_source in id_source_list:
            if r_source == "billing" and r_id in billing_map:
                ordered_items.append(billing_map[r_id])
            elif r_source == "pharmacy" and r_id in pharmacy_map:
                ordered_items.append(pharmacy_map[r_id])
                
        # Map patient details
        from app.models.patient_model import Patient
        p_stmt = select(Patient.id, Patient.first_name, Patient.last_name).where(Patient.is_deleted == False)
        p_res = await self.db.execute(p_stmt)
        patients_map = {row[0]: f"{row[1]} {row[2]}" for row in p_res.all()}

        # Map appointment numbers
        from app.models.appointment_model import Appointment
        a_stmt = select(Appointment.id, Appointment.appointment_number)
        a_res = await self.db.execute(a_stmt)
        appointments_map = {row[0]: row[1] for row in a_res.all()}

        if format_type == "excel":
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Billings Report"
            
            headers = [
                "Sr. No.", "Patient Name", "bill_number", "subtotal", "discount_percent",
                "discount_amount", "gst_rate", "gst_amount", "tax_amount", "total_amount",
                "paid_amount", "balance_amount", "status", "due_date", "notes", "invoice_path",
                "Appointment Number", "created_at", "updated_at", "source",
                "item_id", "item_description", "item_quantity", "item_unit_price",
                "item_gst_rate", "item_gst_amount", "item_line_total", "item_type"
            ]
            ws.append(headers)
            
            for sr_no, b in enumerate(ordered_items, start=1):
                p_name = patients_map.get(b.patient_id, "")
                appt_num_str = appointments_map.get(b.appointment_id, "") if b.appointment_id else ""
                
                # Check if there are items to export. If not, output one row with empty item columns
                if not b.items:
                    row = [
                        sr_no, p_name, b.bill_number, f"₹{b.subtotal:.2f}", b.discount_percent,
                        f"₹{b.discount_amount:.2f}", b.gst_rate, f"₹{b.gst_amount:.2f}", f"₹{b.tax_amount:.2f}",
                        f"₹{b.total_amount:.2f}", f"₹{b.paid_amount:.2f}", f"₹{b.balance_amount:.2f}",
                        b.status, b.due_date.strftime("%Y-%m-%d %H:%M:%S") if isinstance(b.due_date, datetime) else (str(b.due_date) if b.due_date else ""),
                        b.notes or "", b.invoice_path or "", appt_num_str,
                        b.created_at.strftime("%Y-%m-%d %H:%M:%S") if isinstance(b.created_at, datetime) else str(b.created_at),
                        b.updated_at.strftime("%Y-%m-%d %H:%M:%S") if isinstance(b.updated_at, datetime) else str(b.updated_at),
                        b.source,
                        "", "", "", "", "", "", "", ""
                    ]
                    ws.append(row)
                else:
                    for item in b.items:
                        row = [
                            sr_no, p_name, b.bill_number, f"₹{b.subtotal:.2f}", b.discount_percent,
                            f"₹{b.discount_amount:.2f}", b.gst_rate, f"₹{b.gst_amount:.2f}", f"₹{b.tax_amount:.2f}",
                            f"₹{b.total_amount:.2f}", f"₹{b.paid_amount:.2f}", f"₹{b.balance_amount:.2f}",
                            b.status, b.due_date.strftime("%Y-%m-%d %H:%M:%S") if isinstance(b.due_date, datetime) else (str(b.due_date) if b.due_date else ""),
                            b.notes or "", b.invoice_path or "", appt_num_str,
                            b.created_at.strftime("%Y-%m-%d %H:%M:%S") if isinstance(b.created_at, datetime) else str(b.created_at),
                            b.updated_at.strftime("%Y-%m-%d %H:%M:%S") if isinstance(b.updated_at, datetime) else str(b.updated_at),
                            b.source,
                            item.id,
                            item.description,
                            item.quantity,
                            f"₹{item.unit_price:.2f}",
                            item.gst_rate,
                            f"₹{item.gst_amount:.2f}" if getattr(item, "gst_amount", None) is not None else "",
                            f"₹{item.line_total:.2f}" if getattr(item, "line_total", None) is not None else "",
                            item.item_type
                        ]
                        ws.append(row)
                        
            stream = BytesIO()
            wb.save(stream)
            stream.seek(0)
            return stream, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            
        elif format_type == "pdf":
            from jinja2 import Environment, FileSystemLoader
            from app.utils.pdf_generator import html_to_pdf
            from app.utils.helpers import utc_now
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from xhtml2pdf import default
            import os
            
            font_path = os.path.abspath("app/static/fonts/DejaVuSans.ttf")
            if "DejaVuSans" not in pdfmetrics.getRegisteredFontNames() and os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont("DejaVuSans", font_path))
                
            default.DEFAULT_FONT["dejavusans"] = "DejaVuSans"
            default.DEFAULT_FONT["dejavusans-bold"] = "DejaVuSans"
            default.DEFAULT_FONT["dejavusans-oblique"] = "DejaVuSans"
            default.DEFAULT_FONT["dejavusans-boldoblique"] = "DejaVuSans"
            
            env = Environment(loader=FileSystemLoader("app/templates"))
            template = env.get_template("billings_export_template.html")
            
            formatted_billings = []
            for b in ordered_items:
                p_name = patients_map.get(b.patient_id, "")
                
                formatted_billings.append({
                    "id": b.id,
                    "patient_id": b.patient_id,
                    "patient_name": p_name,
                    "bill_number": b.bill_number,
                    "subtotal": b.subtotal,
                    "discount_percent": b.discount_percent,
                    "discount_amount": b.discount_amount,
                    "gst_rate": b.gst_rate,
                    "gst_amount": b.gst_amount,
                    "tax_amount": b.tax_amount,
                    "total_amount": b.total_amount,
                    "paid_amount": b.paid_amount,
                    "balance_amount": b.balance_amount,
                    "status": b.status,
                    "due_date": b.due_date.strftime("%Y-%m-%d") if isinstance(b.due_date, datetime) else (str(b.due_date) if b.due_date else ""),
                    "notes": b.notes,
                    "appointment_id": b.appointment_id,
                    "source": b.source,
                    "created_at": b.created_at.strftime("%Y-%m-%d %H:%M:%S") if isinstance(b.created_at, datetime) else str(b.created_at),
                    "bill_items": [
                        {
                            "id": item.id,
                            "description": item.description,
                            "quantity": item.quantity,
                            "unit_price": item.unit_price,
                            "gst_rate": item.gst_rate,
                            "gst_amount": getattr(item, "gst_amount", 0.0),
                            "line_total": getattr(item, "line_total", 0.0),
                            "item_type": item.item_type
                        }
                        for item in b.items
                    ]
                })
                
            html_content = template.render(
                billings=formatted_billings,
                generated_at=utc_now().strftime("%Y-%m-%d %H:%M:%S")
            )
            
            pdf_data = html_to_pdf(html_content)
            return pdf_data, "application/pdf"
        else:
            raise BadRequestException("Invalid format specified for export")


class InsuranceService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.repo = InsuranceRepository(db)
        self.claim_repo = InsuranceClaimRepository(db)
        self.audit_repo = AuditRepository(db)

    async def create_insurance(self, data: InsuranceCreate, user_id: int) -> InsuranceResponse:
        insurance = Insurance(**data.model_dump())
        insurance = await self.repo.create(insurance)
        await self.audit_repo.create("create", "billing_insurance", user_id=user_id, resource_id=str(insurance.id))
        return InsuranceResponse.model_validate(insurance)

    async def submit_claim(self, data: InsuranceClaimCreate, user_id: int) -> InsuranceClaimResponse:
        claim = InsuranceClaim(
            claim_number=generate_claim_number(),
            submitted_at=utc_now(),
            **data.model_dump(),
        )
        claim = await self.claim_repo.create(claim)
        await self.audit_repo.create("create", "billing_claim", user_id=user_id, resource_id=str(claim.id))

        from app.services.transaction_history_service import TransactionHistoryService
        await TransactionHistoryService(self.db).create_event(
            event_type="INSURANCE_CLAIM",
            reference_no=claim.claim_number,
            description=f"Insurance Claim Submitted: {claim.notes or ''}",
            amount=claim.claimed_amount,
            source_module="insurance",
            source_id=claim.id,
            status="completed",
            user_id=user_id
        )

        return InsuranceClaimResponse.model_validate(claim)
