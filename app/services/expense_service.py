from io import BytesIO
from datetime import date
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import BadRequestException, NotFoundException, ConflictException
from app.models.expense_model import ExpenseCategory, Expense, VendorPayment
from app.repositories.audit_repository import AuditRepository
from app.repositories.expense_repository import (
    ExpenseCategoryRepository,
    ExpenseRepository,
    VendorPaymentRepository
)
from app.repositories.vendor_repository import VendorRepository
from app.schemas.expense_schema import (
    ExpenseCategoryCreate, ExpenseCategoryUpdate, ExpenseCategoryResponse,
    ExpenseCreate, ExpenseUpdate, ExpenseResponse, ExpenseQuery,
    VendorPaymentCreate, VendorPaymentUpdate, VendorPaymentResponse,
    ExpenseSummaryResponse
)
from app.schemas.vendor_schema import VendorResponse
from app.utils.pagination import build_paginated_result


class ExpenseService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.category_repo = ExpenseCategoryRepository(db)
        self.vendor_repo = VendorRepository(db)
        self.expense_repo = ExpenseRepository(db)
        self.payment_repo = VendorPaymentRepository(db)
        self.audit_repo = AuditRepository(db)

    # --- Expense Category Services ---
    async def create_category(self, data: ExpenseCategoryCreate, user_id: int) -> ExpenseCategoryResponse:
        existing = await self.category_repo.get_by_name(data.name)
        if existing:
            raise ConflictException(f"Expense category with name '{data.name}' already exists")

        category = ExpenseCategory(**data.model_dump())
        category = await self.category_repo.create(category)
        await self.audit_repo.create("create", "expense_category", user_id=user_id, resource_id=str(category.id))
        return ExpenseCategoryResponse.model_validate(category)

    async def list_categories(self, page: int = 1, size: int = 20):
        skip = (page - 1) * size
        categories = await self.category_repo.list_all(skip=skip, limit=size)
        total = await self.category_repo.count_all()
        return build_paginated_result(
            [ExpenseCategoryResponse.model_validate(c) for c in categories], total, page, size
        )

    async def get_category(self, category_id: int) -> ExpenseCategoryResponse:
        category = await self.category_repo.get_by_id(category_id)
        if not category:
            raise NotFoundException(f"Expense category with ID {category_id} not found")
        return ExpenseCategoryResponse.model_validate(category)

    async def update_category(self, category_id: int, data: ExpenseCategoryUpdate, user_id: int) -> ExpenseCategoryResponse:
        category = await self.category_repo.get_by_id(category_id)
        if not category:
            raise NotFoundException(f"Expense category with ID {category_id} not found")

        if data.name:
            existing = await self.category_repo.get_by_name(data.name)
            if existing and existing.id != category_id:
                raise ConflictException(f"Expense category with name '{data.name}' already exists")

        for key, value in data.model_dump(exclude_unset=True).items():
            setattr(category, key, value)

        category = await self.category_repo.update(category)
        await self.audit_repo.create("update", "expense_category", user_id=user_id, resource_id=str(category.id))
        return ExpenseCategoryResponse.model_validate(category)

    async def delete_category(self, category_id: int, user_id: int) -> None:
        category = await self.category_repo.get_by_id(category_id)
        if not category:
            raise NotFoundException(f"Expense category with ID {category_id} not found")

        expense_count = await self.expense_repo.count_all(category_id=category_id)
        if expense_count > 0:
            raise BadRequestException("Cannot delete category as it is linked to one or more expenses")

        await self.category_repo.soft_delete(category)
        await self.audit_repo.create("delete", "expense_category", user_id=user_id, resource_id=str(category.id))

    # --- Vendor Services Removed (Centralized in VendorService) ---

    # --- Expense Services ---
    async def create_expense(self, data: ExpenseCreate, user_id: int) -> ExpenseResponse:
        # Validate Category
        category = await self.category_repo.get_by_id(data.category_id)
        if not category:
            raise NotFoundException(f"Expense category with ID {data.category_id} not found")

        # Validate Vendor
        if data.vendor_id is not None:
            vendor = await self.vendor_repo.get_by_id(data.vendor_id)
            if not vendor:
                raise NotFoundException(f"Vendor with ID {data.vendor_id} not found")

        expense = Expense(**data.model_dump())
        expense = await self.expense_repo.create(expense)
        await self.audit_repo.create("create", "expense", user_id=user_id, resource_id=str(expense.id))

        from app.services.transaction_history_service import TransactionHistoryService
        await TransactionHistoryService(self.db).create_event(
            event_type="EXPENSE_RECORDED",
            reference_no=f"EXP-{expense.id}",
            description=f"Expense Recorded: {expense.description or ''}",
            amount=expense.amount,
            source_module="expenses",
            source_id=expense.id,
            status="completed",
            user_id=user_id
        )

        return ExpenseResponse.model_validate(expense)

    async def list_expenses(self, query: ExpenseQuery):
        # Resolve pharmacy category ID
        from app.models.expense_model import ExpenseCategory
        from sqlalchemy import select, func, or_
        from sqlalchemy.orm import selectinload
        from app.utils.helpers import utc_now

        cat_stmt = select(ExpenseCategory.id, ExpenseCategory.name).where(
            func.lower(ExpenseCategory.name) == "pharmacy"
        )
        cat_res = await self.db.execute(cat_stmt)
        cat_row = cat_res.first()
        if cat_row:
            pharm_cat_id = cat_row[0]
            pharm_cat_name = cat_row[1]
        else:
            pharm_cat_id = 999
            pharm_cat_name = "Pharmacy"

        # Determine whether to query general expenses
        query_general = True
        if query.category_id is not None and query.category_id != pharm_cat_id:
            query_general = True
        elif query.category_id == pharm_cat_id:
            query_general = False

        # Determine whether to query pharmacy purchases
        query_pharmacy = True
        if query.category_id is not None and query.category_id == pharm_cat_id:
            query_pharmacy = True
        elif query.category_id is not None and query.category_id != pharm_cat_id:
            query_pharmacy = False

        expenses_list = []
        if query_general:
            # Fetch all matching general expenses up to a reasonable limit for combining in memory
            expenses_list = await self.expense_repo.list_all(
                skip=0,
                limit=100000,
                sort_by=query.sort_by,
                sort_order=query.sort_order,
                category_id=query.category_id,
                vendor_id=query.vendor_id,
                status=query.status,
                start_date=query.start_date,
                end_date=query.end_date,
                description=query.description
            )

        purchases_list = []
        if query_pharmacy:
            from app.models.pharmacy_model import Purchase, Supplier
            from datetime import datetime

            purchase_stmt = select(Purchase).options(
                selectinload(Purchase.supplier)
            )
            if query.vendor_id is not None:
                purchase_stmt = purchase_stmt.where(Purchase.supplier_id == query.vendor_id)
            if query.status:
                p_status = query.status.lower().strip()
                if p_status == "paid":
                    purchase_stmt = purchase_stmt.where(func.lower(Purchase.status) == "received")
                elif p_status == "pending":
                    purchase_stmt = purchase_stmt.where(func.lower(Purchase.status) == "pending")
                else:
                    purchase_stmt = purchase_stmt.where(func.lower(Purchase.status) == p_status)
            if query.start_date:
                start_dt = datetime.combine(query.start_date, datetime.min.time())
                purchase_stmt = purchase_stmt.where(Purchase.ordered_at >= start_dt)
            if query.end_date:
                end_dt = datetime.combine(query.end_date, datetime.max.time())
                purchase_stmt = purchase_stmt.where(Purchase.ordered_at <= end_dt)
            if query.description:
                pattern = f"%{query.description.lower().strip()}%"
                purchase_stmt = purchase_stmt.where(
                    or_(
                        func.lower(Purchase.purchase_number).like(pattern),
                        func.lower(Purchase.notes).like(pattern)
                    )
                )

            purch_res = await self.db.execute(purchase_stmt)
            purchases_list = list(purch_res.scalars().unique().all())

        all_items = []
        for e in expenses_list:
            resp = ExpenseResponse.model_validate(e)
            resp.source = "expense"
            all_items.append(resp)

        from app.schemas.expense_schema import ExpenseCategoryResponse
        from app.schemas.vendor_schema import VendorResponse
        for p in purchases_list:
            vendor_data = None
            if p.supplier:
                vendor_data = VendorResponse(
                    id=p.supplier.id,
                    name=p.supplier.name,
                    vendor_type="supplier",
                    contact_person=p.supplier.contact_person,
                    phone=p.supplier.phone,
                    email=p.supplier.email,
                    address=p.supplier.address,
                    is_active=p.supplier.is_active if p.supplier.is_active is not None else True,
                    created_at=p.supplier.created_at or p.created_at or utc_now(),
                    updated_at=p.supplier.updated_at or p.created_at or utc_now()
                )

            cat_data = ExpenseCategoryResponse(
                id=pharm_cat_id,
                name=pharm_cat_name,
                description="Pharmacy Inventory and Supplies Purchases",
                is_active=True,
                created_at=p.created_at or utc_now(),
                updated_at=p.created_at or utc_now()
            )

            status_val = p.status.capitalize() if p.status else "Pending"
            if status_val == "Received":
                status_val = "Paid"

            desc_val = f"Pharmacy Purchase {p.purchase_number or ''}"
            if p.notes:
                desc_val += f". Notes: {p.notes}"

            resp = ExpenseResponse(
                id=p.id,
                category_id=pharm_cat_id,
                vendor_id=p.supplier_id,
                amount=p.total_amount or 0.0,
                description=desc_val,
                expense_date=p.ordered_at.date() if p.ordered_at else (p.created_at.date() if p.created_at else utc_now().date()),
                status=status_val,
                category=cat_data,
                vendor=vendor_data,
                created_at=p.created_at or utc_now(),
                updated_at=p.created_at or utc_now(),
                source="pharmacy"
            )
            all_items.append(resp)

        # Sort all items
        reverse = (query.sort_order == "desc")
        def get_sort_key(item: ExpenseResponse):
            if query.sort_by == "category":
                return item.category.name.lower() if (item.category and item.category.name) else ""
            elif query.sort_by == "vendor":
                return item.vendor.name.lower() if (item.vendor and item.vendor.name) else ""
            elif query.sort_by == "vendor_id":
                return item.vendor_id or 0
            elif query.sort_by == "description":
                return item.description.lower() if item.description else ""
            elif query.sort_by == "status":
                return item.status.lower() if item.status else ""
            
            val = getattr(item, query.sort_by, None)
            if val is None:
                if query.sort_by == "created_at":
                    val = item.created_at
                elif query.sort_by == "expense_date":
                    val = item.expense_date
                else:
                    val = item.created_at
            return val

        all_items.sort(key=get_sort_key, reverse=reverse)

        # Paginate
        skip = (query.page - 1) * query.size
        total = len(all_items)
        paginated = all_items[skip : skip + query.size]

        return build_paginated_result(paginated, total, query.page, query.size)


    async def get_expense(self, expense_id: int) -> ExpenseResponse:
        expense = await self.expense_repo.get_by_id(expense_id)
        if expense:
            resp = ExpenseResponse.model_validate(expense)
            resp.source = "expense"
            return resp

        # Check in pharmacy purchases
        from app.models.pharmacy_model import Purchase
        from sqlalchemy.orm import selectinload
        from sqlalchemy import select
        
        stmt = select(Purchase).where(Purchase.id == expense_id).options(selectinload(Purchase.supplier))
        res = await self.db.execute(stmt)
        p = res.scalar_one_or_none()
        if p:
            # Resolve pharmacy category
            from app.models.expense_model import ExpenseCategory
            from sqlalchemy import func
            cat_stmt = select(ExpenseCategory.id, ExpenseCategory.name).where(
                func.lower(ExpenseCategory.name) == "pharmacy"
            )
            cat_res = await self.db.execute(cat_stmt)
            cat_row = cat_res.first()
            if cat_row:
                pharm_cat_id = cat_row[0]
                pharm_cat_name = cat_row[1]
            else:
                pharm_cat_id = 999
                pharm_cat_name = "Pharmacy"

            from app.schemas.expense_schema import ExpenseCategoryResponse
            from app.schemas.vendor_schema import VendorResponse
            from app.utils.helpers import utc_now
            
            vendor_data = None
            if p.supplier:
                vendor_data = VendorResponse(
                    id=p.supplier.id,
                    name=p.supplier.name,
                    vendor_type="supplier",
                    contact_person=p.supplier.contact_person,
                    phone=p.supplier.phone,
                    email=p.supplier.email,
                    address=p.supplier.address,
                    is_active=p.supplier.is_active if p.supplier.is_active is not None else True,
                    created_at=p.supplier.created_at or p.created_at or utc_now(),
                    updated_at=p.supplier.updated_at or p.created_at or utc_now()
                )

            cat_data = ExpenseCategoryResponse(
                id=pharm_cat_id,
                name=pharm_cat_name,
                description="Pharmacy Inventory and Supplies Purchases",
                is_active=True,
                created_at=p.created_at or utc_now(),
                updated_at=p.created_at or utc_now()
            )

            status_val = p.status.capitalize() if p.status else "Pending"
            if status_val == "Received":
                status_val = "Paid"

            desc_val = f"Pharmacy Purchase {p.purchase_number or ''}"
            if p.notes:
                desc_val += f". Notes: {p.notes}"

            return ExpenseResponse(
                id=p.id,
                category_id=pharm_cat_id,
                vendor_id=p.supplier_id,
                amount=p.total_amount or 0.0,
                description=desc_val,
                expense_date=p.ordered_at.date() if p.ordered_at else (p.created_at.date() if p.created_at else utc_now().date()),
                status=status_val,
                category=cat_data,
                vendor=vendor_data,
                created_at=p.created_at or utc_now(),
                updated_at=p.created_at or utc_now(),
                source="pharmacy"
            )

        raise NotFoundException(f"Expense with ID {expense_id} not found")

    async def update_expense(self, expense_id: int, data: ExpenseUpdate, user_id: int) -> ExpenseResponse:
        expense = await self.expense_repo.get_by_id(expense_id)
        if not expense:
            # Check if it is a pharmacy purchase to raise a helpful message
            from app.models.pharmacy_model import Purchase
            from sqlalchemy import select
            stmt = select(Purchase).where(Purchase.id == expense_id)
            res = await self.db.execute(stmt)
            p = res.scalar_one_or_none()
            if p:
                raise BadRequestException("Cannot update pharmacy purchase details from the expenses module. Please manage it under the Pharmacy module.")
            raise NotFoundException(f"Expense with ID {expense_id} not found")

        if data.category_id is not None:
            category = await self.category_repo.get_by_id(data.category_id)
            if not category:
                raise NotFoundException(f"Expense category with ID {data.category_id} not found")

        if data.vendor_id is not None:
            vendor = await self.vendor_repo.get_by_id(data.vendor_id)
            if not vendor:
                raise NotFoundException(f"Vendor with ID {data.vendor_id} not found")

        for key, value in data.model_dump(exclude_unset=True).items():
            setattr(expense, key, value)

        expense = await self.expense_repo.update(expense)
        if "status" not in data.model_fields_set:
            await self._update_expense_status(expense.id)
        # Re-fetch with loaded relationships
        expense = await self.expense_repo.get_by_id(expense.id)
        await self.audit_repo.create("update", "expense", user_id=user_id, resource_id=str(expense.id))

        # Sync transaction history
        from sqlalchemy import select
        from app.models.transaction_history_model import TransactionHistory
        stmt = select(TransactionHistory).where(
            TransactionHistory.source_module == "expenses",
            TransactionHistory.source_id == expense.id,
            TransactionHistory.is_deleted.is_(False)
        )
        tx_result = await self.db.execute(stmt)
        tx_history = tx_result.scalar_one_or_none()

        from datetime import datetime, time
        event_datetime = datetime.combine(expense.expense_date, time.min)

        if tx_history:
            tx_history.amount = expense.amount
            tx_history.description = f"Expense Recorded: {expense.description or ''}"
            tx_history.event_date = event_datetime
            tx_history.status = expense.status
            from app.repositories.transaction_history_repository import TransactionHistoryRepository
            await TransactionHistoryRepository(self.db).update(tx_history)
        else:
            from app.services.transaction_history_service import TransactionHistoryService
            await TransactionHistoryService(self.db).create_event(
                event_type="EXPENSE_RECORDED",
                reference_no=f"EXP-{expense.id}",
                description=f"Expense Recorded: {expense.description or ''}",
                amount=expense.amount,
                source_module="expenses",
                source_id=expense.id,
                status=expense.status,
                event_date=event_datetime,
                user_id=user_id
            )

        resp = ExpenseResponse.model_validate(expense)
        resp.source = "expense"
        return resp

    async def delete_expense(self, expense_id: int, user_id: int) -> None:
        expense = await self.expense_repo.get_by_id(expense_id)
        if expense:
            await self.expense_repo.soft_delete(expense)
            await self.audit_repo.create("delete", "expense", user_id=user_id, resource_id=str(expense.id))
            return

        # Check in pharmacy purchases
        from app.models.pharmacy_model import Purchase
        from sqlalchemy import select
        stmt = select(Purchase).where(Purchase.id == expense_id)
        res = await self.db.execute(stmt)
        p = res.scalar_one_or_none()
        if p:
            # We hard delete the purchase since there is no soft delete column on purchases
            await self.db.delete(p)
            await self.db.flush()
            await self.audit_repo.create("delete", "pharmacy_purchase", user_id=user_id, resource_id=str(p.id))
            return

        raise NotFoundException(f"Expense with ID {expense_id} not found")

    async def get_expense_summary(self, start_date: date | None = None, end_date: date | None = None) -> ExpenseSummaryResponse:
        summary = await self.expense_repo.get_summary(start_date, end_date)

        # Now query and aggregate pharmacy purchases (expenses)
        from app.models.pharmacy_model import Purchase, Supplier
        from app.models.expense_model import ExpenseCategory
        from sqlalchemy import select, func
        from datetime import datetime

        purchase_filter = []
        if start_date is not None:
            start_dt = datetime.combine(start_date, datetime.min.time())
            purchase_filter.append(Purchase.ordered_at >= start_dt)
        if end_date is not None:
            end_dt = datetime.combine(end_date, datetime.max.time())
            purchase_filter.append(Purchase.ordered_at <= end_dt)

        purch_stmt = select(
            func.coalesce(func.sum(Purchase.total_amount), 0.0),
            func.count(Purchase.id)
        ).where(*purchase_filter)
        purch_res = await self.db.execute(purch_stmt)
        purch_row = purch_res.first()
        purch_amount = float(purch_row[0] if purch_row else 0.0)
        purch_count = int(purch_row[1] if purch_row else 0)

        if purch_count > 0:
            summary["total_amount"] = round(summary["total_amount"] + purch_amount, 2)
            summary["total_count"] = summary["total_count"] + purch_count

            # 1. Category aggregate
            cat_stmt = select(ExpenseCategory.id, ExpenseCategory.name).where(
                func.lower(ExpenseCategory.name) == "pharmacy"
            )
            cat_res = await self.db.execute(cat_stmt)
            cat_row = cat_res.first()
            if cat_row:
                pharm_cat_id = cat_row[0]
                pharm_cat_name = cat_row[1]
            else:
                pharm_cat_id = 999
                pharm_cat_name = "Pharmacy"

            by_category = list(summary.get("by_category", []))
            found_cat = False
            for cat in by_category:
                if cat.get("category_id") == pharm_cat_id or cat.get("name").lower() == "pharmacy":
                    cat["total_amount"] = round(cat["total_amount"] + purch_amount, 2)
                    cat["count"] = cat["count"] + purch_count
                    found_cat = True
                    break
            if not found_cat:
                by_category.append({
                    "category_id": pharm_cat_id,
                    "name": pharm_cat_name,
                    "total_amount": purch_amount,
                    "count": purch_count
                })
            summary["by_category"] = by_category

            # 2. Vendor aggregate (suppliers mapped as vendors)
            by_vendor = list(summary.get("by_vendor", []))
            supplier_stmt = select(
                Purchase.supplier_id.label("supplier_id"),
                Supplier.name.label("name"),
                func.coalesce(func.sum(Purchase.total_amount), 0.0).label("total_amount"),
                func.count(Purchase.id).label("count")
            ).join(
                Supplier, Purchase.supplier_id == Supplier.id
            ).where(
                *purchase_filter
            ).group_by(
                Purchase.supplier_id, Supplier.name
            )
            supplier_res = await self.db.execute(supplier_stmt)
            for r in supplier_res.all():
                by_vendor.append({
                    "vendor_id": r.supplier_id,
                    "name": r.name,
                    "total_amount": float(r.total_amount),
                    "count": int(r.count)
                })
            summary["by_vendor"] = by_vendor

            # 3. Status aggregate (and update paid/pending totals)
            by_status = list(summary.get("by_status", []))
            status_stmt = select(
                Purchase.status.label("status"),
                func.coalesce(func.sum(Purchase.total_amount), 0.0).label("total_amount"),
                func.count(Purchase.id).label("count")
            ).where(
                *purchase_filter
            ).group_by(
                Purchase.status
            )
            status_res = await self.db.execute(status_stmt)
            for r in status_res.all():
                raw_status = (r.status or "").lower().strip()
                status_key = "Paid" if raw_status == "received" else ("Pending" if raw_status == "pending" else (r.status or "").capitalize())
                
                amt = float(r.total_amount)
                cnt = int(r.count)

                if status_key.lower() == "paid":
                    summary["paid_amount"] = round(summary.get("paid_amount", 0.0) + amt, 2)
                    summary["paid_count"] = summary.get("paid_count", 0) + cnt
                elif status_key.lower() == "pending":
                    summary["pending_amount"] = round(summary.get("pending_amount", 0.0) + amt, 2)
                    summary["pending_count"] = summary.get("pending_count", 0) + cnt

                found_status = False
                for item in by_status:
                    if (item.get("status") or "").capitalize() == status_key:
                        item["total_amount"] = round(item["total_amount"] + amt, 2)
                        item["count"] = item["count"] + cnt
                        found_status = True
                        break
                if not found_status:
                    by_status.append({
                        "status": status_key,
                        "total_amount": amt,
                        "count": cnt
                    })
            summary["by_status"] = by_status

            # Count distinct suppliers from purchases in addition to general vendors
            distinct_supplier_stmt = select(
                func.count(func.distinct(Purchase.supplier_id))
            ).where(*purchase_filter, Purchase.supplier_id.isnot(None))
            distinct_supp_res = await self.db.execute(distinct_supplier_stmt)
            pharm_vendors_count = int(distinct_supp_res.scalar() or 0)
            summary["total_vendors"] = summary.get("total_vendors", 0) + pharm_vendors_count

            # 4. Monthly summary
            monthly_summary = list(summary.get("monthly_summary", []))
            monthly_stmt = select(
                func.year(Purchase.ordered_at).label("year"),
                func.month(Purchase.ordered_at).label("month"),
                func.coalesce(func.sum(Purchase.total_amount), 0.0).label("total_amount"),
                func.count(Purchase.id).label("count")
            ).where(
                *purchase_filter
            ).group_by(
                func.year(Purchase.ordered_at),
                func.month(Purchase.ordered_at)
            )
            monthly_res = await self.db.execute(monthly_stmt)
            for r in monthly_res.all():
                if r.year is None or r.month is None:
                    continue
                year_val = int(r.year)
                month_val = int(r.month)
                found_month = False
                for item in monthly_summary:
                    if item.get("year") == year_val and item.get("month") == month_val:
                        item["total_amount"] = round(item["total_amount"] + float(r.total_amount), 2)
                        item["count"] = item["count"] + int(r.count)
                        found_month = True
                        break
                if not found_month:
                    monthly_summary.append({
                        "year": year_val,
                        "month": month_val,
                        "total_amount": float(r.total_amount),
                        "count": int(r.count)
                    })
            monthly_summary.sort(key=lambda x: (x.get("year"), x.get("month")), reverse=True)
            summary["monthly_summary"] = monthly_summary

        return ExpenseSummaryResponse.model_validate(summary)

    # --- Vendor Payment Services ---
    async def create_payment(self, data: VendorPaymentCreate, user_id: int) -> VendorPaymentResponse:
        # Validate vendor
        vendor = await self.vendor_repo.get_by_id(data.vendor_id)
        if not vendor:
            raise NotFoundException(f"Vendor with ID {data.vendor_id} not found")

        # Validate expense
        expense = await self.expense_repo.get_by_id(data.expense_id)
        if not expense:
            raise NotFoundException(f"Expense with ID {data.expense_id} not found")

        payment = VendorPayment(**data.model_dump())
        payment = await self.payment_repo.create(payment)

        # Update expense status
        await self._update_expense_status(expense.id)

        await self.audit_repo.create("create", "vendor_payment", user_id=user_id, resource_id=str(payment.id))
        return VendorPaymentResponse.model_validate(payment)

    async def list_payments(self, page: int = 1, size: int = 20, vendor_id: int | None = None, expense_id: int | None = None):
        skip = (page - 1) * size
        payments = await self.payment_repo.list_all(skip=skip, limit=size, vendor_id=vendor_id, expense_id=expense_id)
        total = await self.payment_repo.count_all(vendor_id=vendor_id, expense_id=expense_id)
        return build_paginated_result(
            [VendorPaymentResponse.model_validate(p) for p in payments], total, page, size
        )

    async def get_payment(self, payment_id: int) -> VendorPaymentResponse:
        payment = await self.payment_repo.get_by_id(payment_id)
        if not payment:
            raise NotFoundException(f"Vendor payment with ID {payment_id} not found")
        return VendorPaymentResponse.model_validate(payment)

    async def update_payment(self, payment_id: int, data: VendorPaymentUpdate, user_id: int) -> VendorPaymentResponse:
        payment = await self.payment_repo.get_by_id(payment_id)
        if not payment:
            raise NotFoundException(f"Vendor payment with ID {payment_id} not found")

        old_expense_id = payment.expense_id

        if data.vendor_id is not None:
            vendor = await self.vendor_repo.get_by_id(data.vendor_id)
            if not vendor:
                raise NotFoundException(f"Vendor with ID {data.vendor_id} not found")

        if data.expense_id is not None:
            expense = await self.expense_repo.get_by_id(data.expense_id)
            if not expense:
                raise NotFoundException(f"Expense with ID {data.expense_id} not found")

        for key, value in data.model_dump(exclude_unset=True).items():
            setattr(payment, key, value)

        payment = await self.payment_repo.update(payment)

        # Update old and new expense statuses
        await self._update_expense_status(old_expense_id)
        if payment.expense_id != old_expense_id:
            await self._update_expense_status(payment.expense_id)

        await self.audit_repo.create("update", "vendor_payment", user_id=user_id, resource_id=str(payment.id))
        return VendorPaymentResponse.model_validate(payment)

    async def delete_payment(self, payment_id: int, user_id: int) -> None:
        payment = await self.payment_repo.get_by_id(payment_id)
        if not payment:
            raise NotFoundException(f"Vendor payment with ID {payment_id} not found")

        expense_id = payment.expense_id
        await self.payment_repo.soft_delete(payment)

        # Recalculate expense status
        await self._update_expense_status(expense_id)

        await self.audit_repo.create("delete", "vendor_payment", user_id=user_id, resource_id=str(payment.id))

    # --- Private Helper Methods ---
    async def _update_expense_status(self, expense_id: int) -> None:
        expense = await self.expense_repo.get_by_id(expense_id)
        if not expense:
            return

        payments = await self.payment_repo.get_payments_by_expense(expense_id)
        total_paid = sum(p.amount for p in payments)

        new_status = "Paid" if total_paid >= expense.amount else "Pending"
        if expense.status != new_status:
            expense.status = new_status
            await self.expense_repo.update(expense)

    async def generate_expense_bulk_template(self) -> BytesIO:
        import openpyxl
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Expenses Bulk Import"
        
        headers = [
            "category_id", "vendor_id", "amount", "description", "expense_date", "status"
        ]
        ws.append(headers)
        
        # Add one valid sample row
        ws.append([
            1,
            2,
            1500.50,
            "Office utilities payment",
            "2026-08-15",
            "Paid"
        ])
        
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream

    async def import_expenses_from_excel(self, file, user_id: int) -> dict:
        from pydantic import ValidationError
        import openpyxl
        from datetime import date, datetime
        
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents))
        ws = wb.active
        
        # Read headers
        header_row = next(ws.iter_rows(max_row=1, values_only=True), None)
        if not header_row:
            raise BadRequestException("The uploaded file is empty or has no headers.")
            
        headers = [str(h).strip().lower() for h in header_row if h is not None]
        required_headers = {"category_id", "amount", "description", "expense_date"}
        if not required_headers.issubset(set(headers)):
            raise BadRequestException("Missing required headers in the upload template.")
            
        total_rows = 0
        created = 0
        failed = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None for cell in row):
                continue
                
            total_rows += 1
            row_dict = {}
            for header, val in zip(headers, row):
                if val is None or str(val).strip() == "":
                    row_dict[header] = None
                else:
                    row_dict[header] = val
                    
            try:
                category_id_raw = row_dict.get("category_id")
                if category_id_raw is None:
                    raise BadRequestException("category_id is required.")
                try:
                    category_id = int(float(category_id_raw))
                except ValueError:
                    raise BadRequestException(f"Invalid category_id value: {category_id_raw}")
                    
                vendor_id = None
                vendor_id_raw = row_dict.get("vendor_id")
                if vendor_id_raw is not None:
                    try:
                        vendor_id = int(float(vendor_id_raw))
                    except ValueError:
                        raise BadRequestException(f"Invalid vendor_id value: {vendor_id_raw}")
                        
                amount_raw = row_dict.get("amount")
                if amount_raw is None:
                    raise BadRequestException("amount is required.")
                try:
                    amount = float(amount_raw)
                except ValueError:
                    raise BadRequestException(f"Invalid amount value: {amount_raw}")
                    
                desc_raw = row_dict.get("description")
                if not desc_raw:
                    raise BadRequestException("description is required.")
                description = str(desc_raw).strip()
                
                expense_date = None
                date_raw = row_dict.get("expense_date")
                if date_raw is None:
                    raise BadRequestException("expense_date is required.")
                if isinstance(date_raw, (datetime, date)):
                    expense_date = date_raw if isinstance(date_raw, date) else date_raw.date()
                else:
                    try:
                        expense_date = datetime.strptime(str(date_raw).strip(), "%Y-%m-%d").date()
                    except ValueError:
                        try:
                            expense_date = datetime.fromisoformat(str(date_raw).strip()).date()
                        except ValueError:
                            raise BadRequestException(f"Invalid expense_date format: {date_raw}. Use YYYY-MM-DD.")
                            
                status = "Pending"
                status_raw = row_dict.get("status")
                if status_raw is not None:
                    status = str(status_raw).strip()
                    
                # 2. Build ExpenseCreate schema
                expense_create = ExpenseCreate(
                    category_id=category_id,
                    vendor_id=vendor_id,
                    amount=amount,
                    description=description,
                    expense_date=expense_date,
                    status=status
                )
                
                # 3. Call service creation method
                await self.create_expense(expense_create, user_id)
                created += 1
                
            except ValidationError as e:
                failed += 1
                err_msg = "; ".join([f"{'.'.join(str(loc) for loc in error['loc'])}: {error['msg']}" for error in e.errors()])
                errors.append({
                    "row": row_idx,
                    "error": err_msg
                })
            except ConflictException as e:
                failed += 1
                errors.append({
                    "row": row_idx,
                    "error": str(e.detail)
                })
            except NotFoundException as e:
                failed += 1
                errors.append({
                    "row": row_idx,
                    "error": str(e.detail)
                })
            except BadRequestException as e:
                failed += 1
                errors.append({
                    "row": row_idx,
                    "error": str(e.detail)
                })
            except Exception as e:
                failed += 1
                errors.append({
                    "row": row_idx,
                    "error": str(e)
                })
                
        await self.db.flush()
        return {
            "total_rows": total_rows,
            "created": created,
            "failed": failed,
            "errors": errors
        }

    async def _get_all_combined_expenses(self) -> list[ExpenseResponse]:
        from app.models.expense_model import ExpenseCategory
        from sqlalchemy import select, func, or_
        from sqlalchemy.orm import selectinload
        from app.utils.helpers import utc_now

        # Resolve pharmacy category ID
        cat_stmt = select(ExpenseCategory.id, ExpenseCategory.name).where(
            func.lower(ExpenseCategory.name) == "pharmacy"
        )
        cat_res = await self.db.execute(cat_stmt)
        cat_row = cat_res.first()
        if cat_row:
            pharm_cat_id = cat_row[0]
            pharm_cat_name = cat_row[1]
        else:
            pharm_cat_id = 999
            pharm_cat_name = "Pharmacy"

        # Fetch all general expenses
        expenses_list = await self.expense_repo.list_all(
            skip=0,
            limit=1000000,
            sort_by="created_at",
            sort_order="desc"
        )

        # Fetch all pharmacy purchases
        from app.models.pharmacy_model import Purchase, Supplier
        purchase_stmt = select(Purchase).options(
            selectinload(Purchase.supplier)
        ).where(Purchase.is_deleted == False)
        purch_res = await self.db.execute(purchase_stmt)
        purchases_list = list(purch_res.scalars().unique().all())

        all_items = []
        for e in expenses_list:
            resp = ExpenseResponse.model_validate(e)
            resp.source = "expense"
            all_items.append(resp)

        from app.schemas.expense_schema import ExpenseCategoryResponse
        from app.schemas.vendor_schema import VendorResponse
        for p in purchases_list:
            vendor_data = None
            if p.supplier:
                vendor_data = VendorResponse(
                    id=p.supplier.id,
                    name=p.supplier.name,
                    vendor_type="supplier",
                    contact_person=p.supplier.contact_person,
                    phone=p.supplier.phone,
                    email=p.supplier.email,
                    address=p.supplier.address,
                    is_active=p.supplier.is_active if p.supplier.is_active is not None else True,
                    created_at=p.supplier.created_at or p.created_at or utc_now(),
                    updated_at=p.supplier.updated_at or p.created_at or utc_now()
                )

            cat_data = ExpenseCategoryResponse(
                id=pharm_cat_id,
                name=pharm_cat_name,
                description="Pharmacy Inventory and Supplies Purchases",
                is_active=True,
                created_at=p.created_at or utc_now(),
                updated_at=p.created_at or utc_now()
            )

            status_val = p.status.capitalize() if p.status else "Pending"
            if status_val == "Received":
                status_val = "Paid"

            desc_val = f"Pharmacy Purchase {p.purchase_number or ''}"
            if p.notes:
                desc_val += f". Notes: {p.notes}"

            resp = ExpenseResponse(
                id=p.id,
                category_id=pharm_cat_id,
                vendor_id=p.supplier_id,
                amount=p.total_amount or 0.0,
                description=desc_val,
                expense_date=p.ordered_at.date() if p.ordered_at else (p.created_at.date() if p.created_at else utc_now().date()),
                status=status_val,
                category=cat_data,
                vendor=vendor_data,
                created_at=p.created_at or utc_now(),
                updated_at=p.created_at or utc_now(),
                source="pharmacy"
            )
            all_items.append(resp)

        # Sort chronologically DESC by created_at or fallback expense_date
        def get_date_sort(item: ExpenseResponse):
            return item.created_at or utc_now()

        all_items.sort(key=get_date_sort, reverse=True)
        return all_items

    async def export_expenses(self, format_type: str) -> tuple[BytesIO | bytes, str]:
        from datetime import date, datetime
        
        expenses = await self._get_all_combined_expenses()
        
        if format_type == "excel":
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Expenses Export"
            
            headers = [
                "Sr. No.", "category_id", "vendor_id", "amount", "description",
                "expense_date", "status", "category_name", "vendor_name",
                "created_at", "updated_at", "source"
            ]
            ws.append(headers)
            
            for sr_no, exp in enumerate(expenses, start=1):
                cat_name = exp.category.name if exp.category else ""
                vendor_name = exp.vendor.name if exp.vendor else ""
                
                row = [
                    sr_no,
                    exp.category_id,
                    exp.vendor_id,
                    f"₹{exp.amount:.2f}",
                    exp.description or "",
                    exp.expense_date.strftime("%Y-%m-%d") if isinstance(exp.expense_date, (date, datetime)) else str(exp.expense_date),
                    exp.status,
                    cat_name,
                    vendor_name,
                    exp.created_at.strftime("%Y-%m-%d %H:%M:%S") if isinstance(exp.created_at, datetime) else str(exp.created_at),
                    exp.updated_at.strftime("%Y-%m-%d %H:%M:%S") if isinstance(exp.updated_at, datetime) else str(exp.updated_at),
                    exp.source
                ]
                ws.append(row)
                
            stream = BytesIO()
            wb.save(stream)
            stream.seek(0)
            return stream, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            
        elif format_type == "pdf":
            from jinja2 import Environment, FileSystemLoader
            from app.utils.pdf_generator import html_to_pdf
            from app.utils.helpers import utc_now
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from xhtml2pdf import default
            import os

            font_path = os.path.abspath("app/static/fonts/DejaVuSans.ttf")
            if "DejaVuSans" not in pdfmetrics.getRegisteredFontNames() and os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont("DejaVuSans", font_path))
            
            # Map dejavusans variants in xhtml2pdf fontList registry
            default.DEFAULT_FONT["dejavusans"] = "DejaVuSans"
            default.DEFAULT_FONT["dejavusans-bold"] = "DejaVuSans"
            default.DEFAULT_FONT["dejavusans-oblique"] = "DejaVuSans"
            default.DEFAULT_FONT["dejavusans-boldoblique"] = "DejaVuSans"
            
            env = Environment(loader=FileSystemLoader("app/templates"))
            template = env.get_template("expenses_export_template.html")
            
            formatted_expenses = []
            for exp in expenses:
                cat_name = exp.category.name if exp.category else "-"
                vendor_name = exp.vendor.name if exp.vendor else "-"
                
                exp_date_str = exp.expense_date.strftime("%Y-%m-%d") if isinstance(exp.expense_date, (date, datetime)) else str(exp.expense_date)
                created_str = exp.created_at.strftime("%Y-%m-%d %H:%M:%S") if isinstance(exp.created_at, datetime) else str(exp.created_at)
                updated_str = exp.updated_at.strftime("%Y-%m-%d %H:%M:%S") if isinstance(exp.updated_at, datetime) else str(exp.updated_at)
                
                formatted_expenses.append({
                    "id": exp.id,
                    "category_id": exp.category_id,
                    "vendor_id": exp.vendor_id if exp.vendor_id else "-",
                    "amount": f"₹{exp.amount:.2f}",
                    "description": exp.description or "-",
                    "expense_date": exp_date_str,
                    "status": exp.status,
                    "category_name": cat_name,
                    "vendor_name": vendor_name,
                    "created_at": created_str,
                    "updated_at": updated_str,
                    "source": exp.source
                })
                
            html = template.render(
                expenses=formatted_expenses,
                generated_at=utc_now().strftime("%Y-%m-%d %H:%M:%S")
            )
            
            pdf_bytes = html_to_pdf(html)
            return pdf_bytes, "application/pdf"
            
        else:
            raise BadRequestException("Invalid format specified for export")
