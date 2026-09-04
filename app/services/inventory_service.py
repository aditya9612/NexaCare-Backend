from sqlalchemy import select, func
from app.models.inventory_model import WarehouseStock
from fastapi import HTTPException
from io import BytesIO
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.constants import ReorderAlertStatus, StockTransactionType
from app.core.exceptions import BadRequestException, NotFoundException, ConflictException
from app.services.stock_movement_service import StockMovementService
from app.models.inventory_model import InventoryItem, ReorderAlert, StockTransaction, Warehouse
from app.models.vendor_model import Vendor
from app.repositories.audit_repository import AuditRepository
from app.repositories.department_repository import DepartmentRepository
from app.repositories.inventory_repository import (
    InventoryRepository,
    ReorderAlertRepository,
    StockTransactionRepository,
    WarehouseRepository,
)
from app.repositories.vendor_repository import VendorRepository
from app.schemas.inventory_schema import (
    ConsumptionReport,
    InventoryItemCreate,
    InventoryItemResponse,
    InventoryItemUpdate,
    ReorderAlertResponse,
    StockSummary,
    StockTransactionCreate,
    StockTransactionResponse,
    StockTransactionUpdate,
    # Vendor schemas removed (centralized)
    WarehouseCreate,
    WarehouseResponse,
    WarehouseUpdate,
    InventoryDashboardResponse,
)
from app.utils.helpers import generate_code, generate_stock_transaction_number, utc_now
from app.utils.pagination import build_paginated_result


class InventoryService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.item_repo = InventoryRepository(db)
        self.transaction_repo = StockTransactionRepository(db)
        self.vendor_repo = VendorRepository(db)
        self.warehouse_repo = WarehouseRepository(db)
        self.alert_repo = ReorderAlertRepository(db)
        self.audit_repo = AuditRepository(db)
        self.dept_repo = DepartmentRepository(db)

    async def _validate_department(self, department_id: int | None) -> None:
        if department_id is not None:
            dept = await self.dept_repo.get_by_id(department_id)
            if not dept:
                raise NotFoundException(f"Department with ID {department_id} not found")

    async def _validate_warehouse(self, warehouse_id: int | None) -> None:
        if warehouse_id is not None:
            wh = await self.warehouse_repo.get_by_id(warehouse_id)
            if not wh:
                raise NotFoundException(f"Warehouse with ID {warehouse_id} not found")

    async def _validate_vendor(self, vendor_id: int | None) -> None:
        if vendor_id is not None:
            vendor = await self.vendor_repo.get_by_id(vendor_id)
            if not vendor:
                raise NotFoundException(f"Vendor with ID {vendor_id} not found")

    async def create_item(self, data: InventoryItemCreate, user_id: int) -> InventoryItemResponse:
        await self._validate_department(data.department_id)
        await self._validate_warehouse(data.warehouse_id)
        await self._validate_vendor(data.vendor_id)

        # Check duplicate name case-insensitive ignoring leading/trailing spaces
        existing_name = await self.item_repo.get_by_name(data.name)
        if existing_name:
            raise ConflictException("Inventory item with this name already exists.")

        sku = data.sku
        if sku:
            existing = await self.item_repo.get_by_sku(sku)
            if existing:
                raise ConflictException("Inventory item with this SKU already exists")
        else:
            while True:
                sku = generate_code("SKU")
                existing = await self.item_repo.get_by_sku(sku)
                if not existing:
                    break

        barcode = data.barcode
        if barcode:
            existing_barcode = await self.item_repo.get_by_barcode(barcode)
            if existing_barcode:
                raise ConflictException("Inventory item with this barcode already exists")

        item = InventoryItem(sku=sku, **data.model_dump(exclude={"sku"}))
        item = await self.item_repo.create(item)
        await self._check_reorder_alert(item)
        await self.audit_repo.create("create", "inventory", user_id=user_id, resource_id=str(item.id))
        return InventoryItemResponse.model_validate(item)

    async def list_items(
        self, page: int = 1, size: int = 20, sort_by: str = "created_at",
        sort_order: str = "desc", category: str | None = None, warehouse_id: int | None = None,
    ):
        skip = (page - 1) * size
        items = await self.item_repo.list_all(
            skip=skip, limit=size, sort_by=sort_by, sort_order=sort_order,
            category=category, warehouse_id=warehouse_id,
        )
        total = await self.item_repo.count_all(category=category, warehouse_id=warehouse_id)
        return build_paginated_result([InventoryItemResponse.model_validate(i) for i in items], total, page, size)

    async def search_items(self, q: str, page: int = 1, size: int = 20):
        skip = (page - 1) * size
        items = await self.item_repo.search(q, skip=skip, limit=size)
        total = await self.item_repo.count_search(q)
        return build_paginated_result([InventoryItemResponse.model_validate(i) for i in items], total, page, size)

    async def get_item(self, item_id: int) -> InventoryItemResponse:
        item = await self.item_repo.get_by_id(item_id)
        if not item:
            raise NotFoundException("Inventory item not found")
        return InventoryItemResponse.model_validate(item)

    async def update_item(self, item_id: int, data: InventoryItemUpdate, user_id: int) -> InventoryItemResponse:
        item = await self.item_repo.get_by_id(item_id)
        if not item:
            raise NotFoundException("Inventory item not found")
        await self._validate_department(data.department_id)
        await self._validate_warehouse(data.warehouse_id)
        await self._validate_vendor(data.vendor_id)

        if data.barcode:
            existing_barcode = await self.item_repo.get_by_barcode(data.barcode)
            if existing_barcode and existing_barcode.id != item_id:
                raise ConflictException("Inventory item with this barcode already exists")

        for key, value in data.model_dump(exclude_unset=True).items():
            setattr(item, key, value)
        item = await self.item_repo.update(item)
        await self._check_reorder_alert(item)
        await self.audit_repo.create("update", "inventory", user_id=user_id, resource_id=str(item.id))
        return InventoryItemResponse.model_validate(item)

    async def delete_item(self, item_id: int, user_id: int) -> None:
        item = await self.item_repo.get_by_id(item_id)
        if not item:
            raise NotFoundException("Inventory item not found")
        await self.item_repo.soft_delete(item)
        await self.audit_repo.create("delete", "inventory", user_id=user_id, resource_id=str(item.id))

    async def _check_reorder_alert(self, item: InventoryItem) -> None:
        if item.quantity <= item.reorder_level:
            alert = ReorderAlert(
                item_id=item.id,
                current_quantity=item.quantity,
                reorder_level=item.reorder_level,
                status=ReorderAlertStatus.ACTIVE,
            )
            await self.alert_repo.create(alert)
        else:
            await self.alert_repo.resolve_for_item(item.id)

    async def create_transaction(self, data: StockTransactionCreate, user_id: int) -> StockTransactionResponse:
        item = await self.item_repo.get_by_id(data.item_id)
        if not item:
            raise NotFoundException("Inventory item not found")

        # Use StockMovementService for all physical mutations
        direction = "IN"
        if data.transaction_type in (StockTransactionType.OUTWARD, StockTransactionType.CONSUMPTION, StockTransactionType.TRANSFER):
            direction = "OUT"

        quantity = data.quantity
        if data.transaction_type == StockTransactionType.ADJUSTMENT:
            # We must reject direct API usage of adjustment here, it's safer to only allow explicit references
            raise BadRequestException("Direct adjustment transactions are not supported. Use physical IN/OUT.")

        transaction = await StockMovementService.create_movement(
            db=self.db,
            item_id=data.item_id,
            warehouse_id=data.warehouse_id,
            transaction_type=data.transaction_type,
            direction=direction,
            quantity=abs(quantity),
            batch_id=data.batch_id,
            unit_cost=data.unit_cost,
            reference_type=data.reference_type,
            reference_id=data.reference_id,
            notes=data.notes,
            performed_by=user_id
        )
        return self._to_transaction_response(transaction)

    def _to_transaction_response(self, transaction: StockTransaction) -> StockTransactionResponse:
        data = StockTransactionResponse.model_validate(transaction)
        data.type = transaction.transaction_type
        if hasattr(transaction, "item") and transaction.item:
            data.item_name = transaction.item.name
        if hasattr(transaction, "warehouse") and transaction.warehouse:
            data.warehouse_name = transaction.warehouse.name
        data.total_value = round(abs(transaction.quantity) * transaction.unit_cost, 2)
        return data

    async def list_transactions(
        self, page: int = 1, size: int = 20, item_id: int | None = None, transaction_type: str | None = None
    ):
        skip = (page - 1) * size
        items = await self.transaction_repo.list_all(
            skip=skip, limit=size, item_id=item_id, transaction_type=transaction_type
        )
        total = await self.transaction_repo.count_all(item_id=item_id, transaction_type=transaction_type)
        return build_paginated_result(
            [self._to_transaction_response(t) for t in items], total, page, size
        )

    async def get_transaction(self, transaction_id: int) -> StockTransactionResponse:
        transaction = await self.transaction_repo.get_by_id(transaction_id)
        if not transaction:
            raise NotFoundException("Stock transaction not found")
        return self._to_transaction_response(transaction)

    async def update_transaction(self, transaction_id: int, data: StockTransactionUpdate, user_id: int) -> StockTransactionResponse:
        raise BadRequestException("Stock transactions are immutable and cannot be updated.")

    async def delete_stock_transaction(self, transaction_id: int, user_id: int) -> None:
        raise BadRequestException("Stock transactions are immutable and cannot be deleted.")

    async def get_transaction(self, transaction_id: int) -> StockTransactionResponse:
        transaction = await self.transaction_repo.get_by_id(transaction_id)
        if not transaction:
            raise NotFoundException("Stock transaction not found")
        return StockTransactionResponse.model_validate(transaction)

    async def get_dashboard_summary(self) -> InventoryDashboardResponse:
        total_registered_items = await self.item_repo.count_all()
        stock_alerts = await self.alert_repo.count_active()
        active_warehouse_units = await self.warehouse_repo.count_active()
        total_vendors = await self.vendor_repo.count_all()
        return InventoryDashboardResponse(
            total_registered_items=total_registered_items,
            stock_alerts=stock_alerts,
            active_warehouse_units=active_warehouse_units,
            total_vendors=total_vendors,
        )

    async def generate_items_bulk_template(self) -> BytesIO:
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventory Items Bulk Import"

        headers = [
            "name", "sku", "barcode", "category", "quantity", "unit",
            "unit_cost", "reorder_level", "expiry_date", "warehouse_id",
            "vendor_id", "department_id", "description"
        ]
        ws.append(headers)

        
        # Configure Header Row Height
        ws.row_dimensions[1].height = 32
        
        # Header Style definition
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        header_font = Font(name="Calibri", size=12, bold=True, color="000000")
        header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_side = Side(style="thin", color="D3D3D3")
        header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = header_border
        
        # Add sample row
        sample_row = [
            "Disposable Syringes 10ml",
            "SKU-SYRINGE-10ML",
            "8901234567890",
            "Surgicals",
            150,
            "Box",
            12.50,
            20,
            "2028-12-31",
            1,
            1,
            1,
            "10ml syringes box of 100 units"
        ]
        ws.append(sample_row)

        
        # Freeze panes at A2
        ws.freeze_panes = "A2"
        
        # AutoFilter (using openpyxl get_column_letter dynamically based on actual header length)
        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}2"
        
        # Set custom column widths
        col_widths = {
            1: 25,  # name
            2: 20,  # sku
            3: 18,  # barcode
            4: 15,  # category
            5: 12,  # quantity
            6: 12,  # unit
            7: 12,  # unit_cost
            8: 15,  # reorder_level
            9: 15,  # expiry_date
            10: 15, # warehouse_id
            11: 15, # vendor_id
            12: 15, # department_id
            13: 30  # description
        }
        for col_idx, width in col_widths.items():
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width
        
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream

    async def import_items_from_excel(self, file, user_id: int) -> dict:
        from pydantic import ValidationError
        import openpyxl
        from datetime import date, datetime

        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents))
        ws = wb.active

        # Read headers
        header_row = next(ws.iter_rows(max_row=1, values_only=True), None)
        if not header_row:
            raise BadRequestException("The uploaded file is empty or has no headers.")

        headers = [str(h).strip().lower() for h in header_row if h is not None]
        required_headers = {"name", "category", "unit"}
        if not required_headers.issubset(set(headers)):
            raise BadRequestException("Missing required headers. File must contain name, category, and unit headers.")

        total_rows = 0
        created = 0
        failed = 0
        errors = []
        batch_skus = set()
        batch_barcodes = set()
        batch_names = set()

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None for cell in row):
                continue

            total_rows += 1
            row_dict = {}
            for header, val in zip(headers, row):
                if val is None or str(val).strip() == "":
                    row_dict[header] = None
                else:
                    row_dict[header] = val

            try:
                # Normalization
                barcode_raw = row_dict.get("barcode")
                if barcode_raw is not None:
                    if isinstance(barcode_raw, float):
                        barcode_raw = str(int(barcode_raw))
                    else:
                        barcode_raw = str(barcode_raw).strip()
                    row_dict["barcode"] = barcode_raw

                expiry_raw = row_dict.get("expiry_date")
                if expiry_raw is not None:
                    if isinstance(expiry_raw, (datetime, date)):
                        row_dict["expiry_date"] = expiry_raw.strftime("%Y-%m-%d")
                    else:
                        row_dict["expiry_date"] = str(expiry_raw).strip()

                cost_raw = row_dict.get("unit_cost")
                if cost_raw is not None:
                    try:
                        row_dict["unit_cost"] = float(cost_raw)
                    except ValueError:
                        pass

                qty_raw = row_dict.get("quantity")
                if qty_raw is not None:
                    try:
                        row_dict["quantity"] = int(float(qty_raw))
                    except ValueError:
                        pass

                reorder_raw = row_dict.get("reorder_level")
                if reorder_raw is not None:
                    try:
                        row_dict["reorder_level"] = int(float(reorder_raw))
                    except ValueError:
                        pass

                wh_raw = row_dict.get("warehouse_id")
                if wh_raw is not None:
                    try:
                        row_dict["warehouse_id"] = int(float(wh_raw))
                    except ValueError:
                        pass

                vendor_raw = row_dict.get("vendor_id")
                if vendor_raw is not None:
                    try:
                        row_dict["vendor_id"] = int(float(vendor_raw))
                    except ValueError:
                        pass

                dept_raw = row_dict.get("department_id")
                if dept_raw is not None:
                    try:
                        row_dict["department_id"] = int(float(dept_raw))
                    except ValueError:
                        pass

                # Validation using schema
                validated_data = InventoryItemCreate(**row_dict)

                # Check entity existence (department, warehouse, vendor)
                await self._validate_department(validated_data.department_id)
                await self._validate_warehouse(validated_data.warehouse_id)
                await self._validate_vendor(validated_data.vendor_id)

                # Name uniqueness
                name_key = validated_data.name.strip().lower()
                if name_key in batch_names:
                    raise ConflictException("Duplicate name in the uploaded file")
                existing_name = await self.item_repo.get_by_name(validated_data.name)
                if existing_name:
                    raise ConflictException("Inventory item with this name already exists")
                batch_names.add(name_key)

                # SKU uniqueness
                sku = validated_data.sku
                if sku:
                    sku_key = sku.strip().lower()
                    if sku_key in batch_skus:
                        raise ConflictException("Duplicate SKU in the uploaded file")
                    existing_sku = await self.item_repo.get_by_sku(sku)
                    if existing_sku:
                        raise ConflictException("Inventory item with this SKU already exists")
                    batch_skus.add(sku_key)
                else:
                    while True:
                        sku = generate_code("SKU")
                        existing_sku = await self.item_repo.get_by_sku(sku)
                        if not existing_sku and sku.strip().lower() not in batch_skus:
                            break
                    batch_skus.add(sku.strip().lower())

                # Barcode uniqueness
                barcode = validated_data.barcode
                if barcode:
                    barcode_key = barcode.strip().lower()
                    if barcode_key in batch_barcodes:
                        raise ConflictException("Duplicate barcode in the uploaded file")
                    existing_barcode = await self.item_repo.get_by_barcode(barcode)
                    if existing_barcode:
                        raise ConflictException("Inventory item with this barcode already exists")
                    batch_barcodes.add(barcode_key)

                # Insert item
                item = InventoryItem(sku=sku, **validated_data.model_dump(exclude={"sku"}))
                item = await self.item_repo.create(item)
                await self.audit_repo.create("create", "inventory", user_id=user_id, resource_id=str(item.id))
                created += 1

            except ValidationError as e:
                failed += 1
                err_msg = "; ".join([f"{'.'.join(str(loc) for loc in error['loc'])}: {error['msg']}" for error in e.errors()])
                errors.append({"row": row_idx, "error": err_msg})
            except ConflictException as e:
                failed += 1
                errors.append({"row": row_idx, "error": str(e.detail)})
            except NotFoundException as e:
                failed += 1
                errors.append({"row": row_idx, "error": str(e.detail)})
            except Exception as e:
                failed += 1
                errors.append({"row": row_idx, "error": str(e)})

        await self.db.flush()
        return {
            "total_rows": total_rows,
            "created": created,
            "failed": failed,
            "errors": errors,
        }

    async def export_items(self, format_type: str) -> tuple[BytesIO | bytes, str]:
        from datetime import date, datetime

        items = await self.item_repo.get_all_active()

        if format_type == "excel":
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Inventory Items"

            headers = [
                "Sr. No.", "name", "sku", "barcode", "category", "quantity", "unit",
                "unit_cost", "reorder_level", "expiry_date", "warehouse_id",
                "vendor_id", "department_id", "created_at", "updated_at"
            ]
            ws.append(headers)

            
            # Configure Header Row Height
            ws.row_dimensions[1].height = 32
            
            # Header Style definition
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            header_font = Font(name="Calibri", size=12, bold=True, color="000000")
            header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_side = Side(style="thin", color="D3D3D3")
            header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = header_border
            
            for sr_no, item in enumerate(items, start=1):
                row = [
                    sr_no,
                    item.name,
                    item.sku,
                    item.barcode or "",
                    item.category,
                    int(item.quantity) if item.quantity is not None else 0,
                    item.unit,
                    float(item.unit_cost) if item.unit_cost is not None else 0.0,
                    int(item.reorder_level) if item.reorder_level is not None else 0,
                    item.expiry_date.strftime("%Y-%m-%d") if isinstance(item.expiry_date, (date, datetime)) else (item.expiry_date or ""),
                    int(item.warehouse_id) if item.warehouse_id is not None else "",
                    int(item.vendor_id) if item.vendor_id is not None else "",
                    int(item.department_id) if item.department_id is not None else "",
                    item.created_at.strftime("%Y-%m-%d %H:%M:%S") if isinstance(item.created_at, datetime) else str(item.created_at),
                    item.updated_at.strftime("%Y-%m-%d %H:%M:%S") if isinstance(item.updated_at, datetime) else (str(item.updated_at) if item.updated_at else "")
                ]
                ws.append(row)

                
            # Freeze panes at A2
            ws.freeze_panes = "A2"
            
            # AutoFilter (using openpyxl get_column_letter dynamically based on actual header length)
            total_rows = len(items) + 1
            last_col_letter = openpyxl.utils.get_column_letter(len(headers))
            ws.auto_filter.ref = f"A1:{last_col_letter}{total_rows}"
            
            # Set custom column widths
            col_widths = {
                1: 8,   # Sr. No.
                2: 25,  # name
                3: 20,  # sku
                4: 18,  # barcode
                5: 15,  # category
                6: 12,  # quantity
                7: 12,  # unit
                8: 12,  # unit_cost
                9: 15,  # reorder_level
                10: 15, # expiry_date
                11: 15, # warehouse_id
                12: 15, # vendor_id
                13: 15, # department_id
                14: 22, # created_at
                15: 22  # updated_at
            }
            for col_idx, width in col_widths.items():
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = width
                
            stream = BytesIO()
            wb.save(stream)
            stream.seek(0)
            return stream, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        elif format_type == "pdf":
            from jinja2 import Environment, FileSystemLoader
            from app.utils.pdf_generator import html_to_pdf
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from xhtml2pdf import default
            import os

            font_path = os.path.abspath("app/static/fonts/DejaVuSans.ttf")
            if "DejaVuSans" not in pdfmetrics.getRegisteredFontNames() and os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont("DejaVuSans", font_path))

            default.DEFAULT_FONT["dejavusans"] = "DejaVuSans"
            default.DEFAULT_FONT["dejavusans-bold"] = "DejaVuSans"
            default.DEFAULT_FONT["dejavusans-oblique"] = "DejaVuSans"
            default.DEFAULT_FONT["dejavusans-boldoblique"] = "DejaVuSans"

            env = Environment(loader=FileSystemLoader("app/templates"))
            template = env.get_template("inventory_items_export_template.html")

            formatted_items = []
            for item in items:
                expiry_str = "-"
                if item.expiry_date:
                    if isinstance(item.expiry_date, (date, datetime)):
                        expiry_str = item.expiry_date.strftime("%Y-%m-%d")
                    else:
                        expiry_str = str(item.expiry_date)

                created_str = "-"
                if item.created_at:
                    if isinstance(item.created_at, datetime):
                        created_str = item.created_at.strftime("%Y-%m-%d %H:%M:%S")
                    else:
                        created_str = str(item.created_at)

                updated_str = "-"
                if item.updated_at:
                    if isinstance(item.updated_at, datetime):
                        updated_str = item.updated_at.strftime("%Y-%m-%d %H:%M:%S")
                    else:
                        updated_str = str(item.updated_at)

                formatted_items.append({
                    "id": item.id,
                    "name": item.name,
                    "sku": item.sku,
                    "barcode": item.barcode,
                    "category": item.category,
                    "quantity": int(item.quantity) if item.quantity is not None else 0,
                    "unit": item.unit,
                    "unit_cost": float(item.unit_cost) if item.unit_cost is not None else 0.0,
                    "reorder_level": int(item.reorder_level) if item.reorder_level is not None else 0,
                    "expiry_date": expiry_str,
                    "warehouse_id": item.warehouse_id,
                    "vendor_id": item.vendor_id,
                    "department_id": item.department_id,
                    "warehouse_name": item.warehouse.name if item.warehouse else "-",
                    "vendor_name": item.vendor.name if item.vendor else "-",
                    "department_name": item.department.department_name if item.department else "-",
                    "description": item.description or "",
                    "is_active": "Yes" if item.is_active else "No",
                    "created_at": created_str,
                    "updated_at": updated_str,
                })

            html = template.render(
                items=formatted_items,
                generated_at=utc_now().strftime("%Y-%m-%d %H:%M:%S")
            )

            pdf_bytes = html_to_pdf(html)
            return pdf_bytes, "application/pdf"

        else:
            raise BadRequestException("Invalid format specified for export")



    async def create_warehouse(self, data: WarehouseCreate, hospital_id: int | None) -> WarehouseResponse:
        warehouse = Warehouse(
            name=data.name,
            code=data.code if hasattr(data, 'code') and data.code else generate_code('WH'),
            location=data.location,
            hospital_id=hospital_id
        )
        self.db.add(warehouse)
        await self.db.flush()
        return WarehouseResponse.model_validate(warehouse)

    async def list_warehouses(self, hospital_id: int | None, page: int = 1, size: int = 20):
        query = select(Warehouse)
        if hospital_id is not None:
            query = query.where(Warehouse.hospital_id == hospital_id)
        else:
            query = query.where(Warehouse.hospital_id.is_(None))
        result = await self.db.execute(query.offset((page - 1) * size).limit(size))
        items = result.scalars().all()

        total_query = select(func.count(Warehouse.id))
        if hospital_id is not None:
            total_query = total_query.where(Warehouse.hospital_id == hospital_id)
        else:
            total_query = total_query.where(Warehouse.hospital_id.is_(None))

        total = await self.db.execute(total_query)
        total_count = total.scalar() or 0
        return build_paginated_result([WarehouseResponse.model_validate(w) for w in items], total_count, page, size)

    async def get_warehouse(self, warehouse_id: int, hospital_id: int | None) -> WarehouseResponse:
        warehouse = await self.db.get(Warehouse, warehouse_id)
        if not warehouse or warehouse.hospital_id != hospital_id:
            raise HTTPException(status_code=404, detail="Warehouse not found")
        return WarehouseResponse.model_validate(warehouse)

    async def update_warehouse(self, warehouse_id: int, data: WarehouseUpdate, hospital_id: int | None) -> WarehouseResponse:
        warehouse = await self.db.get(Warehouse, warehouse_id)
        if not warehouse or warehouse.hospital_id != hospital_id:
            raise HTTPException(status_code=404, detail="Warehouse not found")

        if data.name is not None:
            warehouse.name = data.name
        if data.location is not None:
            warehouse.location = data.location

        await self.db.flush()
        return WarehouseResponse.model_validate(warehouse)

    async def delete_warehouse(self, warehouse_id: int, hospital_id: int | None) -> None:
        warehouse = await self.db.get(Warehouse, warehouse_id)
        if not warehouse or warehouse.hospital_id != hospital_id:
            raise HTTPException(status_code=404, detail="Warehouse not found")

        # Soft delete logic or check if empty
        stock_count = await self.db.execute(select(func.count(WarehouseStock.id)).where(WarehouseStock.warehouse_id == warehouse_id, WarehouseStock.quantity > 0))
        if stock_count.scalar() > 0:
            raise HTTPException(status_code=400, detail="Cannot delete warehouse with active stock")

        await self.db.delete(warehouse)
        await self.db.flush()

    async def get_stock_summary(self, hospital_id: int | None) -> StockSummary:
        total_items = await self.db.scalar(select(func.count(WarehouseStock.id)).join(Warehouse).where(Warehouse.hospital_id == hospital_id)) or 0
        total_quantity = await self.db.scalar(select(func.sum(WarehouseStock.quantity)).join(Warehouse).where(Warehouse.hospital_id == hospital_id)) or 0
        total_value = await self.db.scalar(select(func.sum(WarehouseStock.quantity * InventoryItem.unit_cost)).join(Warehouse).join(InventoryItem, WarehouseStock.inventory_item_id == InventoryItem.id).where(Warehouse.hospital_id == hospital_id)) or 0.0
        low_stock_count = await self.db.scalar(select(func.count(WarehouseStock.id)).join(Warehouse).join(InventoryItem, WarehouseStock.inventory_item_id == InventoryItem.id).where(Warehouse.hospital_id == hospital_id, WarehouseStock.quantity < InventoryItem.reorder_level)) or 0

        return StockSummary(
            total_items=total_items,
            total_quantity=int(total_quantity),
            low_stock_count=low_stock_count,
            expired_count=0,
            total_value=float(total_value),
            total_registered_items=total_items,
            stock_alerts=low_stock_count,
            active_warehouse_units=0,
            inactive_warehouse_units=0,
            total_vendors=0
        )

