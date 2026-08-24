from datetime import date
from fastapi import UploadFile
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.core.exceptions import NotFoundException, ConflictException, BadRequestException
from app.models.patient_model import FamilyMember, Patient, PatientDocument
from app.repositories.audit_repository import AuditRepository
from app.repositories.patient_repository import PatientRepository
from app.schemas.patient_schema import (
    FamilyMemberCreate,
    FamilyMemberResponse,
    PatientCreate,
    PatientDocumentResponse,
    PatientResponse,
    PatientCreateResponse,
    PatientUpdate,
)
from app.utils.file_upload import save_upload_file
from app.utils.helpers import generate_mrn
from app.utils.pagination import build_paginated_result


class PatientService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.repo = PatientRepository(db)
        self.audit_repo = AuditRepository(db)

    async def _enrich_assigned_patients(self, patient_responses: list[PatientResponse], patient_ids: list[int], current_user) -> None:
        """
        Enriches PatientResponse objects with bed allocation and clinical status details
        specifically for the logged-in nurse, performing batch queries to avoid N+1 issues.
        """
        if not patient_ids or not current_user or not current_user.role or current_user.role.name.lower() != "nurse":
            return

        # 1. Resolve nurse_id first
        from app.models.nurse_model import Nurse
        from sqlalchemy import select
        res = await self.db.execute(select(Nurse.id).where(Nurse.user_id == current_user.id))
        nurse_id = res.scalar_one_or_none()
        if not nurse_id:
            return

        # 2. Fetch Bed allocation details (Bed, Room, Floor)
        from app.models.bed_allocation_model import Bed, Room, Floor
        bed_query = (
            select(Bed, Room, Floor)
            .join(Room, Room.id == Bed.room_id)
            .join(Floor, Floor.id == Room.floor_id)
            .where(
                Bed.patient_id.in_(patient_ids),
                Bed.status.in_(["Occupied", "Reserved"])
            )
        )
        bed_results = await self.db.execute(bed_query)
        bed_lookup = {}
        for bed, room, floor in bed_results.all():
            from app.schemas.patient_schema import PatientBedAllocationResponse
            bed_lookup[bed.patient_id] = PatientBedAllocationResponse(
                bed_id=bed.id,
                bed_name=bed.name,
                bed_type=bed.type,
                room_id=room.id,
                room_number=room.number,
                room_name=room.name,
                floor_id=floor.id,
                floor_number=floor.number,
                floor_name=floor.name,
                allocation_time=bed.allocation_time,
                admission_date=bed.admission_date
            )

        # 3. Fetch NursePatientAssignment status
        from app.models.nurse_model import NursePatientAssignment
        assign_query = (
            select(NursePatientAssignment)
            .where(
                NursePatientAssignment.nurse_id == nurse_id,
                NursePatientAssignment.patient_id.in_(patient_ids),
                NursePatientAssignment.status == "Active"
            )
        )
        assign_results = await self.db.execute(assign_query)
        status_lookup = {
            a.patient_id: a.patient_status 
            for a in assign_results.scalars().all()
        }

        # 4. Map back to responses
        for p_res in patient_responses:
            p_res.bed_allocation = bed_lookup.get(p_res.id)
            p_res.condition_status = status_lookup.get(p_res.id)

    async def list_patients(
        self,
        page: int = 1,
        size: int = 20,
        sort_by: str = "created_at",
        sort_order: str = "desc",
        start_date: date | None = None,
        end_date: date | None = None,
        current_user = None,
    ):
        if start_date and end_date and start_date > end_date:
            raise BadRequestException("Start date cannot be greater than end date")

        # Resolve nurse_id if role is Nurse
        nurse_id = None
        if current_user and current_user.role and current_user.role.name.lower() == "nurse":
            from app.models.nurse_model import Nurse
            from sqlalchemy import select
            res = await self.db.execute(select(Nurse.id).where(Nurse.user_id == current_user.id))
            nurse_id = res.scalar_one_or_none()
            if nurse_id is None:
                return {
                    "items": [],
                    "total": 0,
                    "page": page,
                    "size": size,
                    "pages": 0,
                    "active_count": 0,
                    "inactive_count": 0,
                    "cities_count": 0,
                }

        skip = (page - 1) * size
        items = await self.repo.list_all(
            skip=skip,
            limit=size,
            sort_by=sort_by,
            sort_order=sort_order,
            start_date=start_date,
            end_date=end_date,
            nurse_id=nurse_id,
        )
        total = await self.repo.count_all(start_date=start_date, end_date=end_date, nurse_id=nurse_id)
        
        patient_responses = [PatientResponse.model_validate(p) for p in items]
        patient_ids = [p.id for p in items]
        await self._enrich_assigned_patients(patient_responses, patient_ids, current_user)

        paginated = build_paginated_result(
            patient_responses, total, page, size
        )
        stats = await self.repo.get_patient_stats(nurse_id=nurse_id)
        return {
            "items": paginated.items,
            "total": paginated.total,
            "page": paginated.page,
            "size": paginated.size,
            "pages": paginated.pages,
            **stats
        }

    async def get_by_id(self, patient_id: int) -> PatientResponse:
        patient = await self.repo.get_by_id(patient_id)
        if not patient:
            raise NotFoundException("Patient not found")
        return PatientResponse.model_validate(patient)

    async def create(
        self, data: PatientCreate, user_id: int, consent_file: UploadFile | None = None
    ) -> PatientCreateResponse:
        if data.phone:
            existing_phone = await self.repo.get_by_phone(data.phone)
            if existing_phone:
                raise ConflictException("Patient with this phone number already exists")

        if data.email:
            existing_email = await self.repo.get_by_email(data.email)
            if existing_email:
                raise ConflictException("Patient with this email already exists")

        patient = Patient(patient_code=generate_mrn(), **data.model_dump())
        patient = await self.repo.create(patient)

        if consent_file:
            file_path = await save_upload_file(consent_file, settings.UPLOAD_DIR)
            doc = PatientDocument(
                patient_id=patient.id,
                document_name=consent_file.filename or "consent_form",
                document_type="Consent Form",
                file_path=file_path,
                uploaded_by=user_id,
            )
            await self.repo.add_document(doc)

        await self.audit_repo.create("create", "patients", user_id=user_id, resource_id=str(patient.id))
        return PatientCreateResponse.model_validate(patient)

    async def update(self, patient_id: int, data: PatientUpdate, user_id: int) -> PatientResponse:
        patient = await self.repo.get_by_id(patient_id)
        if not patient:
            raise NotFoundException("Patient not found")

        if data.phone is not None and data.phone != patient.phone:
            existing_phone = await self.repo.get_by_phone(data.phone)
            if existing_phone:
                raise ConflictException("Patient with this phone number already exists")

        if data.email is not None and data.email != patient.email:
            existing_email = await self.repo.get_by_email(data.email)
            if existing_email:
                raise ConflictException("Patient with this email already exists")

        for key, value in data.model_dump(exclude_unset=True).items():
            setattr(patient, key, value)
        patient = await self.repo.update(patient)
        await self.audit_repo.create("update", "patients", user_id=user_id, resource_id=str(patient.id))
        return PatientResponse.model_validate(patient)

    async def delete(self, patient_id: int, user_id: int) -> None:
        patient = await self.repo.get_by_id(patient_id)
        if not patient:
            raise NotFoundException("Patient not found")
        await self.repo.soft_delete(patient)
        
        # Free any beds allocated to this patient
        from app.models.bed_allocation_model import Bed, BedActivityLog
        from sqlalchemy import select
        result = await self.db.execute(select(Bed).where(Bed.patient_id == patient_id))
        beds = result.scalars().all()
        for bed in beds:
            bed.status = "Available"
            bed.patient_id = None
            bed.allocation_time = None
            bed.admission_date = None
            
            # Log the release activity
            log = BedActivityLog(
                type="release",
                message=f"Automatically released Bed {bed.name} because patient {patient.first_name} {patient.last_name} was deleted.",
                floor_id=None,
                room_id=bed.room_id,
                bed_id=bed.id,
                patient_id=patient_id,
            )
            self.db.add(log)
            
        if patient.user_id:
            from app.models.user_model import User
            user = await self.db.get(User, patient.user_id)
            if user:
                user.is_active = False
        await self.audit_repo.create("delete", "patients", user_id=user_id, resource_id=str(patient.id))

    async def search(self, q: str, page: int = 1, size: int = 20, current_user = None):
        nurse_id = None
        if current_user and current_user.role and current_user.role.name.lower() == "nurse":
            from app.models.nurse_model import Nurse
            from sqlalchemy import select
            res = await self.db.execute(select(Nurse.id).where(Nurse.user_id == current_user.id))
            nurse_id = res.scalar_one_or_none()
            if nurse_id is None:
                return build_paginated_result([], 0, page, size)

        skip = (page - 1) * size
        items = await self.repo.search(q, skip=skip, limit=size, nurse_id=nurse_id)
        total = await self.repo.count_search(q, nurse_id=nurse_id)
        
        patient_responses = [PatientResponse.model_validate(p) for p in items]
        patient_ids = [p.id for p in items]
        await self._enrich_assigned_patients(patient_responses, patient_ids, current_user)

        return build_paginated_result(
            patient_responses, total, page, size
        )

    async def filter_patients(
        self,
        gender: str | None = None,
        blood_group: str | None = None,
        city: str | None = None,
        state: str | None = None,
        status: str | None = None,
        page: int = 1,
        size: int = 20,
        current_user = None,
    ):
        nurse_id = None
        if current_user and current_user.role and current_user.role.name.lower() == "nurse":
            from app.models.nurse_model import Nurse
            from sqlalchemy import select
            res = await self.db.execute(select(Nurse.id).where(Nurse.user_id == current_user.id))
            nurse_id = res.scalar_one_or_none()
            if nurse_id is None:
                return build_paginated_result([], 0, page, size)

        skip = (page - 1) * size
        items = await self.repo.filter_patients(
            gender=gender, blood_group=blood_group, city=city, state=state, status=status,
            skip=skip, limit=size, nurse_id=nurse_id,
        )
        total = await self.repo.count_filter(
            gender=gender, blood_group=blood_group, city=city, state=state, status=status,
            nurse_id=nurse_id,
        )
        
        patient_responses = [PatientResponse.model_validate(p) for p in items]
        patient_ids = [p.id for p in items]
        await self._enrich_assigned_patients(patient_responses, patient_ids, current_user)

        return build_paginated_result(
            patient_responses, total, page, size
        )

    async def get_appointments(self, patient_id: int):
        await self.get_by_id(patient_id)
        from app.schemas.appointment_schema import AppointmentResponse

        appointments = await self.repo.get_appointments(patient_id)
        return [AppointmentResponse.model_validate(a) for a in appointments]

    async def get_history(self, patient_id: int):
        return await self.get_appointments(patient_id)

    async def add_family_member(self, patient_id: int, data: FamilyMemberCreate, user_id: int) -> FamilyMemberResponse:
        await self.get_by_id(patient_id)
        member = FamilyMember(patient_id=patient_id, **data.model_dump())
        member = await self.repo.add_family_member(member)
        await self.audit_repo.create("create", "family_members", user_id=user_id, resource_id=str(member.id))
        return FamilyMemberResponse.model_validate(member)

    async def list_family_members(self, patient_id: int) -> list[FamilyMemberResponse]:
        await self.get_by_id(patient_id)
        members = await self.repo.list_family_members(patient_id)
        return [FamilyMemberResponse.model_validate(m) for m in members]

    async def upload_document(
        self, patient_id: int, file: UploadFile, document_type: str, user_id: int
    ) -> PatientDocumentResponse:
        await self.get_by_id(patient_id)
        file_path = await save_upload_file(file, settings.UPLOAD_DIR)
        doc = PatientDocument(
            patient_id=patient_id,
            document_name=file.filename or "document",
            document_type=document_type,
            file_path=file_path,
            uploaded_by=user_id,
        )
        doc = await self.repo.add_document(doc)
        await self.audit_repo.create("upload", "patient_documents", user_id=user_id, resource_id=str(doc.id))
        return PatientDocumentResponse.model_validate(doc)

    async def list_documents(self, patient_id: int) -> list[PatientDocumentResponse]:
        await self.get_by_id(patient_id)
        docs = await self.repo.list_documents(patient_id)
        return [PatientDocumentResponse.model_validate(d) for d in docs]

    async def get_document(self, patient_id: int, document_id: int) -> PatientDocument:
        await self.get_by_id(patient_id)
        doc = await self.repo.get_document(document_id)
        if not doc:
            raise NotFoundException("Document not found")
        if doc.patient_id != patient_id:
            raise BadRequestException("Document does not belong to this patient")
        return doc

    async def delete_document(self, patient_id: int, document_id: int, user_id: int) -> None:
        doc = await self.get_document(patient_id, document_id)
        import os
        if os.path.exists(doc.file_path):
            try:
                os.remove(doc.file_path)
            except Exception:
                pass
        await self.repo.delete_document(doc)
        await self.audit_repo.create("delete", "patient_documents", user_id=user_id, resource_id=str(document_id))

    async def delete_family_member(self, patient_id: int, member_id: int, user_id: int) -> None:
        await self.get_by_id(patient_id)
        member = await self.repo.get_family_member(member_id)
        if not member:
            raise NotFoundException("Family member not found")
        if member.patient_id != patient_id:
            raise BadRequestException("Family member does not belong to this patient")
        await self.repo.delete_family_member(member)
        await self.audit_repo.create("delete", "family_members", user_id=user_id, resource_id=str(member_id))

    async def generate_patient_bulk_template(self):
        from io import BytesIO
        import openpyxl
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Patient Bulk Template"
        
        headers = [
            "First Name", "Last Name", "Gender", "Date of Birth", "Blood Group", 
            "Marital Status", "Phone", "Email", "Address", "City", "State", 
            "Pincode", "Emergency Contact Name", "Emergency Contact Number", 
            "Allergies", "Medical History", "Chronic Disease", "Diagnosis", 
            "Insurance Provider", "Insurance Number", "Status", "Preferred Language"
        ]
        ws.append(headers)
        
        ws.append([
            "John", "Doe", "Male", "1985-05-15", "O+", "Married", "9876543210", 
            "john.doe@example.com", "123 Health Street", "Mumbai", "Maharashtra", 
            "400001", "Jane Doe", "9876543211", "Peanuts", "Hypertension", "None", 
            "Routine Checkup", "Star Health", "SH123456", "active", "English"
        ])
        
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream

    async def import_patients_from_excel(self, file, user_id: int) -> dict:
        from io import BytesIO
        from pydantic import ValidationError
        import openpyxl
        
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents))
        ws = wb.active
        
        header_row = next(ws.iter_rows(max_row=1, values_only=True), None)
        if not header_row:
            raise BadRequestException("The uploaded file is empty or has no headers.")
            
        headers = [str(h).strip().lower() for h in header_row if h is not None]
        required_headers = {"first name", "last name", "diagnosis"}
        if not required_headers.issubset(set(headers)):
            raise BadRequestException("Missing required headers in the upload template.")
            
        total_rows = 0
        created = 0
        failed = 0
        errors = []
        
        seen_phones = set()
        seen_emails = set()
        
        header_mapping = {
            "first name": "first_name",
            "last name": "last_name",
            "gender": "gender",
            "date of birth": "dob",
            "blood group": "blood_group",
            "marital status": "marital_status",
            "phone": "phone",
            "email": "email",
            "address": "address",
            "city": "city",
            "state": "state",
            "pincode": "pincode",
            "emergency contact name": "emergency_contact_name",
            "emergency contact number": "emergency_contact_number",
            "allergies": "allergies",
            "medical history": "medical_history",
            "chronic disease": "chronic_disease",
            "diagnosis": "diagnosis",
            "insurance provider": "insurance_provider",
            "insurance_number": "insurance_number",
            "insurance number": "insurance_number",
            "status": "status",
            "preferred language": "preferred_language"
        }
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None for cell in row):
                continue
                
            total_rows += 1
            row_dict = {}
            for header, val in zip(headers, row):
                mapped_key = header_mapping.get(header)
                if not mapped_key:
                    continue
                if val is None or str(val).strip() == "" or str(val).strip().lower() == "none":
                    row_dict[mapped_key] = None
                else:
                    row_dict[mapped_key] = str(val).strip()
                    
            try:
                dob_raw = row_dict.get("dob")
                if dob_raw:
                    from datetime import datetime, date
                    try:
                        if isinstance(dob_raw, (datetime, date)):
                            row_dict["dob"] = dob_raw
                        else:
                            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
                                try:
                                    row_dict["dob"] = datetime.strptime(dob_raw, fmt).date()
                                    break
                                except ValueError:
                                    continue
                            else:
                                raise ValueError(f"Invalid date format '{dob_raw}'. Expected YYYY-MM-DD.")
                    except Exception as ex:
                        raise BadRequestException(str(ex))
                
                phone = row_dict.get("phone")
                if phone:
                    if phone in seen_phones:
                        raise BadRequestException(f"Duplicate phone number '{phone}' found in upload file.")
                    seen_phones.add(phone)
                    
                email = row_dict.get("email")
                if email:
                    if email.lower() in seen_emails:
                        raise BadRequestException(f"Duplicate email '{email}' found in upload file.")
                    seen_emails.add(email.lower())
                
                patient_create = PatientCreate(**row_dict)
                
                await self.create(patient_create, user_id)
                created += 1
                
            except ValidationError as e:
                failed += 1
                err_msg = "; ".join([f"{'.'.join(str(loc) for loc in error['loc'])}: {error['msg']}" for error in e.errors()])
                errors.append({
                    "row": row_idx,
                    "error": err_msg
                })
            except ConflictException as e:
                failed += 1
                errors.append({
                    "row": row_idx,
                    "error": str(e.detail)
                })
            except NotFoundException as e:
                failed += 1
                errors.append({
                    "row": row_idx,
                    "error": str(e.detail)
                })
            except BadRequestException as e:
                failed += 1
                errors.append({
                    "row": row_idx,
                    "error": str(e.detail)
                })
            except Exception as e:
                failed += 1
                errors.append({
                    "row": row_idx,
                    "error": str(e)
                })
                
        await self.db.flush()
        return {
            "total_rows": total_rows,
            "created": created,
            "failed": failed,
            "errors": errors
        }

    async def export_patients(self, format_type: str, status: str | None = None):
        from io import BytesIO
        from datetime import datetime, date
        
        def format_full_address(address: str | None, city: str | None, state: str | None, pincode: str | None) -> str:
            parts = []
            if address and str(address).strip() and str(address).lower() not in ("none", "null"):
                parts.append(str(address).strip())
            if city and str(city).strip() and str(city).lower() not in ("none", "null"):
                parts.append(str(city).strip())
            if state and str(state).strip() and str(state).lower() not in ("none", "null"):
                parts.append(str(state).strip())
            
            main_addr = ", ".join(parts)
            
            pincode_clean = None
            if pincode and str(pincode).strip() and str(pincode).lower() not in ("none", "null"):
                pincode_clean = str(pincode).strip()
                
            if main_addr:
                if pincode_clean:
                    return f"{main_addr} - {pincode_clean}"
                return main_addr
            elif pincode_clean:
                return pincode_clean
            return "-"

        def format_emergency_contact(name: str | None, phone: str | None, line_break: str = "\n") -> str:
            name_clean = str(name).strip() if name and str(name).strip() and str(name).lower() not in ("none", "null") else None
            phone_clean = str(phone).strip() if phone and str(phone).strip() and str(phone).lower() not in ("none", "null") else None
            
            if name_clean and phone_clean:
                return f"{name_clean}{line_break}{phone_clean}"
            elif name_clean:
                return name_clean
            elif phone_clean:
                return phone_clean
            return "-"
        
        if status:
            patients = await self.repo.filter_patients(status=status, limit=10000)
        else:
            patients = await self.repo.list_all(limit=10000)
        
        if format_type == "excel":
            import openpyxl
            from openpyxl.styles import Alignment
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Patients Export"
            
            headers = [
                "Sr. No.", "Patient Code", "Full Name", "Gender", "DOB", 
                "Blood Group", "Phone", "Email", "Full Address", 
                "Emergency Contact", "Diagnosis", "Created At"
            ]
            ws.append(headers)
            
            for sr_no, p in enumerate(patients, start=1):
                row = [
                    sr_no,
                    p.patient_code,
                    f"{p.first_name} {p.last_name}".strip(),
                    p.gender or "",
                    str(p.dob) if p.dob else "",
                    p.blood_group or "",
                    p.phone or "",
                    p.email or "",
                    format_full_address(p.address, p.city, p.state, p.pincode),
                    format_emergency_contact(p.emergency_contact_name, p.emergency_contact_number, "\n"),
                    p.diagnosis or "",
                    p.created_at.strftime("%Y-%m-%d %H:%M:%S") if isinstance(p.created_at, datetime) else str(p.created_at)
                ]
                ws.append(row)
                
            # Enable wrap text for "Emergency Contact" column
            col_idx = headers.index("Emergency Contact") + 1
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=col_idx)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                
            stream = BytesIO()
            wb.save(stream)
            stream.seek(0)
            return stream, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            
        elif format_type == "pdf":
            from jinja2 import Environment, FileSystemLoader
            from app.utils.pdf_generator import html_to_pdf
            from app.utils.helpers import utc_now
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from xhtml2pdf import default
            import os
            
            font_path = os.path.abspath("app/static/fonts/DejaVuSans.ttf")
            if "DejaVuSans" not in pdfmetrics.getRegisteredFontNames() and os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont("DejaVuSans", font_path))
                
            default.DEFAULT_FONT["dejavusans"] = "DejaVuSans"
            default.DEFAULT_FONT["dejavusans-bold"] = "DejaVuSans"
            default.DEFAULT_FONT["dejavusans-oblique"] = "DejaVuSans"
            default.DEFAULT_FONT["dejavusans-boldoblique"] = "DejaVuSans"
            
            env = Environment(loader=FileSystemLoader("app/templates"))
            template = env.get_template("patients_export_template.html")
            
            formatted_patients = []
            for p in patients:
                formatted_patients.append({
                    "id": p.id,
                    "patient_code": p.patient_code,
                    "first_name": p.first_name,
                    "last_name": p.last_name,
                    "gender": p.gender,
                    "dob": str(p.dob) if p.dob else "",
                    "blood_group": p.blood_group,
                    "phone": p.phone,
                    "email": p.email,
                    "full_address": format_full_address(p.address, p.city, p.state, p.pincode),
                    "status": p.status,
                    "diagnosis": p.diagnosis,
                    "emergency_contact": format_emergency_contact(p.emergency_contact_name, p.emergency_contact_number, "<br/>"),
                    "created_at": p.created_at.strftime("%Y-%m-%d %H:%M:%S") if isinstance(p.created_at, datetime) else str(p.created_at),
                })
                
            html_content = template.render(
                patients=formatted_patients,
                generated_at=utc_now().strftime("%Y-%m-%d %H:%M:%S")
            )
            
            pdf_data = html_to_pdf(html_content)
            return pdf_data, "application/pdf"
