from datetime import date, datetime, time, timedelta
from io import BytesIO
from typing import Optional

# pyrefly: ignore [missing-import]
from sqlalchemy.ext.asyncio import AsyncSession


from app.core.constants import PharmacyStatus, PurchaseStatus
from app.core.exceptions import BadRequestException, ConflictException, ForbiddenException, NotFoundException
from app.models.pharmacy_model import (
    Medicine,
    PharmacyInvoice,
    PharmacyInvoiceItem,
    Prescription,
    PrescriptionItem,
    Purchase,
    PurchaseItem,
    Supplier,
)
from app.models.user_model import User
from app.repositories.audit_repository import AuditRepository
from app.repositories.patient_repository import PatientRepository
from app.repositories.pharmacy_repository import (
    MedicineRepository,
    PharmacyInvoiceRepository,
    PrescriptionRepository,
    PurchaseRepository,
    SupplierRepository,
    PharmacyDashboardRepository,
)
from app.schemas.pharmacy_schema import (
    ExpiryAlert,
    LowStockAlert,
    MedicineCreate,
    MedicineResponse,
    MedicineUpdate,
    PharmacyInvoiceCreate,
    PharmacyInvoiceUpdate,
    PharmacyInvoiceResponse,
    PharmacyInvoiceItemResponse,
    PharmacyDashboardResponse,
    PharmacyInventoryOverviewResponse,

    PrescriptionCreate,
    PrescriptionItemResponse,
    PrescriptionResponse,
    PrescriptionUpdate,
    PurchaseCreate,
    PurchaseItemResponse,
    PrescriptionStatusUpdate,
    PurchaseResponse,
    SalesReport,
    SupplierCreate,
    SupplierResponse,
    SupplierUpdate,
    PharmacyDashboardResponse,
    LowStockItemAlert,
    PharmacySalesTrendPoint,
    InventoryStatusMix,
    InventoryHealthProgress,
)
from app.utils.helpers import (
    calculate_gst_amount,
    generate_medicine_sku,
    generate_pharmacy_invoice_number,
    generate_prescription_number,
    generate_purchase_number,
    utc_now,
)
from app.utils.pagination import build_paginated_result


class PharmacyService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.medicine_repo = MedicineRepository(db)
        self.prescription_repo = PrescriptionRepository(db)
        self.invoice_repo = PharmacyInvoiceRepository(db)
        self.supplier_repo = SupplierRepository(db)
        self.purchase_repo = PurchaseRepository(db)
        self.dashboard_repo = PharmacyDashboardRepository(db)
        self.audit_repo = AuditRepository(db)
        self.patient_repo = PatientRepository(db)

    def get_date_range(
        self,
        time_filter: str,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> tuple[Optional[datetime], Optional[datetime]]:
        now_dt = utc_now()
        today_date = now_dt.date()

        if time_filter == "today":
            start = datetime.combine(today_date, time.min)
            end = datetime.combine(today_date, time.max)
            return start, end
        elif time_filter in ("7_days", "last_7_days"):
            start_date_val = today_date - timedelta(days=7)
            start = datetime.combine(start_date_val, time.min)
            end = now_dt
            return start, end
        elif time_filter in ("30_days", "last_30_days"):
            start_date_val = today_date - timedelta(days=30)
            start = datetime.combine(start_date_val, time.min)
            end = now_dt
            return start, end
        elif time_filter in ("month", "month_to_date"):
            start = datetime.combine(today_date.replace(day=1), time.min)
            end = now_dt
            return start, end
        elif time_filter == "3_month":
            start_date_val = today_date - timedelta(days=90)
            start = datetime.combine(start_date_val, time.min)
            end = now_dt
            return start, end
        elif time_filter == "custom":
            if not start_date or not end_date:
                start = datetime.combine(today_date, time.min)
                end = datetime.combine(today_date, time.max)
                return start, end
            start = datetime.combine(start_date, time.min)
            end = datetime.combine(end_date, time.max)
            return start, end
        else:
            return None, None

    # Duplicate get_dashboard_summary method removed to avoid method overriding



    # --- Medicines ---
    async def create_medicine(self, data: MedicineCreate, user_id: int) -> MedicineResponse:
        if data.barcode:
            existing = await self.medicine_repo.get_by_barcode(data.barcode)
            if existing:
                raise ConflictException("Medicine with this barcode already exists")
        medicine = Medicine(sku=generate_medicine_sku(), **data.model_dump())
        medicine = await self.medicine_repo.create(medicine)
        await self.audit_repo.create("create", "pharmacy", user_id=user_id, resource_id=str(medicine.id))
        return MedicineResponse.model_validate(medicine)

    async def list_medicines(
        self, page: int = 1, size: int = 20, sort_by: str = "created_at",
        sort_order: str = "desc", category: str | None = None,
    ):
        skip = (page - 1) * size
        items = await self.medicine_repo.list_all(skip=skip, limit=size, sort_by=sort_by,
                                                   sort_order=sort_order, category=category)
        total = await self.medicine_repo.count_all(category=category)
        return build_paginated_result([MedicineResponse.model_validate(m) for m in items], total, page, size)

    async def search_medicines(self, q: str, page: int = 1, size: int = 20):
        skip = (page - 1) * size
        items = await self.medicine_repo.search(q, skip=skip, limit=size)
        total = await self.medicine_repo.count_search(q)
        return build_paginated_result([MedicineResponse.model_validate(m) for m in items], total, page, size)

    async def get_medicine(self, medicine_id: int) -> MedicineResponse:
        medicine = await self.medicine_repo.get_by_id(medicine_id)
        if not medicine:
            raise NotFoundException("Medicine not found")
        return MedicineResponse.model_validate(medicine)

    async def update_medicine(self, medicine_id: int, data: MedicineUpdate, user_id: int) -> MedicineResponse:
        medicine = await self.medicine_repo.get_by_id(medicine_id)
        if not medicine:
            raise NotFoundException("Medicine not found")
        for key, value in data.model_dump(exclude_unset=True).items():
            setattr(medicine, key, value)
        medicine = await self.medicine_repo.update(medicine)
        await self.audit_repo.create("update", "pharmacy", user_id=user_id, resource_id=str(medicine.id))
        return MedicineResponse.model_validate(medicine)

    async def delete_medicine(self, medicine_id: int, user_id: int) -> None:
        medicine = await self.medicine_repo.get_by_id(medicine_id)
        if not medicine:
            raise NotFoundException("Medicine not found")
        await self.medicine_repo.soft_delete(medicine)
        await self.audit_repo.create("delete", "pharmacy", user_id=user_id, resource_id=str(medicine.id))

    async def get_low_stock(self) -> list[LowStockAlert]:
        medicines = await self.medicine_repo.get_low_stock()
        return [
            LowStockAlert(
                medicine_id=m.id, name=m.name, sku=m.sku,
                stock_quantity=m.stock_quantity, reorder_level=m.reorder_level,
            )
            for m in medicines
        ]

    async def get_expiry_alerts(self, days: int = 30) -> list[ExpiryAlert]:
        from app.utils.helpers import get_today_ist
        today = get_today_ist()
        medicines = await self.medicine_repo.get_expiry_alerts(reference_date=today, days=days)
        return [
            ExpiryAlert(
                medicine_id=m.id, name=m.name, sku=m.sku,
                expiry_date=m.expiry_date,
                stock_quantity=m.stock_quantity,
                days_until_expiry=(m.expiry_date - today).days if m.expiry_date else 0,
            )
            for m in medicines
        ]

    # --- Prescriptions ---

    async def create_prescription(self, data: PrescriptionCreate, user_id: int) -> PrescriptionResponse:
        if not data.appointment_id:
            raise BadRequestException("Appointment ID is required to create prescription")
        if not data.items:
            raise BadRequestException("Prescription must contain at least one medicine item")

        from sqlalchemy import select
        from app.models.doctor_model import Doctor
        from app.models.patient_model import Patient
        from app.models.appointment_model import Appointment

        # 1. Verify Doctor exists
        doctor_exists = await self.db.scalar(
            select(Doctor).where(Doctor.id == data.doctor_id)
        )
        if not doctor_exists:
            raise NotFoundException("Doctor not found")

        # 2. Verify Patient exists
        patient_exists = await self.db.scalar(
            select(Patient).where(Patient.id == data.patient_id)
        )
        if not patient_exists:
            raise NotFoundException("Patient not found")

        # 3. Verify Appointment exists and matches
        appointment_result = await self.db.execute(
            select(Appointment).where(Appointment.id == data.appointment_id)
        )
        appointment = appointment_result.scalar_one_or_none()
        if not appointment:
            raise NotFoundException("Appointment not found")

        if appointment.patient_id != data.patient_id:
            raise BadRequestException("Patient ID does not match the appointment")
        if appointment.doctor_id != data.doctor_id:
            raise BadRequestException("Doctor ID does not match the appointment")

        # 3b. Verify appointment date is not in the future
        from datetime import date
        from app.core.constants import AppointmentStatus
        if appointment.appointment_date is not None and appointment.appointment_date > date.today():
            raise BadRequestException("Cannot create prescription for future date appointments")

        # 3c. Verify appointment is confirmed or completed
        appointment_status = (appointment.appointment_status or "").strip().lower()
        if appointment_status not in ["confirmed", "completed"]:
            raise BadRequestException("Prescription can only be created for confirmed or completed appointments.")

        # 4. Verify no prescription exists for this appointment
        existing_rx = await self.db.scalar(
            select(Prescription).where(
                Prescription.appointment_id == data.appointment_id,
                Prescription.is_deleted.is_(False)
            )
        )
        if existing_rx:
            raise BadRequestException("A prescription already exists for this appointment")

        items = [
            PrescriptionItem(**item.model_dump()) for item in data.items
        ]
        prescription = Prescription(
            patient_id=data.patient_id,
            doctor_id=data.doctor_id,
            appointment_id=data.appointment_id,
            prescription_number=generate_prescription_number(),
            instructions=data.instructions,
        )
        prescription = await self.prescription_repo.create(prescription, items)
        prescription = await self.prescription_repo.get_by_id(prescription.id)
        await self.audit_repo.create(
            "create",
            "pharmacy_prescription",
            user_id=user_id,
            resource_id=str(prescription.id)
        )
        return self._prescription_response(prescription)
    async def list_prescriptions(
        self,
        page: int = 1,
        size: int = 20,
        status: str | None = None,
        doctor_id: int | None = None,
        patient_id: int | None = None,
        appointment_id: int | None = None,
        department_id: int | None = None,
        assigned_patient_ids: Optional[list[int]] = None,
    ):
        if assigned_patient_ids == []:
            return build_paginated_result([], 0, page, size)

        skip = (page - 1) * size
        items = await self.prescription_repo.list_all(
            skip=skip,
            limit=size,
            status=status,
            doctor_id=doctor_id,
            patient_id=patient_id,
            appointment_id=appointment_id,
            department_id=department_id,
            assigned_patient_ids=assigned_patient_ids
        )
        total = await self.prescription_repo.count_all(
            status=status,
            doctor_id=doctor_id,
            patient_id=patient_id,
            appointment_id=appointment_id,
            department_id=department_id,
            assigned_patient_ids=assigned_patient_ids
        )

        return build_paginated_result(
            [self._prescription_response(p) for p in items],
            total,
            page,
            size
        )

    async def get_prescription(self, prescription_id: int, doctor_id: int | None = None) -> PrescriptionResponse:
        prescription = await self.prescription_repo.get_by_id(prescription_id)
        if not prescription:
            raise NotFoundException("Prescription not found")
        if doctor_id is not None and prescription.doctor_id != doctor_id:
            raise ForbiddenException("You do not have permission to access this prescription")
        return self._prescription_response(prescription)

    async def update_prescription(
        self,
        prescription_id: int,
        data: PrescriptionUpdate,
        doctor_id: int,
        user_id: int,
        current_user: User
    ) -> PrescriptionResponse:
        prescription = await self.prescription_repo.get_by_id(prescription_id)
        if not prescription:
            raise NotFoundException("Prescription not found")
        if prescription.doctor_id != doctor_id:
            raise ForbiddenException("You do not have permission to modify this prescription")

        from app.core.constants import UserRole
        if current_user.role and current_user.role.name == UserRole.DOCTOR:
            if data.status is not None:
                status_clean = data.status.strip().lower()
                if status_clean not in {"pending", "sent_to_pharmacy"}:
                    raise BadRequestException("Doctors can only update prescription status to 'pending' or 'sent_to_pharmacy'.")

        if data.patient_id is not None:
            prescription.patient_id = data.patient_id
        if data.instructions is not None:
            prescription.instructions = data.instructions
        if data.status is not None:
            prescription.status = data.status

        items = None
        if data.items is not None:
            items = [PrescriptionItem(**item.model_dump()) for item in data.items]

        prescription = await self.prescription_repo.update(prescription, items)
        prescription = await self.prescription_repo.get_by_id(prescription.id)
        await self.audit_repo.create("update", "pharmacy_prescription", user_id=user_id, resource_id=str(prescription.id))
        return self._prescription_response(prescription)

    async def update_prescription_status(
        self,
        prescription_id: int,
        data: PrescriptionStatusUpdate,
        user_id: int
    ) -> PrescriptionResponse:
        prescription = await self.prescription_repo.get_by_id(prescription_id)
        if not prescription:
            raise NotFoundException("Prescription not found")

        prescription.status = data.status
        if data.status.lower() in ("completed", "dispensed"):
            prescription.dispensed_at = utc_now()
        else:
            prescription.dispensed_at = None

        prescription = await self.prescription_repo.update(prescription)
        prescription = await self.prescription_repo.get_by_id(prescription.id)

        await self.audit_repo.create(
            "update_status",
            "pharmacy_prescription",
            user_id=user_id,
            resource_id=str(prescription.id)
        )
        return self._prescription_response(prescription)


    async def delete_prescription(self, prescription_id: int, doctor_id: int, user_id: int) -> None:
        prescription = await self.prescription_repo.get_by_id(prescription_id)
        if not prescription:
            raise NotFoundException("Prescription not found")
        if prescription.doctor_id != doctor_id:
            raise ForbiddenException("You do not have permission to delete this prescription")
        await self.prescription_repo.soft_delete(prescription)
        await self.audit_repo.create("delete", "pharmacy_prescription", user_id=user_id, resource_id=str(prescription.id))


    def _prescription_response(
        self,
        prescription: Prescription
    ) -> PrescriptionResponse:
        resp = PrescriptionResponse.model_validate(prescription)
        resp.items = [
            PrescriptionItemResponse.model_validate(i)
            for i in prescription.items
        ]
        return resp



    # --- Invoices ---
    async def create_invoice(self, data: PharmacyInvoiceCreate, user_id: int) -> PharmacyInvoiceResponse:
        if data.patient_id is not None:
            patient = await self.patient_repo.get_by_id(data.patient_id)
            if not patient:
                raise NotFoundException("Patient not found")

        if data.prescription_id is not None:
            prescription = await self.prescription_repo.get_by_id(data.prescription_id)
            if not prescription:
                raise NotFoundException("Prescription not found")
            if prescription.patient_id != data.patient_id:
                raise BadRequestException("Prescription does not belong to this patient")

        subtotal = 0.0
        invoice_items: list[PharmacyInvoiceItem] = []
        for item_data in data.items:
            medicine = await self.medicine_repo.get_by_id_for_update(item_data.medicine_id)
            if not medicine:
                raise NotFoundException(f"Medicine {item_data.medicine_id} not found")
            if medicine.stock_quantity < item_data.quantity:
                raise BadRequestException(f"Insufficient stock for {medicine.name}")
            if medicine.expiry_date and medicine.expiry_date < date.today():
                raise BadRequestException(f"Medicine {medicine.name} has expired")
            unit_price = item_data.unit_price if item_data.unit_price is not None else medicine.unit_price
            line_total = round(item_data.quantity * unit_price, 2)
            subtotal += line_total
            invoice_items.append(PharmacyInvoiceItem(
                medicine_id=item_data.medicine_id,
                quantity=item_data.quantity,
                unit_price=unit_price,
                line_total=line_total,
            ))
            await self.medicine_repo.update_stock(item_data.medicine_id, -item_data.quantity)

        from app.models.user_model import User
        from app.services.settings_service import SettingsService
        from sqlalchemy import select
        from app.utils.helpers import generate_code

        user_record = await self.db.scalar(select(User).where(User.id == user_id))
        hospital_id = user_record.hospital_id if user_record and user_record.hospital_id else 1
        billing_settings = await SettingsService(self.db).get_billing_settings(hospital_id)

        payment_mode_val = data.payment_mode or billing_settings.get("default_payment_mode", "Cash")
        tax_percentage_val = data.tax_percentage if data.tax_percentage else billing_settings.get("gst_percentage", 0.0)

        discount_amount = round((subtotal * data.discount_percentage) / 100, 2)
        tax_amount = round((subtotal - discount_amount) * tax_percentage_val / 100, 2)
        gst_amount = tax_amount
        total = round(subtotal - discount_amount + tax_amount, 2)
        invoice = PharmacyInvoice(
            invoice_number=generate_code(billing_settings.get("receipt_prefix", "PHR")),
            patient_id=data.patient_id,
            prescription_id=data.prescription_id,
            payment_mode=payment_mode_val,
            subtotal=subtotal,
            discount_percentage=data.discount_percentage,
            discount_amount=discount_amount,
            tax_percentage=tax_percentage_val,
            tax_amount=tax_amount,
            gst_amount=gst_amount,
            total_amount=total,
            paid_amount=total,
            status="paid",
            created_by=user_id,
        )
        invoice = await self.invoice_repo.create(invoice, invoice_items)
        if data.prescription_id is not None:
            prescription.status = "completed"
            prescription.dispensed_at = utc_now()
            await self.prescription_repo.update(prescription)
        await self.audit_repo.create("create", "pharmacy_invoice", user_id=user_id, resource_id=str(invoice.id))

        from app.services.transaction_history_service import TransactionHistoryService
        tx_service = TransactionHistoryService(self.db)
        await tx_service.create_event(
            event_type="INVOICE_CREATED",
            reference_no=invoice.invoice_number,
            description=f"Pharmacy Invoice Created: {invoice.invoice_number}",
            amount=invoice.total_amount,
            source_module="pharmacy_billing",
            source_id=invoice.id,
            status="completed",
            user_id=user_id
        )
        await tx_service.create_event(
            event_type="PAYMENT_RECEIVED",
            reference_no=invoice.invoice_number,
            description=f"Pharmacy Invoice Paid: {invoice.invoice_number}",
            amount=invoice.total_amount,
            source_module="pharmacy_billing",
            source_id=invoice.id,
            status="completed",
            user_id=user_id
        )

        return self._invoice_response(invoice)

    async def list_invoices(self, page: int = 1, size: int = 20):
        skip = (page - 1) * size
        items = await self.invoice_repo.list_all(skip=skip, limit=size)
        total = await self.invoice_repo.count_all()
        return build_paginated_result([self._invoice_response(i) for i in items], total, page, size)

    def _invoice_response(self, invoice: PharmacyInvoice) -> PharmacyInvoiceResponse:
        resp = PharmacyInvoiceResponse.model_validate(invoice)
        resp.items = [PharmacyInvoiceItemResponse.model_validate(i) for i in invoice.items]
        return resp

    async def get_invoice_by_id(self, invoice_id: int) -> PharmacyInvoiceResponse:
        invoice = await self.invoice_repo.get_by_id(invoice_id)
        if not invoice:
            raise NotFoundException("Pharmacy invoice not found")
        return self._invoice_response(invoice)

    async def update_invoice(self, invoice_id: int, data: PharmacyInvoiceUpdate) -> PharmacyInvoiceResponse:
        invoice = await self.invoice_repo.get_by_id(invoice_id)
        if not invoice:
            raise NotFoundException("Pharmacy invoice not found")

        if data.payment_mode is not None:
            invoice.payment_mode = data.payment_mode
        if data.status is not None:
            invoice.status = data.status
        if data.discount_percentage is not None:
            invoice.discount_percentage = data.discount_percentage
        if data.tax_percentage is not None:
            invoice.tax_percentage = data.tax_percentage

        discount_amount = round((invoice.subtotal * invoice.discount_percentage) / 100, 2)
        tax_amount = round((invoice.subtotal - discount_amount) * invoice.tax_percentage / 100, 2)

        invoice.discount_amount = discount_amount
        invoice.tax_amount = tax_amount
        invoice.gst_amount = tax_amount
        invoice.total_amount = round(invoice.subtotal - discount_amount + tax_amount, 2)
        invoice.paid_amount = invoice.total_amount

        invoice = await self.invoice_repo.update(invoice)

        # Update transaction history
        from app.models.transaction_history_model import TransactionHistory
        from sqlalchemy import select
        tx_result = await self.db.execute(
            select(TransactionHistory).where(
                TransactionHistory.source_module == "pharmacy_billing",
                TransactionHistory.source_id == invoice.id,
                TransactionHistory.is_deleted == False
            )
        )
        tx_records = tx_result.scalars().all()
        for tx in tx_records:
            tx.amount = invoice.total_amount
            await self.db.flush()

        return self._invoice_response(invoice)

    async def delete_invoice(self, invoice_id: int) -> None:
        invoice = await self.invoice_repo.get_by_id(invoice_id)
        if not invoice:
            raise NotFoundException("Pharmacy invoice not found")
        await self.invoice_repo.soft_delete(invoice)

        # Soft delete transaction history
        from app.models.transaction_history_model import TransactionHistory
        from sqlalchemy import select
        tx_result = await self.db.execute(
            select(TransactionHistory).where(
                TransactionHistory.source_module == "pharmacy_billing",
                TransactionHistory.source_id == invoice.id,
                TransactionHistory.is_deleted == False
            )
        )
        tx_records = tx_result.scalars().all()
        for tx in tx_records:
            tx.is_deleted = True
            tx.deleted_at = utc_now()
            await self.db.flush()

    async def download_invoice(self, invoice_id: int):
        from fastapi.responses import Response
        from sqlalchemy import select
        from sqlalchemy.orm import selectinload
        from app.models.pharmacy_model import PharmacyInvoice, PharmacyInvoiceItem
        from app.utils.pdf_generator import html_to_pdf

        stmt = (
            select(PharmacyInvoice)
            .where(
                PharmacyInvoice.id == invoice_id,
                PharmacyInvoice.is_deleted.is_(False),
            )
            .options(
                selectinload(PharmacyInvoice.items).selectinload(PharmacyInvoiceItem.medicine)
            )
        )
        res = await self.db.execute(stmt)
        invoice = res.scalar_one_or_none()

        if not invoice:
            raise NotFoundException("Pharmacy invoice not found")

        patient_name = "Walk-in Patient"
        patient_phone = "-"
        if invoice.patient_id:
            patient = await self.patient_repo.get_by_id(invoice.patient_id)
            if patient:
                patient_name = f"{patient.first_name} {patient.last_name}".strip()
                patient_phone = patient.phone or "-"

        item_rows = ""
        for idx, item in enumerate(invoice.items, start=1):
            med_name = item.medicine.name if (item.medicine and item.medicine.name) else f"Medicine #{item.medicine_id}"
            unit_price = float(item.unit_price or 0.0)
            line_total = float(item.line_total or 0.0)
            item_rows += f"""
            <tr>
                <td style="padding: 8px; border: 1px solid #ddd; text-align: center;">{idx}</td>
                <td style="padding: 8px; border: 1px solid #ddd;">{med_name}</td>
                <td style="padding: 8px; border: 1px solid #ddd; text-align: center;">{item.quantity}</td>
                <td style="padding: 8px; border: 1px solid #ddd; text-align: right;">{unit_price:.2f}</td>
                <td style="padding: 8px; border: 1px solid #ddd; text-align: right;">{line_total:.2f}</td>
            </tr>
            """

        payment_mode_str = invoice.payment_mode or "Cash"
        created_str = invoice.created_at.strftime("%Y-%m-%d %H:%M") if invoice.created_at else "-"

        subtotal = float(invoice.subtotal or 0.0)
        discount_amount = float(invoice.discount_amount or 0.0)
        tax_amount = float(invoice.tax_amount or 0.0)
        total_amount = float(invoice.total_amount or 0.0)

        html_content = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>Pharmacy Invoice #{invoice.invoice_number}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; color: #333; }}
        h1 {{ color: #2563eb; margin-bottom: 5px; }}
        .header {{ width: 100%; border-bottom: 2px solid #2563eb; padding-bottom: 10px; margin-bottom: 20px; }}
        .meta-table {{ width: 100%; margin-bottom: 20px; background: #f8fafc; border: 1px solid #e2e8f0; padding: 10px; }}
        .items-table {{ width: 100%; border-collapse: collapse; margin-bottom: 20px; }}
        .items-table th {{ background: #2563eb; color: #ffffff; padding: 8px; border: 1px solid #2563eb; text-align: left; }}
        .totals-table {{ width: 300px; float: right; border-collapse: collapse; margin-top: 10px; }}
        .totals-table td {{ padding: 6px; }}
        .grand-total {{ font-weight: bold; font-size: 16px; color: #2563eb; border-top: 2px solid #333; border-bottom: 2px solid #333; }}
    </style>
</head>
<body>
    <div class="header">
        <h1>NexaCare Pharmacy</h1>
        <p style="margin:0;">Advanced Healthcare & Pharmacy Billing</p>
        <h3 style="margin-top:10px;">INVOICE #{invoice.invoice_number}</h3>
        <p style="margin:0;">Date: {created_str}</p>
    </div>

    <table class="meta-table">
        <tr>
            <td style="vertical-align: top;">
                <strong>Patient Name:</strong> {patient_name}<br>
                <strong>Phone:</strong> {patient_phone}
            </td>
            <td style="vertical-align: top; text-align: right;">
                <strong>Status:</strong> {invoice.status}<br>
                <strong>Payment Mode:</strong> {payment_mode_str}
            </td>
        </tr>
    </table>

    <table class="items-table">
        <thead>
            <tr>
                <th style="text-align: center; width: 40px;">#</th>
                <th>Medicine Name</th>
                <th style="text-align: center; width: 60px;">Qty</th>
                <th style="text-align: right; width: 100px;">Unit Price</th>
                <th style="text-align: right; width: 100px;">Line Total</th>
            </tr>
        </thead>
        <tbody>
            {item_rows}
        </tbody>
    </table>

    <table class="totals-table">
        <tr>
            <td>Subtotal:</td>
            <td style="text-align: right;">{subtotal:.2f}</td>
        </tr>
        <tr>
            <td>Discount ({invoice.discount_percentage or 0}%):</td>
            <td style="text-align: right;">-{discount_amount:.2f}</td>
        </tr>
        <tr>
            <td>GST / Tax ({invoice.tax_percentage or 0}%):</td>
            <td style="text-align: right;">+{tax_amount:.2f}</td>
        </tr>
        <tr class="grand-total">
            <td>Total Amount:</td>
            <td style="text-align: right;">{total_amount:.2f}</td>
        </tr>
    </table>

    <div style="clear: both; padding-top: 40px; text-align: center; color: #777;">
        <p>Thank you for choosing NexaCare Pharmacy! Wish you good health.</p>
    </div>
</body>
</html>"""

        try:
            pdf_bytes = html_to_pdf(html_content)
            return Response(
                content=pdf_bytes,
                media_type="application/pdf",
                headers={
                    "Content-Disposition": f"attachment; filename=pharmacy_invoice_{invoice.invoice_number}.pdf"
                },
            )
        except Exception as exc:
            import logging
            logging.getLogger(__name__).warning("PDF conversion failed, returning HTML fallback: %s", exc)
            from fastapi.responses import HTMLResponse
            return HTMLResponse(content=html_content)


    # --- Suppliers ---
    async def create_supplier(self, data: SupplierCreate, user_id: int) -> SupplierResponse:
        if data.phone:
            existing_phone = await self.supplier_repo.get_by_phone(data.phone)
            if existing_phone:
                raise ConflictException("Supplier with this phone number already exists")

        if data.email:
            existing_email = await self.supplier_repo.get_by_email(data.email)
            if existing_email:
                raise ConflictException("Supplier with this email already exists")

        if data.gst_number:
            existing_gst = await self.supplier_repo.get_by_gst(data.gst_number)
            if existing_gst:
                raise ConflictException("Supplier with this GST number already exists")

        supplier = Supplier(**data.model_dump())
        supplier = await self.supplier_repo.create(supplier)
        await self.audit_repo.create("create", "pharmacy_supplier", user_id=user_id, resource_id=str(supplier.id))
        return SupplierResponse.model_validate(supplier)

    async def list_suppliers(self, page: int = 1, size: int = 20):
        skip = (page - 1) * size
        items = await self.supplier_repo.list_all(skip=skip, limit=size)
        total = await self.supplier_repo.count_all()
        return build_paginated_result([SupplierResponse.model_validate(s) for s in items], total, page, size)


    async def get_supplier(self, supplier_id: int) -> SupplierResponse:
        supplier = await self.supplier_repo.get_by_id(supplier_id)

        if not supplier:
            raise NotFoundException("Supplier not found")

        return SupplierResponse.model_validate(supplier)

    async def update_supplier(self, supplier_id: int, data: SupplierUpdate, user_id: int) -> SupplierResponse:
        supplier = await self.supplier_repo.get_by_id(supplier_id)

        if not supplier:
            raise NotFoundException("Supplier not found")

        for key, value in data.model_dump(exclude_unset=True).items():
            setattr(supplier, key, value)

        supplier = await self.supplier_repo.update(supplier)
        await self.audit_repo.create(
            "update",
            "pharmacy_supplier",
             user_id=user_id,
             resource_id=str(supplier.id),
    )

        return SupplierResponse.model_validate(supplier)

    async def delete_supplier(self, supplier_id: int, user_id: int) -> None:
        supplier = await self.supplier_repo.get_by_id(supplier_id)

        if not supplier:
            raise NotFoundException("Supplier not found")

        await self.supplier_repo.soft_delete(supplier)
        await self.audit_repo.create(
            "delete",
            "pharmacy_supplier",
            user_id=user_id,
            resource_id=str(supplier.id),
        )
    # --- Purchases ---
    async def create_purchase(self, data: PurchaseCreate, user_id: int) -> PurchaseResponse:
        # Validate Supplier
        supplier = await self.supplier_repo.get_by_id(data.supplier_id)
        if not supplier:
            raise NotFoundException(f"Supplier with ID {data.supplier_id} not found")

        total = 0.0
        purchase_items: list[PurchaseItem] = []
        for item_data in data.items:
            # Validate Medicine
            medicine = await self.medicine_repo.get_by_id(item_data.medicine_id)
            if not medicine:
                raise NotFoundException(f"Medicine with ID {item_data.medicine_id} not found")

            line_total = round(item_data.quantity * item_data.unit_price, 2)
            total += line_total
            purchase_items.append(PurchaseItem(
                medicine_id=item_data.medicine_id,
                quantity=item_data.quantity,
                unit_price=item_data.unit_price,
                expiry_date=item_data.expiry_date,
                line_total=line_total,
            ))
        purchase = Purchase(
            purchase_number=generate_purchase_number(),
            supplier_id=data.supplier_id,
            total_amount=total,
            ordered_at=utc_now(),
            notes=data.notes,
            status=data.status or "Pending",
            created_by=user_id,
        )
        purchase = await self.purchase_repo.create(purchase, purchase_items)

        purchase = await self.purchase_repo.get_by_id(purchase.id)
        await self.audit_repo.create("create", "pharmacy_purchase", user_id=user_id, resource_id=str(purchase.id))

        from app.services.transaction_history_service import TransactionHistoryService
        await TransactionHistoryService(self.db).create_event(
            event_type="EXPENSE_RECORDED",
            reference_no=purchase.purchase_number,
            description=f"Pharmacy Purchase: {purchase.purchase_number}",
            amount=purchase.total_amount,
            source_module="pharmacy_purchases",
            source_id=purchase.id,
            status="completed",
            user_id=user_id
        )

        return self._purchase_response(purchase)

    async def list_purchases(self, page: int = 1, size: int = 20):
        skip = (page - 1) * size
        items = await self.purchase_repo.list_all(skip=skip, limit=size)
        total = await self.purchase_repo.count_all()
        return build_paginated_result([self._purchase_response(p) for p in items], total, page, size)

    def _purchase_response(self, purchase: Purchase) -> PurchaseResponse:
        items_resp = []
        for i in purchase.items:
            items_resp.append(PurchaseItemResponse(
                id=i.id,
                purchase_id=i.purchase_id,
                medicine_id=i.medicine_id,
                quantity=i.quantity or 0,
                unit_price=i.unit_price or 0.0,
                expiry_date=i.expiry_date,
                line_total=i.line_total or 0.0,
                created_at=i.created_at or utc_now(),
                updated_at=i.updated_at or utc_now()
            ))

        resp = PurchaseResponse(
            id=purchase.id,
            purchase_number=purchase.purchase_number or "",
            supplier_id=purchase.supplier_id,
            total_amount=purchase.total_amount or 0.0,
            status=purchase.status or "Pending",
            ordered_at=purchase.ordered_at or purchase.created_at or utc_now(),
            received_at=purchase.received_at,
            notes=purchase.notes,
            created_by=purchase.created_by,
            received_by=purchase.received_by,
            items=items_resp,
            created_at=purchase.created_at or utc_now()
        )
        return resp

    async def get_purchase(self, purchase_id: int) -> PurchaseResponse:
        purchase = await self.purchase_repo.get_by_id(purchase_id)

        if not purchase:
            raise NotFoundException("Purchase not found")

        return self._purchase_response(purchase)


    async def update_purchase(
        self,
        purchase_id: int,
        data: PurchaseCreate,
        user_id: int,
    ) -> PurchaseResponse:
        purchase = await self.purchase_repo.get_by_id(purchase_id)

        if not purchase:
            raise NotFoundException("Purchase not found")

        purchase.supplier_id = data.supplier_id
        purchase.notes = data.notes
        if data.status:
            allowed_statuses = {PurchaseStatus.ORDERED, PurchaseStatus.RECEIVED, PurchaseStatus.CANCELLED}
            if data.status not in allowed_statuses:
                raise BadRequestException(
                    f"Invalid purchase status. Must be one of: {', '.join(allowed_statuses)}"
                )
            purchase.status = data.status

        purchase = await self.purchase_repo.update(purchase)

        await self.audit_repo.create(
            "update",
            "pharmacy_purchase",
            user_id=user_id,
            resource_id=str(purchase.id),
        )

        # Update transaction history
        from app.models.transaction_history_model import TransactionHistory
        from sqlalchemy import select
        tx_result = await self.db.execute(
            select(TransactionHistory).where(
                TransactionHistory.source_module == "pharmacy_purchases",
                TransactionHistory.source_id == purchase.id,
                TransactionHistory.is_deleted == False
            )
        )
        tx_records = tx_result.scalars().all()
        for tx in tx_records:
            tx.amount = purchase.total_amount
            tx.description = f"Pharmacy Purchase: {purchase.purchase_number}"
            await self.db.flush()

        purchase = await self.purchase_repo.get_by_id(purchase.id)
        return self._purchase_response(purchase)


    async def delete_purchase(self, purchase_id: int, user_id: int) -> None:
        purchase = await self.purchase_repo.get_by_id(purchase_id)

        if not purchase:
            raise NotFoundException("Purchase not found")

        await self.purchase_repo.soft_delete(purchase)

        await self.audit_repo.create(
            "delete",
            "pharmacy_purchase",
            user_id=user_id,
            resource_id=str(purchase.id),
        )

        # Soft delete transaction history
        from app.models.transaction_history_model import TransactionHistory
        from sqlalchemy import select
        tx_result = await self.db.execute(
            select(TransactionHistory).where(
                TransactionHistory.source_module == "pharmacy_purchases",
                TransactionHistory.source_id == purchase.id,
                TransactionHistory.is_deleted == False
            )
        )
        tx_records = tx_result.scalars().all()
        for tx in tx_records:
            tx.is_deleted = True
            tx.deleted_at = utc_now()
            await self.db.flush()

    async def receive_purchase_order(
        self,
        purchase_order_id: int,
        current_user,
    ) -> PurchaseResponse:
        purchase = await self.purchase_repo.get_by_id(purchase_order_id)

        if not purchase:
            raise NotFoundException("Purchase not found")

        if getattr(purchase, "is_deleted", False):
            raise NotFoundException("Purchase not found")

        current_status = purchase.status.lower() if purchase.status else ""
        if current_status == "received":
            raise BadRequestException("Purchase Order already received")

        allowed_statuses = {"ordered", "pending", "partially_received"}
        if current_status not in allowed_statuses:
            raise BadRequestException("Invalid Purchase Order status")

        purchase.status = "received"
        purchase.received_at = utc_now()
        purchase.received_by = current_user.id

        for item in purchase.items:
            await self.medicine_repo.update_stock(item.medicine_id, item.quantity)

        await self.purchase_repo.update(purchase)

        await self.audit_repo.create(
            "receive",
            "pharmacy_purchase",
            user_id=current_user.id,
            resource_id=str(purchase.id),
        )

        await self.db.flush()
        return self._purchase_response(purchase)

    async def get_sales_report(self, period: str = "all") -> SalesReport:
        from datetime import timedelta
        from app.core.exceptions import BadRequestException

        if not period or period.strip() == "":
            period = "all"

        valid_periods = {"daily", "weekly", "monthly", "yearly", "all", "overall"}
        if period not in valid_periods:
            raise BadRequestException(
                f"Invalid period parameter. Allowed values: {', '.join(sorted(valid_periods))}"
            )

        now = utc_now()
        start = None
        if period == "daily":
            start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        elif period == "weekly":
            start = now - timedelta(days=7)
        elif period == "monthly":
            start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        elif period == "yearly":
            start = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        elif period in ("all", "overall"):
            start = None

        data = await self.invoice_repo.get_sales_report(start, now)
        return SalesReport(period=period, **data)

    async def get_dashboard_summary(
        self,
        time_filter: str = "7_days",
        start_date: date | None = None,
        end_date: date | None = None,
    ) -> PharmacyDashboardResponse:
        start_dt, end_dt = self.get_date_range(time_filter, start_date, end_date)
        return await self.get_dashboard_overview(time_filter=time_filter, start_dt=start_dt, end_dt=end_dt)

    async def get_dashboard_overview(
        self,
        time_filter: str = "7_days",
        start_dt: Optional[datetime] = None,
        end_dt: Optional[datetime] = None,
    ) -> PharmacyDashboardResponse:
        from datetime import timezone, timedelta, time, date as dt_date
        from sqlalchemy import select, func, or_, cast, Date
        from app.utils.helpers import get_today_ist

        today_ist = get_today_ist()

        # 1. Total Medicines (active, i.e., is_deleted=False and is_active=True)
        total_medicines = (await self.db.scalar(
            select(func.count(Medicine.id)).where(
                Medicine.is_deleted.is_(False),
                Medicine.is_active.is_(True)
            )
        )) or 0

        # 2. Low Stock Alerts (Medicine.stock_quantity <= Medicine.reorder_level)
        low_stock_alerts = (await self.db.scalar(
            select(func.count(Medicine.id)).where(
                Medicine.is_deleted.is_(False),
                Medicine.is_active.is_(True),
                Medicine.stock_quantity <= Medicine.reorder_level
            )
        )) or 0

        # 3. Expired Alerts (Near expiry and expired: Medicine.expiry_date <= today_ist + 30 days, stock_quantity > 0)
        threshold_date = today_ist + timedelta(days=30)
        expired_alerts = (await self.db.scalar(
            select(func.count(Medicine.id)).where(
                Medicine.is_deleted.is_(False),
                Medicine.is_active.is_(True),
                Medicine.stock_quantity > 0,
                Medicine.expiry_date.isnot(None),
                Medicine.expiry_date <= threshold_date
            )
        )) or 0

        # Expired medicines count (strictly expired: Medicine.expiry_date < today_ist, stock_quantity > 0)
        expired_medicines_alerts = (await self.db.scalar(
            select(func.count(Medicine.id)).where(
                Medicine.is_deleted.is_(False),
                Medicine.is_active.is_(True),
                Medicine.stock_quantity > 0,
                Medicine.expiry_date.isnot(None),
                Medicine.expiry_date < today_ist
            )
        )) or 0

        # 4. Today Sales (Invoice/billing amount created today, OR in the filtered period)
        today_sales_query = select(func.coalesce(func.sum(PharmacyInvoice.total_amount), 0.0)).where(
            PharmacyInvoice.is_deleted.is_(False),
            PharmacyInvoice.status != "cancelled",
        )
        if start_dt:
            today_sales_query = today_sales_query.where(PharmacyInvoice.created_at >= start_dt)
        if end_dt:
            today_sales_query = today_sales_query.where(PharmacyInvoice.created_at <= end_dt)
            
        today_sales = (await self.db.scalar(today_sales_query)) or 0.0

        # 5. Monthly Sales (Invoice/billing amount for current month, OR in the filtered period)
        month_start = datetime.combine(today_ist.replace(day=1), time.min)
        if month_start.month == 12:
            next_month_start = month_start.replace(year=month_start.year + 1, month=1)
        else:
            next_month_start = month_start.replace(month=month_start.month + 1)
            
        monthly_sales_query = select(func.coalesce(func.sum(PharmacyInvoice.total_amount), 0.0)).where(
            PharmacyInvoice.is_deleted.is_(False),
            PharmacyInvoice.status != "cancelled",
        )
        if time_filter in ("overall", "30_days", "7_days", "3_month", "custom"):
            if start_dt:
                monthly_sales_query = monthly_sales_query.where(PharmacyInvoice.created_at >= start_dt)
            if end_dt:
                monthly_sales_query = monthly_sales_query.where(PharmacyInvoice.created_at <= end_dt)
        else:
            monthly_sales_query = monthly_sales_query.where(
                PharmacyInvoice.created_at >= month_start,
                PharmacyInvoice.created_at < next_month_start
            )
            
        monthly_sales = (await self.db.scalar(monthly_sales_query)) or 0.0

        # 6. Pending Purchases (Count status: Pending, Ordered)
        pending_purchases = (await self.db.scalar(
            select(func.count(Purchase.id)).where(
                Purchase.status.in_(["Pending", "Ordered"])
            )
        )) or 0

        # 7. Total Suppliers (Count active suppliers)
        total_suppliers = (await self.db.scalar(
            select(func.count(Supplier.id)).where(
                Supplier.is_deleted.is_(False)
            )
        )) or 0

        # 8. Prescriptions (Count active prescriptions pending to be dispensed)
        prescriptions = (await self.db.scalar(
            select(func.count(Prescription.id)).where(
                Prescription.is_deleted.is_(False),
                Prescription.status == "pending"
            )
        )) or 0

        # 9. Low Stock Items (Max 10 medicines ordered by stock ascending)
        low_stock_query = (
            select(Medicine)
            .where(
                Medicine.is_deleted.is_(False),
                Medicine.is_active.is_(True),
                Medicine.stock_quantity <= Medicine.reorder_level
            )
            .order_by(Medicine.stock_quantity.asc())
            .limit(10)
        )
        low_stock_res = await self.db.execute(low_stock_query)
        low_stock_items = [
            {
                "medicine_id": m.id,
                "medicine_name": m.name,
                "current_stock": m.stock_quantity,
                "minimum_stock": m.reorder_level,
                # Backward compatibility fields
                "id": m.id,
                "name": m.name,
                "stock_quantity": m.stock_quantity,
                "reorder_level": m.reorder_level,
                "unit": m.unit or "Unit",
                "status_label": f"{m.stock_quantity} Left" if m.stock_quantity > 0 else "Out of Stock"
            }
            for m in low_stock_res.scalars().all()
        ]

        # 10. Today Sales Trend (Hourly sales for today)
        today_start = datetime.combine(today_ist, time.min)
        tomorrow_start = today_start + timedelta(days=1)
        today_trend_query = (
            select(
                func.extract('hour', PharmacyInvoice.created_at).label("hr"),
                func.sum(PharmacyInvoice.total_amount).label("amt")
            )
            .where(
                PharmacyInvoice.is_deleted.is_(False),
                PharmacyInvoice.status != "cancelled",
                PharmacyInvoice.created_at >= today_start,
                PharmacyInvoice.created_at < tomorrow_start
            )
            .group_by(func.extract('hour', PharmacyInvoice.created_at))
            .order_by(func.extract('hour', PharmacyInvoice.created_at).asc())
        )
        today_trend_res = await self.db.execute(today_trend_query)
        today_sales_trend = []
        for row in today_trend_res.all():
            hr_val = 0
            if row.hr is not None:
                try:
                    hr_val = int(row.hr)
                except (ValueError, TypeError):
                    pass
            today_sales_trend.append({
                "hour": f"{hr_val:02d}",
                "amount": float(row.amt or 0.0),
                "label": f"{hr_val:02d}:00"
            })

        # 11. Monthly Sales Trend (Daily sales for selected period/current month)
        monthly_trend_query = (
            select(
                cast(PharmacyInvoice.created_at, Date).label("dt"),
                func.sum(PharmacyInvoice.total_amount).label("amt")
            )
            .where(
                PharmacyInvoice.is_deleted.is_(False),
                PharmacyInvoice.status != "cancelled",
            )
        )
        if start_dt:
            monthly_trend_query = monthly_trend_query.where(PharmacyInvoice.created_at >= start_dt)
        if end_dt:
            monthly_trend_query = monthly_trend_query.where(PharmacyInvoice.created_at <= end_dt)
        else:
            # Fallback to current month if no dates (e.g. if overall is somehow not returning None)
            pass
            
        monthly_trend_query = (
            monthly_trend_query
            .group_by(cast(PharmacyInvoice.created_at, Date))
            .order_by(cast(PharmacyInvoice.created_at, Date).asc())
        )
        monthly_trend_res = await self.db.execute(monthly_trend_query)
        monthly_sales_trend = []
        for row in monthly_trend_res.all():
            dt_str = str(row.dt) if row.dt is not None else ""
            monthly_sales_trend.append({
                "date": dt_str,
                "amount": float(row.amt or 0.0),
                "label": dt_str
            })

        today_sales = round(float(today_sales or 0.0), 2)
        monthly_sales = round(float(monthly_sales or 0.0), 2)
        status_mix_raw = await self.dashboard_repo.get_inventory_status_mix(reference_date=today_ist)
        inventory_status_mix = InventoryStatusMix(**status_mix_raw)

        total_mix = (
            inventory_status_mix.expiring_soon
            + inventory_status_mix.in_stock
            + inventory_status_mix.low_stock
            + inventory_status_mix.out_of_stock
        )
        if total_mix > 0:
            in_stock_pct = round((inventory_status_mix.in_stock / total_mix) * 100, 2)
            low_stock_pct = round((inventory_status_mix.low_stock / total_mix) * 100, 2)
            out_of_stock_pct = round((inventory_status_mix.out_of_stock / total_mix) * 100, 2)
            expiring_soon_pct = round((inventory_status_mix.expiring_soon / total_mix) * 100, 2)
        else:
            in_stock_pct = 0.0
            low_stock_pct = 0.0
            out_of_stock_pct = 0.0
            expiring_soon_pct = 0.0

        inventory_health_progress = InventoryHealthProgress(
            in_stock=in_stock_pct,
            low_stock=low_stock_pct,
            out_of_stock=out_of_stock_pct,
            expiring_soon=expiring_soon_pct,
        )

        return PharmacyDashboardResponse(
            total_medicines=total_medicines,
            low_stock_alerts=low_stock_alerts,
            expired_alerts=expired_alerts,
            today_sales=today_sales,
            monthly_sales=monthly_sales,
            pending_purchases=pending_purchases,
            total_suppliers=total_suppliers,
            prescriptions=prescriptions,
            prescriptions_count=prescriptions,
            expired_medicines_alerts=expired_medicines_alerts,
            daily_sales=today_sales,
            low_stock_items=low_stock_items,
            today_sales_trend=today_sales_trend,
            monthly_sales_trend=monthly_sales_trend,
            inventory_status_mix=inventory_status_mix,
            inventory_health_progress=inventory_health_progress,
        )

    async def get_inventory_overview(self) -> PharmacyInventoryOverviewResponse:
        from app.utils.helpers import get_today_ist
        today = get_today_ist()
        counts = await self.medicine_repo.get_inventory_counts(reference_date=today)
        daily_deductions = await self.invoice_repo.get_daily_stock_deductions()
        most_selling = await self.invoice_repo.get_most_selling_medicines()
        date_wise = await self.invoice_repo.get_date_wise_medicines()

        return PharmacyInventoryOverviewResponse(
            **counts,
            daily_stock_deductions=daily_deductions,
            most_selling_medicines=most_selling,
            date_wise_medicines=date_wise
        )

    async def generate_medicine_bulk_template(self) -> BytesIO:
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Medicines Template"

        headers = [
            "name", "generic_name", "category", "barcode", "batch_number",
            "expiry_date", "manufacturer", "unit", "unit_price",
            "stock_quantity", "reorder_level", "description"
        ]
        ws.append(headers)

        # Add a valid sample medicine row
        sample_row = [
            "Paracetamol", "Acetaminophen", "Analgesics", "8901234567890", "BATCH-001",
            "2028-12-31", "Nexa Pharma", "Tablets", 5.5, 100, 10,
            "Take one tablet as directed by physician"
        ]
        ws.append(sample_row)

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream

    async def import_medicines_from_excel(self, file, user_id: int) -> dict:
        from io import BytesIO
        import openpyxl
        from datetime import date, datetime
        from pydantic import ValidationError
        from app.schemas.pharmacy_schema import MedicineCreate
        from app.models.pharmacy_model import Medicine
        from app.core.exceptions import ConflictException

        content = await file.read()
        try:
            wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
            ws = wb.active
        except Exception as e:
            raise BadRequestException(f"Invalid Excel file format: {str(e)}")

        if not ws:
            raise BadRequestException("The uploaded Excel workbook contains no active worksheet")

        # Extract headers from the first row
        headers = [str(cell.value).strip().lower() for cell in ws[1] if cell.value is not None]
        required_headers = {"name", "category", "unit"}
        missing_headers = required_headers - set(headers)
        if missing_headers:
            raise BadRequestException(f"Missing required columns in Excel: {', '.join(missing_headers)}")

        total_rows = 0
        created = 0
        failed = 0
        errors = []

        # We will track barcodes processed in this batch to prevent duplicate barcode insert within the same file
        batch_barcodes = set()

        # Iterate through rows starting from row 2
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Check if the row is completely empty/None
            if all(cell is None for cell in row):
                continue

            total_rows += 1
            row_dict = {}
            for header, val in zip(headers, row):
                # normalize empty/None values
                if val is None or str(val).strip() == "":
                    row_dict[header] = None
                else:
                    row_dict[header] = val

            try:
                # Pre-processing/normalization for the schema:
                # 1. Barcode: handle float numeric representations safely (e.g. 1234567890123.0 -> "1234567890123")
                barcode_raw = row_dict.get("barcode")
                if barcode_raw is not None:
                    if isinstance(barcode_raw, float):
                        # Convert float to int string without adding padding/zeros or removing digits
                        barcode_raw = str(int(barcode_raw))
                    else:
                        barcode_raw = str(barcode_raw).strip()
                    row_dict["barcode"] = barcode_raw

                # 2. Expiry date: handle datetime.datetime or datetime.date objects from Excel cells
                expiry_raw = row_dict.get("expiry_date")
                if expiry_raw is not None:
                    if isinstance(expiry_raw, (datetime, date)):
                        row_dict["expiry_date"] = expiry_raw.strftime("%Y-%m-%d")
                    else:
                        row_dict["expiry_date"] = str(expiry_raw).strip()

                # 3. Numeric values: unit_price, stock_quantity, reorder_level
                price_raw = row_dict.get("unit_price")
                if price_raw is not None:
                    try:
                        row_dict["unit_price"] = float(price_raw)
                    except ValueError:
                        pass

                qty_raw = row_dict.get("stock_quantity")
                if qty_raw is not None:
                    try:
                        row_dict["stock_quantity"] = int(float(qty_raw))
                    except ValueError:
                        pass

                reorder_raw = row_dict.get("reorder_level")
                if reorder_raw is not None:
                    try:
                        row_dict["reorder_level"] = int(float(reorder_raw))
                    except ValueError:
                        pass

                # Validate using MedicineCreate schema
                validated_data = MedicineCreate(**row_dict)

                # Check duplicate barcode in DB and current batch
                if validated_data.barcode:
                    if validated_data.barcode in batch_barcodes:
                        raise ConflictException("Duplicate barcode in the uploaded file")

                    existing = await self.medicine_repo.get_by_barcode(validated_data.barcode)
                    if existing:
                        raise ConflictException("Medicine with this barcode already exists")

                    batch_barcodes.add(validated_data.barcode)

                # Create the medicine
                medicine = Medicine(sku=generate_medicine_sku(), **validated_data.model_dump())
                medicine = await self.medicine_repo.create(medicine)
                await self.audit_repo.create("create", "pharmacy", user_id=user_id, resource_id=str(medicine.id))
                created += 1

            except ValidationError as e:
                failed += 1
                err_msg = "; ".join([f"{'.'.join(str(loc) for loc in error['loc'])}: {error['msg']}" for error in e.errors()])
                errors.append({"row": row_idx, "error": err_msg})
            except ConflictException as e:
                failed += 1
                errors.append({"row": row_idx, "error": str(e.detail)})
            except Exception as e:
                failed += 1
                errors.append({"row": row_idx, "error": str(e)})

        await self.db.flush()

        return {
            "total_rows": total_rows,
            "created": created,
            "failed": failed,
            "errors": errors,
        }

    async def export_medicines(self, format_type: str) -> tuple[BytesIO | bytes, str]:
        from io import BytesIO
        from app.utils.helpers import utc_now

        medicines = await self.medicine_repo.get_all_active()

        if format_type == "excel":
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Medicines Inventory"

            headers = [
                "sku", "name", "generic_name", "category", "barcode", "batch_number",
                "expiry_date", "manufacturer", "unit", "unit_price",
                "stock_quantity", "reorder_level", "description"
            ]
            ws.append(headers)

            for item in medicines:
                row = [
                    item.sku,
                    item.name,
                    item.generic_name or "",
                    item.category,
                    item.barcode or "",
                    item.batch_number or "",
                    item.expiry_date.strftime("%Y-%m-%d") if isinstance(item.expiry_date, (date, datetime)) else (item.expiry_date or ""),
                    item.manufacturer or "",
                    item.unit,
                    float(item.unit_price) if item.unit_price is not None else 0.0,
                    int(item.stock_quantity) if item.stock_quantity is not None else 0,
                    int(item.reorder_level) if item.reorder_level is not None else 0,
                    item.description or ""
                ]
                ws.append(row)

            stream = BytesIO()
            wb.save(stream)
            stream.seek(0)
            return stream, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        elif format_type == "pdf":
            from jinja2 import Environment, FileSystemLoader
            from app.utils.pdf_generator import html_to_pdf
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from xhtml2pdf import default
            import os
            
            font_path = os.path.abspath("app/static/fonts/DejaVuSans.ttf")
            if "DejaVuSans" not in pdfmetrics.getRegisteredFontNames() and os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont("DejaVuSans", font_path))
                
            default.DEFAULT_FONT["dejavusans"] = "DejaVuSans"
            default.DEFAULT_FONT["dejavusans-bold"] = "DejaVuSans"
            default.DEFAULT_FONT["dejavusans-oblique"] = "DejaVuSans"
            default.DEFAULT_FONT["dejavusans-boldoblique"] = "DejaVuSans"
            
            env = Environment(loader=FileSystemLoader("app/templates"))
            template = env.get_template("medicines_export_template.html")

            # Format datetime dates for the Jinja template rendering
            formatted_medicines = []
            for item in medicines:
                expiry_str = "-"
                if item.expiry_date:
                    if isinstance(item.expiry_date, (date, datetime)):
                        expiry_str = item.expiry_date.strftime("%Y-%m-%d")
                    else:
                        expiry_str = str(item.expiry_date)

                formatted_medicines.append({
                    "id": item.id,
                    "sku": item.sku,
                    "name": item.name,
                    "generic_name": item.generic_name,
                    "category": item.category,
                    "barcode": item.barcode,
                    "batch_number": item.batch_number,
                    "expiry_date": expiry_str,
                    "unit": item.unit,
                    "unit_price": float(item.unit_price) if item.unit_price is not None else 0.0,
                    "stock_quantity": int(item.stock_quantity) if item.stock_quantity is not None else 0,
                    "reorder_level": int(item.reorder_level) if item.reorder_level is not None else 0,
                    "manufacturer": item.manufacturer or "-",
                    "description": item.description or "-",
                    "is_active": "Yes" if item.is_active else "No",
                    "created_at": item.created_at.strftime("%Y-%m-%d %H:%M:%S") if item.created_at else "-",
                    "updated_at": item.updated_at.strftime("%Y-%m-%d %H:%M:%S") if item.updated_at else "-",
                })


            html = template.render(
                medicines=formatted_medicines,
                generated_at=utc_now().strftime("%Y-%m-%d %H:%M:%S")
            )

            pdf_bytes = html_to_pdf(html)
            return pdf_bytes, "application/pdf"

        else:
            raise BadRequestException("Invalid format specified for export")

    async def generate_supplier_bulk_template(self):
        from io import BytesIO
        import openpyxl
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Suppliers Bulk Import"
        
        headers = [
            "name", "contact_person", "phone", "email", "address", "gst_number"
        ]
        ws.append(headers)
        
        # One valid sample row
        ws.append([
            "Alpha Pharma Distributors",
            "Jane Doe",
            "9876543211",
            "janedoe@alphapharma.com",
            "456 Medical Park, Sector 4",
            "27AAAAA1111A1Z1"
        ])
        
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream

    async def import_suppliers_from_excel(self, file, user_id: int) -> dict:
        from io import BytesIO
        from pydantic import ValidationError
        import openpyxl
        
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents))
        ws = wb.active
        
        header_row = next(ws.iter_rows(max_row=1, values_only=True), None)
        if not header_row:
            raise BadRequestException("The uploaded file is empty or has no headers.")
            
        headers = [str(h).strip().lower() for h in header_row if h is not None]
        required_headers = {"name"}
        if not required_headers.issubset(set(headers)):
            raise BadRequestException("Missing required headers in the upload template.")
            
        total_rows = 0
        created = 0
        failed = 0
        errors = []
        
        seen_phones = set()
        seen_emails = set()
        seen_gsts = set()
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None for cell in row):
                continue
                
            total_rows += 1
            row_dict = {}
            for header, val in zip(headers, row):
                if val is None or str(val).strip() == "":
                    row_dict[header] = None
                else:
                    row_dict[header] = str(val).strip()
                    
            try:
                name_raw = row_dict.get("name")
                if not name_raw:
                    raise BadRequestException("name is required.")
                name = str(name_raw).strip()
                
                phone_raw = row_dict.get("phone")
                email_raw = row_dict.get("email")
                gst_raw = row_dict.get("gst_number")
                
                # Check duplicate in file
                if phone_raw:
                    norm_phone = phone_raw.strip()
                    if norm_phone.startswith("+91"):
                        raw_ph = norm_phone[3:]
                    elif norm_phone.startswith("91") and len(norm_phone) == 12:
                        raw_ph = norm_phone[2:]
                    else:
                        raw_ph = norm_phone
                    if raw_ph in seen_phones:
                        raise BadRequestException(f"Duplicate phone number '{phone_raw}' found in upload file.")
                    seen_phones.add(raw_ph)
                    
                if email_raw:
                    norm_email = email_raw.strip().lower()
                    if norm_email in seen_emails:
                        raise BadRequestException(f"Duplicate email '{email_raw}' found in upload file.")
                    seen_emails.add(norm_email)
                    
                if gst_raw:
                    norm_gst = gst_raw.strip().upper()
                    if norm_gst in seen_gsts:
                        raise BadRequestException(f"Duplicate GST number '{gst_raw}' found in upload file.")
                    seen_gsts.add(norm_gst)
                
                # Build SupplierCreate model to run validations
                supplier_create = SupplierCreate(
                    name=name,
                    contact_person=row_dict.get("contact_person"),
                    phone=phone_raw,
                    email=email_raw,
                    address=row_dict.get("address"),
                    gst_number=gst_raw
                )
                
                await self.create_supplier(supplier_create, user_id)
                created += 1
                
            except ValidationError as e:
                failed += 1
                err_msg = "; ".join([f"{'.'.join(str(loc) for loc in error['loc'])}: {error['msg']}" for error in e.errors()])
                errors.append({
                    "row": row_idx,
                    "error": err_msg
                })
            except ConflictException as e:
                failed += 1
                errors.append({
                    "row": row_idx,
                    "error": str(e.detail)
                })
            except NotFoundException as e:
                failed += 1
                errors.append({
                    "row": row_idx,
                    "error": str(e.detail)
                })
            except BadRequestException as e:
                failed += 1
                errors.append({
                    "row": row_idx,
                    "error": str(e.detail)
                })
            except Exception as e:
                failed += 1
                errors.append({
                    "row": row_idx,
                    "error": str(e)
                })
                
        await self.db.flush()
        return {
            "total_rows": total_rows,
            "created": created,
            "failed": failed,
            "errors": errors
        }

    async def export_suppliers(self, format_type: str):
        from io import BytesIO
        from datetime import datetime, date
        
        suppliers = await self.supplier_repo.get_all_active()
        
        if format_type == "excel":
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Suppliers Export"
            
            headers = [
                "Sr. No.", "name", "contact_person", "phone", "email",
                "address", "gst_number", "is_active", "created_at"
            ]
            ws.append(headers)
            
            for sr_no, s in enumerate(suppliers, start=1):
                row = [
                    sr_no,
                    s.name,
                    s.contact_person or "",
                    s.phone or "",
                    s.email or "",
                    s.address or "",
                    s.gst_number or "",
                    "Active" if s.is_active else "Inactive",
                    s.created_at.strftime("%Y-%m-%d %H:%M:%S") if isinstance(s.created_at, datetime) else str(s.created_at)
                ]
                ws.append(row)
                
            stream = BytesIO()
            wb.save(stream)
            stream.seek(0)
            return stream, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            
        elif format_type == "pdf":
            from jinja2 import Environment, FileSystemLoader
            from app.utils.pdf_generator import html_to_pdf
            from app.utils.helpers import utc_now
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from xhtml2pdf import default
            import os
            
            font_path = os.path.abspath("app/static/fonts/DejaVuSans.ttf")
            if "DejaVuSans" not in pdfmetrics.getRegisteredFontNames() and os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont("DejaVuSans", font_path))
                
            default.DEFAULT_FONT["dejavusans"] = "DejaVuSans"
            default.DEFAULT_FONT["dejavusans-bold"] = "DejaVuSans"
            default.DEFAULT_FONT["dejavusans-oblique"] = "DejaVuSans"
            default.DEFAULT_FONT["dejavusans-boldoblique"] = "DejaVuSans"
            
            env = Environment(loader=FileSystemLoader("app/templates"))
            template = env.get_template("suppliers_export_template.html")
            
            formatted_suppliers = []
            for s in suppliers:
                formatted_suppliers.append({
                    "id": s.id,
                    "name": s.name,
                    "contact_person": s.contact_person,
                    "phone": s.phone,
                    "email": s.email,
                    "address": s.address,
                    "gst_number": s.gst_number,
                    "is_active": s.is_active,
                    "created_at": s.created_at.strftime("%Y-%m-%d %H:%M:%S") if isinstance(s.created_at, datetime) else str(s.created_at)
                })
                
            html_content = template.render(
                suppliers=formatted_suppliers,
                generated_at=utc_now().strftime("%Y-%m-%d %H:%M:%S")
            )
            
            pdf_data = html_to_pdf(html_content)
            return pdf_data, "application/pdf"
        else:
            raise BadRequestException("Invalid format specified for export")
