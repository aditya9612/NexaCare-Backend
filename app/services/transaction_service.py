from datetime import date, datetime
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.constants import BillingStatus
from app.core.exceptions import BadRequestException, NotFoundException
from app.models.billing_model import Payment
from app.repositories.audit_repository import AuditRepository
from app.repositories.billing_repository import BillingRepository
from app.repositories.transaction_repository import TransactionRepository
from app.schemas.transaction_schema import TransactionCreate, TransactionUpdate, TransactionResponse
from app.services.billing_service import BillingService
from app.utils.helpers import utc_now
from app.utils.pagination import build_paginated_result


class TransactionService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.repo = TransactionRepository(db)
        self.billing_repo = BillingRepository(db)
        self.audit_repo = AuditRepository(db)

    async def create_transaction(self, data: TransactionCreate, user_id: int) -> TransactionResponse:
        billing = await self.billing_repo.get_by_id(data.billing_id)
        if not billing:
            raise NotFoundException("Billing record not found")

        if billing.status == BillingStatus.CANCELLED:
            raise BadRequestException("Cannot add transaction to a cancelled bill")

        is_refund = data.is_refund or False
        is_completed = data.status is None or data.status.lower() == "completed"

        if is_completed:
            total_amt = billing.total_amount or 0.0
            paid_amt = billing.paid_amount or 0.0
            if is_refund:
                if data.amount > paid_amt:
                    raise BadRequestException("Refund amount exceeds paid amount")
                billing.paid_amount = round(paid_amt - data.amount, 2)
            else:
                balance_due = round(total_amt - paid_amt, 2)
                if data.amount > balance_due:
                    raise BadRequestException("Payment amount exceeds balance due")
                billing.paid_amount = round(paid_amt + data.amount, 2)

        payment = Payment(
            billing_id=data.billing_id,
            amount=data.amount,
            payment_method=data.payment_method,
            transaction_ref=data.transaction_ref,
            payment_date=data.payment_date or utc_now(),
            status=data.status or "completed",
            is_refund=is_refund,
            refund_reason=data.refund_reason,
            received_by=user_id,
        )

        payment = await self.repo.create(payment)

        # Create corresponding transaction history event
        from app.services.transaction_history_service import TransactionHistoryService
        event_type = "REFUND_ISSUED" if is_refund else "PAYMENT_RECEIVED"
        ref_prefix = "REF" if is_refund else "PAY"
        desc_action = "Refund Issued" if is_refund else "Payment Received"
        
        await TransactionHistoryService(self.db).create_event(
            event_type=event_type,
            reference_no=payment.transaction_ref or f"{ref_prefix}-{payment.id}",
            description=f"{desc_action} on bill {billing.bill_number or ''} via {payment.payment_method}",
            amount=payment.amount,
            source_module="refunds" if is_refund else "payments",
            source_id=payment.id,
            status=payment.status,
            user_id=user_id
        )

        if is_completed:
            await BillingService(self.db)._recalculate_billing(billing)

        await self.audit_repo.create("create", "transaction", user_id=user_id, resource_id=str(payment.id))
        return TransactionResponse.model_validate(payment)

    async def get_transaction(self, transaction_id: int) -> TransactionResponse:
        payment = await self.repo.get_by_id(transaction_id)
        if not payment:
            raise NotFoundException("Transaction not found")
        return TransactionResponse.model_validate(payment)

    async def update_transaction(self, transaction_id: int, data: TransactionUpdate, user_id: int) -> TransactionResponse:
        payment = await self.repo.get_by_id(transaction_id)
        if not payment:
            raise NotFoundException("Transaction not found")

        billing = await self.billing_repo.get_by_id(payment.billing_id)
        if not billing:
            raise NotFoundException("Billing record not found")

        old_completed = payment.status.lower() == "completed"
        old_is_refund = payment.is_refund or False
        old_amount = payment.amount

        # Update the payment properties
        for key, value in data.model_dump(exclude_unset=True).items():
            setattr(payment, key, value)

        new_completed = payment.status.lower() == "completed"
        new_is_refund = payment.is_refund or False
        new_amount = payment.amount

        # Recalculate billing balance & status if completion or amount changed
        if old_completed or new_completed:
            temp_paid = billing.paid_amount
            if old_completed:
                if old_is_refund:
                    temp_paid = round(temp_paid + old_amount, 2)
                else:
                    temp_paid = round(temp_paid - old_amount, 2)

            if new_completed:
                if new_is_refund:
                    temp_paid = round(temp_paid - new_amount, 2)
                else:
                    temp_paid = round(temp_paid + new_amount, 2)

            if temp_paid < 0:
                raise BadRequestException("Invalid transaction update: paid amount cannot be negative")
            if temp_paid > round(billing.total_amount, 2):
                raise BadRequestException("Invalid transaction update: paid amount exceeds bill total amount")

            billing.paid_amount = temp_paid
            await BillingService(self.db)._recalculate_billing(billing)

        payment = await self.repo.update(payment)

        # Update corresponding transaction history entry
        from app.models.transaction_history_model import TransactionHistory
        from sqlalchemy import select
        
        source_module = "refunds" if payment.is_refund else "payments"
        hist_res = await self.db.execute(
            select(TransactionHistory).where(
                TransactionHistory.source_module == source_module,
                TransactionHistory.source_id == payment.id,
                TransactionHistory.is_deleted == False
            )
        )
        hist = hist_res.scalar_one_or_none()
        if hist:
            hist.amount = payment.amount
            hist.status = payment.status
            hist.event_type = "REFUND_ISSUED" if payment.is_refund else "PAYMENT_RECEIVED"
            if payment.transaction_ref:
                hist.reference_no = payment.transaction_ref
        else:
            from app.services.transaction_history_service import TransactionHistoryService
            event_type = "REFUND_ISSUED" if payment.is_refund else "PAYMENT_RECEIVED"
            ref_prefix = "REF" if payment.is_refund else "PAY"
            desc_action = "Refund Issued" if payment.is_refund else "Payment Received"
            await TransactionHistoryService(self.db).create_event(
                event_type=event_type,
                reference_no=payment.transaction_ref or f"{ref_prefix}-{payment.id}",
                description=f"{desc_action} on bill {billing.bill_number or ''} via {payment.payment_method}",
                amount=payment.amount,
                source_module=source_module,
                source_id=payment.id,
                status=payment.status,
                user_id=user_id
            )

        await self.audit_repo.create("update", "transaction", user_id=user_id, resource_id=str(payment.id))
        return TransactionResponse.model_validate(payment)

    async def delete_transaction(self, transaction_id: int, user_id: int) -> None:
        payment = await self.repo.get_by_id(transaction_id)
        if not payment:
            raise NotFoundException("Transaction not found")

        billing = await self.billing_repo.get_by_id(payment.billing_id)
        if not billing:
            raise NotFoundException("Billing record not found")

        if payment.status.lower() == "completed":
            if payment.is_refund:
                billing.paid_amount = round(billing.paid_amount + payment.amount, 2)
            else:
                billing.paid_amount = round(billing.paid_amount - payment.amount, 2)

            if billing.paid_amount < 0:
                billing.paid_amount = 0.0

            await BillingService(self.db)._recalculate_billing(billing)

        await self.repo.delete(payment)

        # Soft delete corresponding transaction history entry
        from app.models.transaction_history_model import TransactionHistory
        from sqlalchemy import update
        
        source_module = "refunds" if payment.is_refund else "payments"
        stmt = update(TransactionHistory).where(
            TransactionHistory.source_module == source_module,
            TransactionHistory.source_id == payment.id,
            TransactionHistory.is_deleted == False
        ).values(is_deleted=True, deleted_at=utc_now())
        await self.db.execute(stmt)

        await self.audit_repo.create("delete", "transaction", user_id=user_id, resource_id=str(transaction_id))

    async def list_transactions(
        self,
        page: int = 1,
        size: int = 20,
        sort_by: str = "created_at",
        sort_order: str = "desc",
        billing_id: int | None = None,
        payment_method: str | None = None,
        status: str | None = None,
        start_date: date | datetime | None = None,
        end_date: date | datetime | None = None,
        q: str | None = None,
    ):
        skip = (page - 1) * size
        items = await self.repo.list_all(
            skip=skip,
            limit=size,
            sort_by=sort_by,
            sort_order=sort_order,
            billing_id=billing_id,
            payment_method=payment_method,
            status=status,
            start_date=start_date,
            end_date=end_date,
            q=q,
        )
        total = await self.repo.count_all(
            billing_id=billing_id,
            payment_method=payment_method,
            status=status,
            start_date=start_date,
            end_date=end_date,
            q=q,
        )
        responses = [TransactionResponse.model_validate(item) for item in items]
        return build_paginated_result(responses, total, page, size)

    async def generate_transactions_bulk_template(self):
        from io import BytesIO
        import openpyxl
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions Bulk Import"
        
        headers = [
            "billing_id", "amount", "payment_method", "transaction_ref", "payment_date", "status", "is_refund", "refund_reason"
        ]
        ws.append(headers)
        
        # One valid sample row
        ws.append([
            1,
            1500.00,
            "upi",
            "TXN12345678",
            "2026-08-18 11:30:00",
            "completed",
            "false",
            ""
        ])
        
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream

    async def import_transactions_from_excel(self, file, user_id: int) -> dict:
        from io import BytesIO
        from pydantic import ValidationError
        import openpyxl
        
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents))
        ws = wb.active
        
        header_row = next(ws.iter_rows(max_row=1, values_only=True), None)
        if not header_row:
            raise BadRequestException("The uploaded file is empty or has no headers.")
            
        headers = [str(h).strip().lower() for h in header_row if h is not None]
        required_headers = {"billing_id", "amount", "payment_method"}
        if not required_headers.issubset(set(headers)):
            raise BadRequestException("Missing required headers in the upload template.")
            
        total_rows = 0
        created = 0
        failed = 0
        errors = []
        
        seen_refs = set()
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None for cell in row):
                continue
                
            total_rows += 1
            row_dict = {}
            for header, val in zip(headers, row):
                if val is None or str(val).strip() == "":
                    row_dict[header] = None
                else:
                    row_dict[header] = val
                    
            try:
                # 1. billing_id parsing & validation
                billing_id_raw = row_dict.get("billing_id")
                if billing_id_raw is None:
                    raise BadRequestException("billing_id is required.")
                try:
                    f_val = float(billing_id_raw)
                    if not f_val.is_integer():
                        raise ValueError()
                    billing_id = int(f_val)
                except (ValueError, TypeError):
                    raise BadRequestException(f"Invalid integer value for billing_id: {billing_id_raw}")

                # 2. amount parsing
                amount_raw = row_dict.get("amount")
                if amount_raw is None:
                    raise BadRequestException("amount is required.")
                try:
                    amount = round(float(amount_raw), 2)
                except (ValueError, TypeError):
                    raise BadRequestException(f"Invalid numeric value for amount: {amount_raw}")

                # 3. payment_method parsing
                pm_raw = row_dict.get("payment_method")
                if pm_raw is None:
                    raise BadRequestException("payment_method is required.")
                payment_method = str(pm_raw).strip()

                # 4. transaction_ref parsing & sheet duplicate check
                ref_val = row_dict.get("transaction_ref")
                transaction_ref = None
                if ref_val is not None and str(ref_val).strip() != "":
                    transaction_ref = str(ref_val).strip()
                    if transaction_ref in seen_refs:
                        raise BadRequestException(f"Duplicate transaction_ref found in upload file: {transaction_ref}")
                    seen_refs.add(transaction_ref)

                # 5. payment_date parsing
                payment_date_val = row_dict.get("payment_date")
                payment_date = None
                if payment_date_val is not None:
                    if isinstance(payment_date_val, datetime):
                        payment_date = payment_date_val
                    elif isinstance(payment_date_val, date) and not isinstance(payment_date_val, datetime):
                        payment_date = datetime.combine(payment_date_val, datetime.min.time())
                    elif isinstance(payment_date_val, (int, float)) and not isinstance(payment_date_val, bool):
                        try:
                            from openpyxl.utils.datetime import from_excel
                            payment_date = from_excel(payment_date_val)
                        except Exception:
                            raise BadRequestException(f"Invalid numeric date value for payment_date: {payment_date_val}")
                    else:
                        try:
                            p_str = str(payment_date_val).strip()
                            # Check if string represents an Excel serial date number
                            try:
                                float_val = float(p_str)
                                from openpyxl.utils.datetime import from_excel
                                payment_date = from_excel(float_val)
                            except ValueError:
                                # Try standard string date formats
                                for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d"):
                                    try:
                                        payment_date = datetime.strptime(p_str, fmt)
                                        break
                                    except ValueError:
                                        continue
                                if payment_date is None:
                                    payment_date = datetime.fromisoformat(p_str)
                        except Exception:
                            raise BadRequestException(f"Invalid datetime format for payment_date: {payment_date_val}")

                # 6. status
                status = "completed"
                status_raw = row_dict.get("status")
                if status_raw is not None:
                    status = str(status_raw).strip()

                # 7. is_refund
                is_refund_val = row_dict.get("is_refund")
                is_refund = False
                if is_refund_val is not None:
                    if isinstance(is_refund_val, bool):
                        is_refund = is_refund_val
                    else:
                        is_refund_str = str(is_refund_val).strip().lower()
                        if is_refund_str in ("true", "1", "yes"):
                            is_refund = True
                        elif is_refund_str in ("false", "0", "no"):
                            is_refund = False
                        else:
                            raise BadRequestException(f"Invalid boolean value for is_refund: {is_refund_val}")

                # 8. refund_reason
                refund_reason = None
                rr_raw = row_dict.get("refund_reason")
                if rr_raw is not None:
                    refund_reason = str(rr_raw).strip()

                # Validate using TransactionCreate schema
                txn_create = TransactionCreate(
                    billing_id=billing_id,
                    amount=amount,
                    payment_method=payment_method,
                    transaction_ref=transaction_ref,
                    payment_date=payment_date,
                    status=status,
                    is_refund=is_refund,
                    refund_reason=refund_reason
                )

                # Process row sequentially using existing financial validations
                await self.create_transaction(txn_create, user_id)
                created += 1

            except ValidationError as e:
                failed += 1
                err_msg = "; ".join([f"{'.'.join(str(loc) for loc in error['loc'])}: {error['msg']}" for error in e.errors()])
                errors.append({
                    "row": row_idx,
                    "error": err_msg
                })
            except BadRequestException as e:
                failed += 1
                errors.append({
                    "row": row_idx,
                    "error": str(e.detail)
                })
            except NotFoundException as e:
                failed += 1
                errors.append({
                    "row": row_idx,
                    "error": str(e.detail)
                })
            except Exception as e:
                failed += 1
                errors.append({
                    "row": row_idx,
                    "error": str(e)
                })
                
        await self.db.flush()
        return {
            "total_rows": total_rows,
            "created": created,
            "failed": failed,
            "errors": errors
        }

    async def export_transactions(
        self,
        format_type: str,
        billing_id: int | None = None,
        payment_method: str | None = None,
        status: str | None = None,
        start_date: date | None = None,
        end_date: date | None = None,
        q: str | None = None
    ):
        from io import BytesIO
        from datetime import datetime, date
        from sqlalchemy import select
        
        # Retrieve all active transactions matching filters
        payments = await self.repo.get_all_active(
            billing_id=billing_id,
            payment_method=payment_method,
            status=status,
            start_date=start_date,
            end_date=end_date,
            q=q
        )
        
        # Load associated billing records for bill_number/patient lookup
        from app.models.billing_model import Billing
        from app.models.patient_model import Patient
        
        b_stmt = select(Billing.id, Billing.bill_number, Billing.patient_id)
        b_res = await self.db.execute(b_stmt)
        billing_map = {row[0]: (row[1], row[2]) for row in b_res.all()}
        
        # Fetch patients names
        p_stmt = select(Patient.id, Patient.first_name, Patient.last_name).where(Patient.is_deleted == False)
        p_res = await self.db.execute(p_stmt)
        patients_map = {row[0]: f"{row[1]} {row[2]}" for row in p_res.all()}

        if format_type == "excel":
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Transactions Report"
            
            headers = [
                "Sr. No.", "billing_id", "bill_number", "patient_name", "amount", "payment_method",
                "transaction_ref", "payment_date", "status", "is_refund", "refund_reason",
                "created_at", "updated_at"
            ]
            ws.append(headers)
            
            for sr_no, p in enumerate(payments, start=1):
                bill_number, p_id = billing_map.get(p.billing_id, ("", None))
                pat_name = patients_map.get(p_id, "") if p_id else ""
                
                row = [
                    sr_no,
                    p.billing_id,
                    bill_number,
                    pat_name,
                    f"₹{p.amount:.2f}",
                    p.payment_method,
                    p.transaction_ref or "",
                    p.payment_date.strftime("%Y-%m-%d %H:%M:%S") if isinstance(p.payment_date, datetime) else str(p.payment_date),
                    p.status,
                    "Yes" if p.is_refund else "No",
                    p.refund_reason or "",
                    p.created_at.strftime("%Y-%m-%d %H:%M:%S") if isinstance(p.created_at, datetime) else str(p.created_at),
                    p.updated_at.strftime("%Y-%m-%d %H:%M:%S") if isinstance(p.updated_at, datetime) else str(p.updated_at)
                ]
                ws.append(row)
                
            stream = BytesIO()
            wb.save(stream)
            stream.seek(0)
            return stream, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            
        elif format_type == "pdf":
            from jinja2 import Environment, FileSystemLoader
            from app.utils.pdf_generator import html_to_pdf
            from app.utils.helpers import utc_now
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from xhtml2pdf import default
            import os
            
            font_path = os.path.abspath("app/static/fonts/DejaVuSans.ttf")
            if "DejaVuSans" not in pdfmetrics.getRegisteredFontNames() and os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont("DejaVuSans", font_path))
                
            default.DEFAULT_FONT["dejavusans"] = "DejaVuSans"
            default.DEFAULT_FONT["dejavusans-bold"] = "DejaVuSans"
            default.DEFAULT_FONT["dejavusans-oblique"] = "DejaVuSans"
            default.DEFAULT_FONT["dejavusans-boldoblique"] = "DejaVuSans"
            
            env = Environment(loader=FileSystemLoader("app/templates"))
            template = env.get_template("transactions_export_template.html")
            
            formatted_txns = []
            for p in payments:
                bill_number, p_id = billing_map.get(p.billing_id, ("", None))
                pat_name = patients_map.get(p_id, "") if p_id else ""
                
                formatted_txns.append({
                    "id": p.id,
                    "billing_id": p.billing_id,
                    "bill_number": bill_number,
                    "patient_name": pat_name,
                    "amount": p.amount,
                    "payment_method": p.payment_method,
                    "transaction_ref": p.transaction_ref,
                    "payment_date": p.payment_date.strftime("%Y-%m-%d %H:%M:%S") if isinstance(p.payment_date, datetime) else str(p.payment_date),
                    "status": p.status,
                    "is_refund": p.is_refund,
                    "refund_reason": p.refund_reason,
                    "created_at": p.created_at.strftime("%Y-%m-%d %H:%M:%S") if isinstance(p.created_at, datetime) else str(p.created_at),
                    "updated_at": p.updated_at.strftime("%Y-%m-%d %H:%M:%S") if isinstance(p.updated_at, datetime) else str(p.updated_at)
                })
                
            html_content = template.render(
                transactions=formatted_txns,
                generated_at=utc_now().strftime("%Y-%m-%d %H:%M:%S")
            )
            
            pdf_data = html_to_pdf(html_content)
            return pdf_data, "application/pdf"
        else:
            raise BadRequestException("Invalid format specified for export")
