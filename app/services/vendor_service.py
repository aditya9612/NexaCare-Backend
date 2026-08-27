from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import BadRequestException, NotFoundException, ConflictException
from app.models.vendor_model import Vendor
from app.models.expense_model import Expense
from app.models.inventory_model import InventoryItem
from app.repositories.vendor_repository import VendorRepository
from app.repositories.audit_repository import AuditRepository
from app.schemas.vendor_schema import VendorCreate, VendorUpdate, VendorResponse
from app.utils.pagination import build_paginated_result


class VendorService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.vendor_repo = VendorRepository(db)
        self.audit_repo = AuditRepository(db)

    async def create_vendor(self, data: VendorCreate, user_id: int) -> VendorResponse:
        existing = await self.vendor_repo.get_by_name(data.name)
        if existing:
            raise ConflictException(f"Vendor with name '{data.name}' already exists")

        vendor = Vendor(**data.model_dump())
        vendor = await self.vendor_repo.create(vendor)
        await self.audit_repo.create("create", "vendor", user_id=user_id, resource_id=str(vendor.id))
        return VendorResponse.model_validate(vendor)

    async def list_vendors(self, page: int = 1, size: int = 20, vendor_type: str | None = None):
        skip = (page - 1) * size
        vendors = await self.vendor_repo.list_all(skip=skip, limit=size, vendor_type=vendor_type)
        total = await self.vendor_repo.count_all(vendor_type=vendor_type)
        return build_paginated_result(
            [VendorResponse.model_validate(v) for v in vendors], total, page, size
        )

    async def get_vendor(self, vendor_id: int) -> VendorResponse:
        vendor = await self.vendor_repo.get_by_id(vendor_id)
        if not vendor:
            raise NotFoundException(f"Vendor with ID {vendor_id} not found")
        return VendorResponse.model_validate(vendor)

    async def update_vendor(self, vendor_id: int, data: VendorUpdate, user_id: int) -> VendorResponse:
        vendor = await self.vendor_repo.get_by_id(vendor_id)
        if not vendor:
            raise NotFoundException(f"Vendor with ID {vendor_id} not found")

        if data.name:
            existing = await self.vendor_repo.get_by_name(data.name)
            if existing and existing.id != vendor_id:
                raise ConflictException(f"Vendor with name '{data.name}' already exists")

        for key, value in data.model_dump(exclude_unset=True).items():
            setattr(vendor, key, value)

        vendor = await self.vendor_repo.update(vendor)
        await self.audit_repo.create("update", "vendor", user_id=user_id, resource_id=str(vendor.id))
        return VendorResponse.model_validate(vendor)

    async def delete_vendor(self, vendor_id: int, user_id: int) -> None:
        vendor = await self.vendor_repo.get_by_id(vendor_id)
        if not vendor:
            raise NotFoundException(f"Vendor with ID {vendor_id} not found")

        # Check if linked to any expenses
        expense_exists = await self.db.scalar(
            select(func.count()).select_from(Expense).where(
                Expense.vendor_id == vendor_id,
                Expense.is_deleted.is_(False)
            )
        )
        if expense_exists and expense_exists > 0:
            raise BadRequestException("Cannot delete vendor as it is linked to one or more expenses")

        # Check if linked to any inventory items
        item_exists = await self.db.scalar(
            select(func.count()).select_from(InventoryItem).where(
                InventoryItem.vendor_id == vendor_id,
                InventoryItem.is_deleted.is_(False)
            )
        )
        if item_exists and item_exists > 0:
            raise BadRequestException("Cannot delete vendor as it is linked to one or more inventory items")

        await self.vendor_repo.soft_delete(vendor)
        await self.audit_repo.create("delete", "vendor", user_id=user_id, resource_id=str(vendor.id))

    async def generate_vendor_bulk_template(self):
        from io import BytesIO
        import openpyxl
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vendors Bulk Import"
        
        headers = [
            "name", "vendor_type", "contact_person", "phone", "email", "address", "gst_number", "service_type"
        ]
        ws.append(headers)
        
        # Configure Header Row Height
        ws.row_dimensions[1].height = 32
        
        # Header Style definition
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        header_font = Font(name="Calibri", size=12, bold=True, color="000000")
        header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_side = Side(style="thin", color="D3D3D3")
        header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = header_border
        
        # One valid sample row
        ws.append([
            "Acme Healthcare Solutions",
            "inventory",
            "John Doe",
            "9876543210",
            "johndoe@acme.com",
            "123 Health Ave, Suite 100",
            "27AAAAA1111A1Z1",
            "Medical Equipment"
        ])
        
        # Freeze panes at A2
        ws.freeze_panes = "A2"
        
        # AutoFilter (using openpyxl get_column_letter dynamically based on actual header length)
        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}2"
        
        # Set custom column widths
        col_widths = {
            1: 25,  # name
            2: 15,  # vendor_type
            3: 20,  # contact_person
            4: 15,  # phone
            5: 25,  # email
            6: 30,  # address
            7: 20,  # gst_number
            8: 20   # service_type
        }
        for col_idx, width in col_widths.items():
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width
        
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream

    async def import_vendors_from_excel(self, file, user_id: int) -> dict:
        from io import BytesIO
        from pydantic import ValidationError
        import openpyxl
        
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents))
        ws = wb.active
        
        header_row = next(ws.iter_rows(max_row=1, values_only=True), None)
        if not header_row:
            raise BadRequestException("The uploaded file is empty or has no headers.")
            
        headers = [str(h).strip().lower() for h in header_row if h is not None]
        required_headers = {"name", "vendor_type"}
        if not required_headers.issubset(set(headers)):
            raise BadRequestException("Missing required headers in the upload template.")
            
        total_rows = 0
        created = 0
        failed = 0
        errors = []
        
        seen_names = set()
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None for cell in row):
                continue
                
            total_rows += 1
            row_dict = {}
            for header, val in zip(headers, row):
                if val is None or str(val).strip() == "":
                    row_dict[header] = None
                else:
                    row_dict[header] = str(val).strip()
                    
            try:
                name_raw = row_dict.get("name")
                if not name_raw:
                    raise BadRequestException("name is required.")
                name = str(name_raw).strip()
                
                # Check duplicate in file
                normalized_name = name.lower()
                if normalized_name in seen_names:
                    raise BadRequestException(f"Duplicate vendor name '{name}' found in upload file.")
                seen_names.add(normalized_name)
                
                vendor_type_raw = row_dict.get("vendor_type")
                if not vendor_type_raw:
                    raise BadRequestException("vendor_type is required.")
                vendor_type = str(vendor_type_raw).strip().lower()
                if vendor_type == "expense":
                    vendor_type = "expenses"
                
                # Build VendorCreate model to execute all validators
                vendor_create = VendorCreate(
                    name=name,
                    vendor_type=vendor_type,
                    contact_person=row_dict.get("contact_person"),
                    phone=row_dict.get("phone"),
                    email=row_dict.get("email"),
                    address=row_dict.get("address"),
                    gst_number=row_dict.get("gst_number"),
                    service_type=row_dict.get("service_type")
                )
                
                await self.create_vendor(vendor_create, user_id)
                created += 1
                
            except ValidationError as e:
                failed += 1
                err_msg = "; ".join([f"{'.'.join(str(loc) for loc in error['loc'])}: {error['msg']}" for error in e.errors()])
                errors.append({
                    "row": row_idx,
                    "error": err_msg
                })
            except ConflictException as e:
                failed += 1
                errors.append({
                    "row": row_idx,
                    "error": str(e.detail)
                })
            except NotFoundException as e:
                failed += 1
                errors.append({
                    "row": row_idx,
                    "error": str(e.detail)
                })
            except BadRequestException as e:
                failed += 1
                errors.append({
                    "row": row_idx,
                    "error": str(e.detail)
                })
            except Exception as e:
                failed += 1
                errors.append({
                    "row": row_idx,
                    "error": str(e)
                })
                
        await self.db.flush()
        return {
            "total_rows": total_rows,
            "created": created,
            "failed": failed,
            "errors": errors
        }

    async def export_vendors(self, format_type: str):
        from io import BytesIO
        from datetime import datetime, date
        
        # Retrieve all non-deleted vendors
        vendors = await self.vendor_repo.get_all_active()
        
        if format_type == "excel":
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Vendors Export"
            
            headers = [
                "Sr. No.", "name", "vendor_type", "contact_person", "phone",
                "email", "address", "gst_number", "service_type", "is_active",
                "created_at", "updated_at"
            ]
            ws.append(headers)
            
            # Configure Header Row Height
            ws.row_dimensions[1].height = 32
            
            # Header Style definition
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            header_font = Font(name="Calibri", size=12, bold=True, color="000000")
            header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_side = Side(style="thin", color="D3D3D3")
            header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = header_border
            
            for sr_no, v in enumerate(vendors, start=1):
                row = [
                    sr_no,
                    v.name,
                    v.vendor_type,
                    v.contact_person or "",
                    v.phone or "",
                    v.email or "",
                    v.address or "",
                    v.gst_number or "",
                    v.service_type or "",
                    "Active" if v.is_active else "Inactive",
                    v.created_at.strftime("%Y-%m-%d %H:%M:%S") if isinstance(v.created_at, datetime) else str(v.created_at),
                    v.updated_at.strftime("%Y-%m-%d %H:%M:%S") if isinstance(v.updated_at, datetime) else str(v.updated_at)
                ]
                ws.append(row)
                
            # Freeze panes at A2
            ws.freeze_panes = "A2"
            
            # AutoFilter (using openpyxl get_column_letter dynamically based on actual header length)
            total_rows = len(vendors) + 1
            last_col_letter = openpyxl.utils.get_column_letter(len(headers))
            ws.auto_filter.ref = f"A1:{last_col_letter}{total_rows}"
            
            # Set custom column widths
            col_widths = {
                1: 8,   # Sr. No.
                2: 25,  # name
                3: 15,  # vendor_type
                4: 20,  # contact_person
                5: 15,  # phone
                6: 25,  # email
                7: 30,  # address
                8: 20,  # gst_number
                9: 20,  # service_type
                10: 12, # is_active
                11: 22, # created_at
                12: 22  # updated_at
            }
            for col_idx, width in col_widths.items():
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = width

            stream = BytesIO()
            wb.save(stream)
            stream.seek(0)
            return stream, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            
        elif format_type == "pdf":
            from jinja2 import Environment, FileSystemLoader
            from app.utils.pdf_generator import html_to_pdf
            from app.utils.helpers import utc_now
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from xhtml2pdf import default
            import os
            
            font_path = os.path.abspath("app/static/fonts/DejaVuSans.ttf")
            if "DejaVuSans" not in pdfmetrics.getRegisteredFontNames() and os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont("DejaVuSans", font_path))
                
            default.DEFAULT_FONT["dejavusans"] = "DejaVuSans"
            default.DEFAULT_FONT["dejavusans-bold"] = "DejaVuSans"
            default.DEFAULT_FONT["dejavusans-oblique"] = "DejaVuSans"
            default.DEFAULT_FONT["dejavusans-boldoblique"] = "DejaVuSans"
            
            env = Environment(loader=FileSystemLoader("app/templates"))
            template = env.get_template("vendors_export_template.html")
            
            def wrap_unbroken_text(text: str, max_chars: int = 12) -> str:
                if not text:
                    return ""
                words = text.split(" ")
                processed_words = []
                for word in words:
                    if len(word) > max_chars:
                        chunks = [word[i:i+max_chars] for i in range(0, len(word), max_chars)]
                        processed_words.append("\u200b".join(chunks))
                    else:
                        processed_words.append(word)
                return " ".join(processed_words)

            def wrap_val(val, max_chars: int = 12) -> str:
                if val is None:
                    return ""
                return wrap_unbroken_text(str(val), max_chars)
            
            formatted_vendors = []
            for v in vendors:
                formatted_vendors.append({
                    "id": v.id,
                    "name": wrap_val(v.name, 12),
                    "vendor_type": v.vendor_type.capitalize() if v.vendor_type else "",
                    "contact_person": wrap_val(v.contact_person, 12),
                    "phone": wrap_val(v.phone, 12),
                    "email": wrap_val(v.email, 12),
                    "address": wrap_val(v.address, 12),
                    "gst_number": wrap_val(v.gst_number, 12),
                    "service_type": wrap_val(v.service_type, 12),
                    "is_active": v.is_active,
                    "created_at": v.created_at.strftime("%Y-%m-%d %H:%M:%S") if isinstance(v.created_at, datetime) else str(v.created_at),
                    "updated_at": v.updated_at.strftime("%Y-%m-%d %H:%M:%S") if isinstance(v.updated_at, datetime) else str(v.updated_at)
                })
                
            html_content = template.render(
                vendors=formatted_vendors,
                generated_at=utc_now().strftime("%Y-%m-%d %H:%M:%S")
            )
            
            pdf_data = html_to_pdf(html_content)
            return pdf_data, "application/pdf"
